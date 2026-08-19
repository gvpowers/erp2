"""
GV Powers ERP - Complete Monolithic Application
All models, routes, services, utils, config, and PDF generation in one file.
"""


############################################################
# IMPORTS
############################################################

import os
import io
import re
import sys
import time
import logging
import logging.handlers
import subprocess
import shutil
import csv
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders as email_encoders
from functools import wraps
from typing import Dict, List, Optional, Tuple, Union

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, session, jsonify, send_file, abort, current_app,
    after_this_request,
)
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect
from flask_login import LoginManager, UserMixin, login_user, logout_user, current_user
from sqlalchemy import func, desc, and_, or_, extract, text
from sqlalchemy.exc import IntegrityError, OperationalError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import barcode
from barcode.writer import SVGWriter

load_dotenv()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


############################################################
# PDF FONTS (Unicode TTF - supports the Indian Rupee symbol)
############################################################

_PDF_FONT = 'DejaVuSansCondensed'
_PDF_FONT_BOLD = 'DejaVuSansCondensed-Bold'
_PDF_FONT_OBLIQUE = 'DejaVuSansCondensed-Oblique'


# Fixed GV POWERS business identity. These values are NOT configurable from the
# admin Settings page — they are the canonical identity for this installation and
# are used across invoices, quotations, reports, PDFs and emails.
_FIXED_COMPANY = {
    'name': 'GV Powers',
    'gstin': '33AGEPV1534G2ZJ',
    'phone': '+91 98940 79090',
    'mobile': '+91 98940 79095',
    'website': 'https://gvpowers.in',
    'email': 'gvpowerssalem@gmail.com',
    'state': 'Tamil Nadu',
    'state_code': 33,
    'address': 'No. 10, Kadharkhan Street, Opp. Railway Junction, Salem - 636005, Tamil Nadu, India',
    'city': 'Salem',
    'pincode': '636005',
    'country': 'India',
    'tagline': 'Powering A Better Tomorrow',
    'services': 'Solar Energy | UPS Systems | Inverters | RO Solutions | Electricals',
}


def _register_pdf_fonts():
    """Register embedded Unicode fonts for reliable PDF text rendering.

    Uses the condensed DejaVu Sans family so glyph widths closely match the previous
    Helvetica layout while providing full Unicode support. Currency is displayed as
    "Rs." (ASCII) in generated PDFs so it can never render as a black square.
    """
    font_dir = os.path.join(BASE_DIR, 'fonts')
    pdfmetrics.registerFont(TTFont(_PDF_FONT, os.path.join(font_dir, 'DejaVuSansCondensed.ttf')))
    pdfmetrics.registerFont(TTFont(_PDF_FONT_BOLD, os.path.join(font_dir, 'DejaVuSansCondensed-Bold.ttf')))
    pdfmetrics.registerFont(TTFont(_PDF_FONT_OBLIQUE, os.path.join(font_dir, 'DejaVuSansCondensed-Oblique.ttf')))
    pdfmetrics.registerFontFamily(
        _PDF_FONT,
        normal=_PDF_FONT,
        bold=_PDF_FONT_BOLD,
        italic=_PDF_FONT_OBLIQUE,
        boldItalic=_PDF_FONT_BOLD,
    )


_register_pdf_fonts()


############################################################
# CONFIGURATION
############################################################


class Config:
    SECRET_KEY = os.getenv("SECRET_KEY", "change-me-in-production")
    SQLALCHEMY_DATABASE_URI = os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/gv_powers_erp")
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    SQLALCHEMY_ENGINE_OPTIONS = {"pool_pre_ping": True, "pool_recycle": 300, "pool_size": 10, "max_overflow": 20}
    PERMANENT_SESSION_LIFETIME = timedelta(hours=8)
    BASE_DIR = BASE_DIR
    UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
    BACKUP_FOLDER = os.path.join(BASE_DIR, "backups", "database")
    PDF_FOLDER = os.path.join(BASE_DIR, "pdf")
    EXPORT_FOLDER = os.path.join(BASE_DIR, "exports")
    LOG_FOLDER = os.path.join(BASE_DIR, "logs")
    MAX_CONTENT_LENGTH = int(os.getenv("MAX_CONTENT_SIZE", 16 * 1024 * 1024))
    COMPANY_NAME = os.getenv("COMPANY_NAME", "GV Powers")
    COMPANY_TAGLINE = os.getenv("COMPANY_TAGLINE", "Powering A Better Tomorrow")
    COMPANY_SERVICES = os.getenv("COMPANY_SERVICES", "Solar Energy | UPS Systems | Inverters | RO Solutions | Electricals")
    COMPANY_GSTIN = os.getenv("COMPANY_GSTIN", "33AGEPV1534G2ZJ")
    COMPANY_STATE = os.getenv("COMPANY_STATE", "Tamil Nadu")
    COMPANY_STATE_CODE = int(os.getenv("COMPANY_STATE_CODE", "33"))
    COMPANY_ADDRESS = os.getenv("COMPANY_ADDRESS", "No. 10, Kadharkhan Street, Opp. Railway Junction, Salem - 636005, Tamil Nadu")
    COMPANY_PHONE = os.getenv("COMPANY_PHONE", "+91 98940 79090")
    COMPANY_MOBILE = os.getenv("COMPANY_MOBILE", "+91 98940 79095")
    COMPANY_EMAIL = os.getenv("COMPANY_EMAIL", "gvpowerssalem@gmail.com")
    COMPANY_WEBSITE = os.getenv("COMPANY_WEBSITE", "https://gvpowers.in")
    ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
    ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "Admin@123")
    ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "admin@gvpowers.in")
    GST_RATES = [0, 5, 12, 18, 28]

    COMPANY = {}  # populated at runtime in create_app


class DevelopmentConfig(Config):
    DEBUG = True
    SQLALCHEMY_DATABASE_URI = os.getenv("DATABASE_URL", "sqlite:///" + os.path.join(BASE_DIR, "instance", "app.db").replace("\\", "/"))
    SQLALCHEMY_ENGINE_OPTIONS = {"pool_pre_ping": True}


class ProductionConfig(Config):
    DEBUG = False
    TESTING = False
    SQLALCHEMY_ENGINE_OPTIONS = {"pool_pre_ping": True, "pool_recycle": 300, "pool_size": 10, "max_overflow": 20}


config_by_name = {"development": DevelopmentConfig, "production": ProductionConfig, "default": DevelopmentConfig}


def get_config():
    env = os.getenv("FLASK_ENV", "development")
    return config_by_name.get(env, DevelopmentConfig)


############################################################
# ENVIRONMENT VARIABLES
############################################################

GST_STATE_CODES: Dict[int, str] = {
    1: "Jammu & Kashmir", 2: "Himachal Pradesh", 3: "Punjab", 4: "Chandigarh",
    5: "Uttarakhand", 6: "Haryana", 7: "Delhi", 8: "Rajasthan",
    9: "Uttar Pradesh", 10: "Bihar", 11: "Sikkim", 12: "Arunachal Pradesh",
    13: "Nagaland", 14: "Manipur", 15: "Mizoram", 16: "Tripura",
    17: "Meghalaya", 18: "Assam", 19: "West Bengal", 20: "Jharkhand",
    21: "Odisha", 22: "Chhattisgarh", 23: "Madhya Pradesh", 24: "Gujarat",
    25: "Daman & Diu", 26: "Dadra & Nagar Haveli and Daman & Diu",
    27: "Maharashtra", 28: "Andhra Pradesh (old code - Telangana)",
    29: "Karnataka", 30: "Goa", 31: "Lakshadweep", 32: "Kerala",
    33: "Tamil Nadu", 34: "Puducherry", 35: "Andaman & Nicobar Islands",
    36: "Telangana", 37: "Andhra Pradesh", 38: "Ladakh", 97: "Other Territory",
}
GST_STATE_NAMES_TO_CODES: Dict[str, int] = {name: code for code, name in GST_STATE_CODES.items()}
VALID_GST_RATES: Tuple[Decimal, ...] = (Decimal("0"), Decimal("0.25"), Decimal("3"), Decimal("5"), Decimal("12"), Decimal("18"), Decimal("28"))
VALID_GST_RATE_FLOATS: Tuple[float, ...] = (0.0, 0.25, 3.0, 5.0, 12.0, 18.0, 28.0)

_GSTIN_RE = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[A-Z0-9]{1}Z[A-Z0-9]{1}$", re.IGNORECASE)
_GSTIN_CHECKSUM_CHARS = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def validate_gstIN(gstin: str) -> bool:
    if not gstin:
        return False
    gstin = gstin.strip().upper()
    if len(gstin) != 15:
        return False
    if not _GSTIN_RE.match(gstin):
        return False
    state_code = int(gstin[:2])
    if state_code not in GST_STATE_CODES and state_code != 97:
        return False
    return _verify_gstIN_checksum(gstin)


def _verify_gstIN_checksum(gstin: str) -> bool:
    factor = 2
    total = 0
    for idx, char in enumerate(gstin[:14]):
        digit = _GSTIN_CHECKSUM_CHARS.index(char)
        if idx % 2 == 0:
            total += _lmod(digit * factor)
        else:
            total += digit
    remainder = total % 36
    check_code = (36 - remainder) % 36
    return _GSTIN_CHECKSUM_CHARS[check_code] == gstin[14].upper()


def _lmod(value: int) -> int:
    quotient = value // 36
    return value - 36 * quotient


_PAN_RE = re.compile(r"^[A-Z]{5}[0-9]{4}[A-Z]$", re.IGNORECASE)
_PAN_INVALID_CHARS = ("ABF", "IOS", "IDK", "XYZ", "ALK", "AZB", "CID", "DTN", "EOG", "EXH", "FFU", "FJI", "FLK", "FNR", "GBA", "GCU", "GFC", "GFX", "GHA", "GMH", "GMU", "GNC", "GNN", "GNR", "GOA", "GOE", "GPA", "GPH", "GPM", "GPN", "GPR", "GPU", "GPV", "GPW", "GTY", "GZA", "GZB", "GZL", "GZO", "GZP", "GZT", "HFA", "HFM", "HGS", "HHC", "HHU", "HIH", "HJI", "HJL", "HJQ", "HNR", "HOP", "HPK", "HPM", "HPN", "HPR", "HRH", "HRL", "HSB", "HUP", "HWF", "HWG", "HWR", "JAL", "JEW", "JJL", "JMY", "JNA", "JNZ", "JPY", "JRO", "JSL", "JWN")
_PAN_VALID_THIRD_CHAR = ("A", "B", "C", "F", "G", "H", "L", "J", "P", "T")


def validate_pan(pan: str) -> bool:
    if not pan or len(pan.strip()) != 10:
        return False
    pan = pan.strip().upper()
    if not _PAN_RE.match(pan):
        return False
    if pan[:3] in _PAN_INVALID_CHARS:
        return False
    if pan[3] not in _PAN_VALID_THIRD_CHAR:
        return False
    return True


_MOBILE_RE = re.compile(r"^(\+91[\-\s]?)?[6-9]\d{9}$")


def validate_mobile(mobile: str) -> bool:
    if not mobile:
        return False
    cleaned = mobile.strip().replace(" ", "").replace("-", "")
    return bool(_MOBILE_RE.match(cleaned))


_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")


def validate_email(email: str) -> bool:
    if not email or len(email) > 254:
        return False
    return bool(_EMAIL_RE.match(email.strip()))


_ONES = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
_TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]


def _chunk_to_words(n: int) -> str:
    if n == 0: return ""
    if n < 20: return _ONES[n]
    if n < 100: return (_TENS[n // 10] + " " + _ONES[n % 10]).strip()
    return (_ONES[n // 100] + " Hundred " + _chunk_to_words(n % 100)).strip()


def _rupees_in_words(amount: int) -> str:
    words = ""
    if amount >= 1_00_00_000:
        words += _chunk_to_words(amount // 1_00_00_000) + " Crore "
        amount %= 1_00_00_000
    if amount >= 1_00_000:
        words += _chunk_to_words(amount // 1_00_000) + " Lakh "
        amount %= 1_00_000
    if amount >= 1_000:
        words += _chunk_to_words(amount // 1_000) + " Thousand "
        amount %= 1_000
    if amount > 0:
        words += _chunk_to_words(amount) + " "
    return words.strip()


def amount_to_words(amount: Union[int, float, Decimal, str]) -> str:
    """Indian Rupees in words (Crore / Lakh / Thousand) with paise support.

    Always derived from the FINAL GRAND TOTAL. Examples:
        amount_to_words(5900)    -> "Rupees Five Thousand Nine Hundred Only"
        amount_to_words(5900.75) -> "Rupees Five Thousand Nine Hundred and Seventy Five Paise Only"
        amount_to_words(0)       -> "Rupees Zero Only"
    """
    if isinstance(amount, str):
        try:
            amount = Decimal(amount)
        except InvalidOperation:
            return "Rupees Zero Only"
    elif amount is None:
        return "Rupees Zero Only"
    elif not isinstance(amount, Decimal):
        try:
            amount = Decimal(str(amount))
        except (InvalidOperation, ValueError):
            return "Rupees Zero Only"
    if amount < 0:
        return "Minus " + amount_to_words(-amount)
    total_paise = int((amount * 100).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    rupees, paise = divmod(total_paise, 100)
    if total_paise == 0:
        return "Rupees Zero Only"
    rupee_words = _rupees_in_words(rupees)
    if paise == 0:
        if rupees == 1:
            return "Rupee One Only"
        return f"Rupees {rupee_words} Only"
    paise_words = _rupees_in_words(paise)
    if rupees == 0:
        return f"Rupees {paise_words} Paise Only"
    return f"Rupees {rupee_words} and {paise_words} Paise Only"


def get_financial_year(dt: Optional[date] = None) -> Tuple[date, date]:
    if dt is None: dt = date.today()
    if isinstance(dt, datetime): dt = dt.date()
    if dt.month >= 4: return date(dt.year, 4, 1), date(dt.year + 1, 3, 31)
    return date(dt.year - 1, 4, 1), date(dt.year, 3, 31)


def get_financial_year_code(dt: Optional[date] = None) -> str:
    if dt is None: dt = date.today()
    if dt.month >= 4: return f"{dt.year}-{(dt.year + 1) % 100:02d}"
    return f"{dt.year - 1}-{dt.year % 100:02d}"


def get_financial_year_prefix(dt: Optional[date] = None) -> str:
    if dt is None: dt = date.today()
    if dt.month >= 4: return f"{dt.year % 100:02d}{(dt.year + 1) % 100:02d}"
    return f"{(dt.year - 1) % 100:02d}{dt.year % 100:02d}"


def generate_invoice_number(existing_numbers: Optional[List[str]] = None, dt: Optional[date] = None,
                            commit: bool = True) -> str:
    """Allocate and permanently consume the next invoice number.

    The number is only permanent once the caller's transaction is committed.
    ``commit=True`` (default) commits the sequence bump immediately; pass
    ``commit=False`` from invoice-creation code so the allocation is part of the
    same transaction as the invoice and rolls back with it on failure.
    """
    if dt is None: dt = date.today()
    day = dt.strftime('%d%m%Y')
    prefix = f"INV-{day}-"
    seq = _allocate_sequence(prefix, existing_numbers, 'invoices', 'invoice_number', commit)
    return f"{prefix}{seq:03d}"


def peek_next_invoice_number(dt: Optional[date] = None) -> str:
    """Return the next invoice number WITHOUT consuming it (read-only preview).

    Opening the New Invoice form must never reserve a number; the number shown is
    provisional and the real one is allocated and committed only when the invoice
    is actually saved.
    """
    if dt is None: dt = date.today()
    day = dt.strftime('%d%m%Y')
    prefix = f"INV-{day}-"
    try:
        seq = int(db.session.execute(
            text("SELECT last_value FROM invoice_sequences WHERE seq_key = :key"),
            {'key': prefix},
        ).scalar() or 0)
    except Exception:
        seq = 0
    legacy = _sequence_seed_from_table('invoices', 'invoice_number', prefix)
    return f"{prefix}{max(seq, legacy) + 1:03d}"


def generate_quotation_number(existing_numbers: Optional[List[str]] = None, dt: Optional[date] = None) -> str:
    if dt is None: dt = date.today()
    day = dt.strftime('%d%m%Y')
    prefix = f"QTN-{day}-"
    seq = _allocate_sequence(prefix, existing_numbers, 'quotations', 'quotation_number')
    return f"{prefix}{seq:03d}"


def generate_purchase_order_number(existing_numbers: Optional[List[str]] = None, dt: Optional[date] = None) -> str:
    if dt is None: dt = date.today()
    day = dt.strftime('%d%m%Y')
    prefix = f"PO-{day}-"
    seq = _allocate_sequence(prefix, existing_numbers, 'purchase_orders', 'po_number')
    return f"{prefix}{seq:03d}"


def _sequence_seed_from_table(table: str, column: str, prefix: str) -> int:
    """Compute the current max sequence for a numbered column (legacy seed)."""
    try:
        row = db.session.execute(text(
            "SELECT COALESCE(MAX(CAST(SUBSTR(%s, LENGTH(:p) + 1) AS INTEGER)), 0) "
            "FROM %s WHERE %s LIKE :like" % (column, table, column)
        ), {'p': prefix, 'like': prefix + '%'}).scalar()
        return int(row or 0)
    except Exception:
        return 0


def _allocate_sequence(prefix: str, existing_numbers: Optional[List[str]],
                       table: str, column: str, commit: bool = True) -> int:
    """Atomically allocate the next sequence number for a numbering group.

    Serialization is done at the database level through the invoice_sequences
    table: the counter is bumped with a single atomic ``UPDATE ... SET
    last_value = last_value + 1`` so two concurrent requests can never observe
    the same value (works on both PostgreSQL and SQLite). The first allocation
    for a given prefix seeds the counter from any legacy numbers so numbering
    never collides with or recycles existing document numbers.

    When ``commit`` is True (default) the bump is committed immediately, which is
    what standalone number-generator callers want. When ``commit`` is False the
    bump is only flushed, so it stays inside the caller's transaction and is
    rolled back together with it if the caller's save fails — the number is then
    never consumed by a failed invoice creation.

    Falls back to the legacy scan when the table is unavailable, so the app can
    never break on an unusual database state.
    """
    _ts = datetime.utcnow()
    for _ in range(10):
        try:
            upd = db.session.execute(
                text("UPDATE invoice_sequences SET last_value = last_value + 1, "
                     "updated_at = :ts WHERE seq_key = :key"),
                {'ts': _ts, 'key': prefix},
            )
            if upd.rowcount == 1:
                new_val = db.session.execute(
                    text("SELECT last_value FROM invoice_sequences WHERE seq_key = :key"),
                    {'key': prefix},
                ).scalar()
                if commit:
                    db.session.commit()
                return int(new_val)
            seed = _next_sequence(existing_numbers or [], prefix)
            if not existing_numbers:
                seed = _sequence_seed_from_table(table, column, prefix)
            db.session.execute(
                text("INSERT INTO invoice_sequences (seq_key, prefix, period, last_value, updated_at) "
                     "VALUES (:key, :key, :key, :seed, :ts)"),
                {'key': prefix, 'seed': seed, 'ts': _ts},
            )
            if commit:
                db.session.commit()
        except IntegrityError:
            db.session.rollback()
        except OperationalError:
            db.session.rollback()
            time.sleep(0.02)
    db.session.rollback()
    if existing_numbers:
        return _next_sequence(existing_numbers, prefix) + 1
    return _sequence_seed_from_table(table, column, prefix) + 1


def _next_sequence(existing_numbers: List[str], prefix: str) -> int:
    max_seq = 0
    for num in existing_numbers:
        if num and num.startswith(prefix):
            try:
                seq_int = int(num[len(prefix):])
                if seq_int > max_seq: max_seq = seq_int
            except (ValueError, IndexError): continue
    return max_seq


_XSS_PATTERNS = [re.compile(r"<script\b[^>]*>", re.IGNORECASE), re.compile(r"</script>", re.IGNORECASE), re.compile(r"javascript\s*:", re.IGNORECASE), re.compile(r"on\w+\s*=", re.IGNORECASE), re.compile(r"<iframe\b", re.IGNORECASE), re.compile(r"<object\b", re.IGNORECASE), re.compile(r"<embed\b", re.IGNORECASE), re.compile(r"<link\b", re.IGNORECASE), re.compile(r"<style\b", re.IGNORECASE), re.compile(r"expression\s*\(", re.IGNORECASE), re.compile(r"data\s*:\s*text/html", re.IGNORECASE), re.compile(r"<\s*img\b[^>]+onerror", re.IGNORECASE), re.compile(r"<\s*svg\b[^>]+onload", re.IGNORECASE)]
_SQL_INJECTION_PATTERNS = [re.compile(r"(SELECT|INSERT|UPDATE|DELETE|DROP|TRUNCATE|ALTER|CREATE|EXEC|EXECUTE|UNION|FETCH|DECLARE)\b", re.IGNORECASE), re.compile(r"(--|#|/\*|\*/)", re.IGNORECASE), re.compile(r"\b(OR|AND)\b\s+\d+\s*=\s*\d+", re.IGNORECASE), re.compile(r"['\";]\s*(OR|AND)\s+['\"]", re.IGNORECASE), re.compile(r"\bWAITFOR\s+DELAY\b", re.IGNORECASE), re.compile(r"\bBENCHMARK\s*\(", re.IGNORECASE), re.compile(r"\bSLEEP\s*\(", re.IGNORECASE), re.compile(r"\bLOAD_FILE\s*\(", re.IGNORECASE), re.compile(r"\bINTO\s+(OUT|DUMP)FILE\b", re.IGNORECASE)]


def sanitize_input(value: str) -> str:
    if not value: return ""
    cleaned = str(value).strip()
    for p in _XSS_PATTERNS: cleaned = p.sub("", cleaned)
    cleaned = re.sub(r"<[^>]+>", "", cleaned)
    for p in _SQL_INJECTION_PATTERNS: cleaned = p.sub("", cleaned)
    return cleaned


def contains_xss(value: str) -> bool:
    if not value: return False
    for p in _XSS_PATTERNS:
        if p.search(value): return True
    return False


def contains_sql_injection(value: str) -> bool:
    if not value: return False
    for p in _SQL_INJECTION_PATTERNS:
        if p.search(value): return True
    return False


def format_currency(amount, symbol=True, decimal_places=2) -> str:
    """Centralized Indian currency formatter.

    Indian thousands grouping, "Rs." prefix (PDF-safe ASCII symbol):
        format_currency(5900)          -> "Rs. 5,900.00"
        format_currency(100000)        -> "Rs. 1,00,000.00"
        format_currency(5900, False)   -> "5,900.00"
    Uses Decimal everywhere (no float arithmetic); rounds to 2 decimals.
    """
    if isinstance(amount, str):
        try:
            amount = Decimal(amount)
        except InvalidOperation:
            return "Rs. 0.00" if symbol else "0.00"
    if isinstance(amount, (int, float)):
        amount = Decimal(str(amount))
    places = Decimal(10) ** -decimal_places if decimal_places > 0 else Decimal("1")
    amount = amount.quantize(places, rounding=ROUND_HALF_UP)
    negative = amount < 0
    amount = abs(amount)
    int_part = int(amount)
    frac_str = str(amount - int_part)[1:] if decimal_places > 0 else ""
    int_str = str(int_part)
    if len(int_str) <= 3:
        formatted_int = int_str
    else:
        last_three = int_str[-3:]
        remaining = int_str[:-3]
        groups = []
        while remaining:
            groups.append(remaining[-2:])
            remaining = remaining[:-2]
        formatted_int = ",".join(reversed(groups)) + "," + last_three
    result = formatted_int + frac_str
    if symbol:
        result = "Rs. " + result
    if negative:
        result = "-" + result
    return result


def format_indian_currency(amount, symbol=True, decimal_places=2) -> str:
    """Backwards-compatible alias for :func:`format_currency`."""
    return format_currency(amount, symbol=symbol, decimal_places=decimal_places)


def get_csrf_token() -> str:
    try:
        from flask_wtf.csrf import generate_csrf
        return generate_csrf()
    except (ImportError, RuntimeError):
        return session.get("csrf_token", "")


def get_csrf_headers() -> Dict[str, str]:
    return {"X-CSRFToken": get_csrf_token(), "X-Requested-With": "XMLHttpRequest"}


def to_int(value, default=None):
    if value is None or value == '':
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return default


def extract_state_code_from_gstin(gstin: str) -> Optional[int]:
    if not gstin or len(gstin) < 2: return None
    try:
        code = int(gstin[:2])
        if code in GST_STATE_CODES: return code
    except (ValueError, TypeError): pass
    return None


def get_state_name_from_code(code: int) -> Optional[str]:
    return GST_STATE_CODES.get(code)


def get_state_code_from_name(name: str) -> Optional[int]:
    for state_name, code in GST_STATE_NAMES_TO_CODES.items():
        if state_name.lower() == name.strip().lower(): return code
    return None


def determine_business_type(customer_gstin: Optional[str] = None, **kwargs) -> str:
    if customer_gstin and validate_gstIN(customer_gstin): return "B2B"
    return "B2C"


def setup_logging(base_dir: str):
    log_dir = os.path.join(base_dir, "logs")
    os.makedirs(log_dir, exist_ok=True)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(name)s: %(message)s")
    daily = logging.handlers.TimedRotatingFileHandler(os.path.join(log_dir, "daily.log"), when="midnight", interval=1, backupCount=90)
    daily.setFormatter(fmt); daily.setLevel(logging.INFO)
    error = logging.handlers.RotatingFileHandler(os.path.join(log_dir, "error.log"), maxBytes=5*1024*1024, backupCount=10)
    error.setFormatter(fmt); error.setLevel(logging.ERROR)
    inv_log = logging.handlers.RotatingFileHandler(os.path.join(log_dir, "invoice.log"), maxBytes=5*1024*1024, backupCount=10)
    inv_log.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")); inv_log.setLevel(logging.INFO)
    sec_log = logging.handlers.RotatingFileHandler(os.path.join(log_dir, "security.log"), maxBytes=5*1024*1024, backupCount=10)
    sec_log.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")); sec_log.setLevel(logging.WARNING)
    root = logging.getLogger(); root.setLevel(logging.INFO); root.addHandler(daily); root.addHandler(error)
    logging.getLogger("invoice").addHandler(inv_log); logging.getLogger("invoice").setLevel(logging.INFO)
    logging.getLogger("security").addHandler(sec_log); logging.getLogger("security").setLevel(logging.WARNING)


############################################################
# DATABASE
############################################################

db = SQLAlchemy()
csrf = CSRFProtect()
login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.login_message_category = 'warning'
_app_initialized = False


def _ensure_product_columns():
    """Add columns introduced after the products table was first created (SQLite/Postgres safe)."""
    if not hasattr(db, 'engine'):
        return
    dialect = db.engine.dialect.name
    with db.engine.connect() as conn:
        if dialect == 'sqlite':
            cols = [r[1] for r in conn.execute(text("PRAGMA table_info(products)"))]
        else:
            rows = conn.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name = 'products'"
            ))
            cols = [r[0] for r in rows]
        additions = {
            'location': "VARCHAR(100)",
            'status': "VARCHAR(20) DEFAULT 'active'",
            'last_purchase': "DATETIME",
            'last_sale': "DATETIME",
            'last_low_stock_notification_at': "DATETIME",
            'low_stock_alert_active': "BOOLEAN DEFAULT 0",
        }
        for col, ddl in additions.items():
            if col not in cols:
                conn.execute(text("ALTER TABLE products ADD COLUMN %s %s" % (col, ddl)))
        conn.commit()


def _ensure_payment_columns():
    """Add columns introduced for the Accounts Receivable / payment workflow
    without dropping the existing database (SQLite/Postgres safe)."""
    if not hasattr(db, 'engine'):
        return
    dialect = db.engine.dialect.name
    with db.engine.connect() as conn:
        def _cols(table):
            if dialect == 'sqlite':
                return [r[1] for r in conn.execute(text("PRAGMA table_info(%s)" % table))]
            rows = conn.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name = '%s'" % table))
            return [r[0] for r in rows]

        invoice_additions = {'balance_due': "NUMERIC(12,2) DEFAULT 0"}
        for col, ddl in invoice_additions.items():
            if col not in _cols('invoices'):
                conn.execute(text("ALTER TABLE invoices ADD COLUMN %s %s" % (col, ddl)))

        payment_additions = {
            'customer_id': "INTEGER",
            'transaction_id': "VARCHAR(80)",
            'utr': "VARCHAR(50)",
            'remarks': "TEXT",
            'received_by': "VARCHAR(80)",
            'updated_at': "DATETIME",
        }
        for col, ddl in payment_additions.items():
            if col not in _cols('payments'):
                conn.execute(text("ALTER TABLE payments ADD COLUMN %s %s" % (col, ddl)))
        conn.execute(text(
            "UPDATE invoices SET balance_due = "
            "CASE WHEN grand_total IS NULL THEN 0 ELSE grand_total - COALESCE(amount_paid, 0) END "
            "WHERE balance_due IS NULL"))
        conn.commit()


def _ensure_invoice_sequence_columns():
    """Add columns introduced for quotation→invoice conversion and safe
    numbering without dropping the existing database (SQLite/Postgres safe)."""
    if not hasattr(db, 'engine'):
        return
    dialect = db.engine.dialect.name
    with db.engine.connect() as conn:
        def _cols(table):
            if dialect == 'sqlite':
                return [r[1] for r in conn.execute(text("PRAGMA table_info(%s)" % table))]
            rows = conn.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name = '%s'" % table))
            return [r[0] for r in rows]

        if 'quotation_id' not in _cols('invoices'):
            conn.execute(text(
                "ALTER TABLE invoices ADD COLUMN quotation_id INTEGER REFERENCES quotations(id)"))
        conn.execute(text(
            "CREATE UNIQUE INDEX IF NOT EXISTS ix_invoices_quotation_id "
            "ON invoices (quotation_id)"))
        conn.commit()


def _drop_bank_columns():
    """Safely remove bank-related columns (SQLite/Postgres safe, idempotent).

    Bank features have been removed from the application. The columns are
    dropped defensively so stale schema can never leak bank data. Failures are
    logged and swallowed: the application models no longer reference them.
    """
    if not hasattr(db, 'engine'):
        return
    dialect = db.engine.dialect.name
    drops = {
        'payments': ['bank_name', 'cheque_number'],
        'suppliers': ['bank_name', 'bank_account', 'bank_ifsc', 'upi_id'],
    }
    try:
        with db.engine.connect() as conn:
            for table, cols in drops.items():
                if dialect == 'sqlite':
                    existing = [r[1] for r in conn.execute(text("PRAGMA table_info(%s)" % table))]
                else:
                    rows = conn.execute(text(
                        "SELECT column_name FROM information_schema.columns WHERE table_name = '%s'" % table))
                    existing = [r[0] for r in rows]
                for col in cols:
                    if col not in existing:
                        continue
                    try:
                        conn.execute(text("ALTER TABLE %s DROP COLUMN %s" % (table, col)))
                        conn.commit()
                    except Exception:
                        conn.rollback()
                        current_app.logger.warning('Could not drop column %s.%s (non-fatal)', table, col)
            conn.commit()
    except Exception:
        try:
            current_app.logger.warning('Bank column cleanup skipped (non-fatal).')
        except Exception:
            pass


def create_app(config_class=None):
    global _app_initialized
    if _app_initialized:
        return app
    _app_initialized = True
    if config_class is None: config_class = get_config()
    app.config.from_object(config_class)
    app.config['COMPANY'] = dict(_FIXED_COMPANY)
    app.config['SESSION_TYPE'] = 'cookie'
    for d in ['UPLOAD_FOLDER', 'BACKUP_FOLDER', 'PDF_FOLDER', 'EXPORT_FOLDER', 'LOG_FOLDER']:
        os.makedirs(app.config.get(d, d.lower()), exist_ok=True)
    db.init_app(app)
    csrf.init_app(app)
    login_manager.init_app(app)

    @login_manager.user_loader
    def load_user(user_id):
        return db.session.get(User, int(user_id))
    app.jinja_env.globals['format_indian_currency'] = format_indian_currency
    app.jinja_env.globals['format_currency'] = format_currency
    app.jinja_env.globals['get_csrf_token'] = get_csrf_token
    app.jinja_env.globals['GST_STATE_CODES'] = GST_STATE_CODES
    app.jinja_env.globals['_payment_status_label'] = _payment_status_label
    app.jinja_env.globals['_payment_method_label'] = _payment_method_label
    app.jinja_env.globals['_payment_display_ref'] = _payment_display_ref
    app.jinja_env.globals['_payments_newest'] = _payments_newest
    setup_logging(app.config['BASE_DIR'])
    with app.app_context():
        db.create_all()
        _ensure_product_columns()
        _ensure_payment_columns()
        _ensure_invoice_sequence_columns()
        _drop_bank_columns()
        _seed_database()
        _ensure_default_settings()
        _load_company_settings(app)
    return app


app = Flask(__name__)


############################################################
# MODELS
############################################################


class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), default='staff')
    is_active = db.Column(db.Boolean, default=True)
    last_login = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    def set_password(self, password): self.password_hash = generate_password_hash(password)
    def check_password(self, password): return check_password_hash(self.password_hash, password)
    @property
    def is_admin(self): return self.role == 'admin'


class Customer(db.Model):
    __tablename__ = 'customers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    mobile = db.Column(db.String(20))
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    gstin = db.Column(db.String(15))
    state = db.Column(db.String(50))
    state_code = db.Column(db.Integer, default=29)
    total_purchases = db.Column(db.Numeric(12, 2), default=0)
    invoice_count = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    invoices = db.relationship('Invoice', backref='customer', lazy=True)


class Category(db.Model):
    __tablename__ = 'categories'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)
    description = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    products = db.relationship('Product', backref='category', lazy=True)


class Product(db.Model):
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    sku = db.Column(db.String(50), unique=True)
    barcode = db.Column(db.String(50), unique=True, nullable=True)
    hsn = db.Column(db.String(10))
    brand = db.Column(db.String(100), nullable=True)
    category_id = db.Column(db.Integer, db.ForeignKey('categories.id'))
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'), nullable=True)
    description = db.Column(db.Text)
    unit = db.Column(db.String(20), default='pcs')
    purchase_price = db.Column(db.Numeric(12, 2), default=0)
    selling_price = db.Column(db.Numeric(12, 2), default=0)
    gst_rate = db.Column(db.Numeric(5, 2), default=18)
    opening_stock = db.Column(db.Integer, default=0)
    stock_quantity = db.Column(db.Integer, default=0)
    min_stock = db.Column(db.Integer, default=0)
    max_stock = db.Column(db.Integer, default=500)
    last_low_stock_notification_at = db.Column(db.DateTime, nullable=True)
    low_stock_alert_active = db.Column(db.Boolean, default=False)
    location = db.Column(db.String(100), nullable=True)
    status = db.Column(db.String(20), default='active')
    last_purchase = db.Column(db.DateTime, nullable=True)
    last_sale = db.Column(db.DateTime, nullable=True)
    warehouse = db.Column(db.String(100), nullable=True)
    warranty = db.Column(db.String(50), nullable=True)
    image = db.Column(db.String(256), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    @property
    def current_stock(self):
        return self.stock_quantity

    @property
    def price(self):
        return self.selling_price

    @property
    def is_low_stock(self) -> bool:
        return (self.current_stock or 0) <= (self.min_stock or 0)

    @property
    def stock_status(self) -> str:
        if (self.current_stock or 0) <= 0:
            return 'out_of_stock'
        if self.is_low_stock:
            return 'low_stock'
        return 'in_stock'

    @property
    def available_stock_value(self) -> float:
        return round((self.current_stock or 0) * float(self.purchase_price or 0), 2)


class StockMovement(db.Model):
    __tablename__ = 'stock_movements'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    movement_type = db.Column(db.String(20), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    reference_type = db.Column(db.String(50))
    reference_id = db.Column(db.Integer)
    notes = db.Column(db.Text)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    product = db.relationship('Product', backref='movements')
    user = db.relationship('User', backref='stock_movements')


class Supplier(db.Model):
    __tablename__ = 'suppliers'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    contact_person = db.Column(db.String(120))
    mobile = db.Column(db.String(20))
    email = db.Column(db.String(120))
    address = db.Column(db.Text)
    gstin = db.Column(db.String(15))
    state = db.Column(db.String(50))
    state_code = db.Column(db.Integer, default=29)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    products = db.relationship('Product', backref='supplier', lazy=True)


class Invoice(db.Model):
    __tablename__ = 'invoices'
    id = db.Column(db.Integer, primary_key=True)
    invoice_number = db.Column(db.String(30), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customers.id'))
    customer_name = db.Column(db.String(120))
    customer_mobile = db.Column(db.String(20))
    customer_email = db.Column(db.String(120))
    customer_address = db.Column(db.Text)
    customer_gstin = db.Column(db.String(15))
    customer_state = db.Column(db.String(50))
    customer_state_code = db.Column(db.Integer)
    invoice_date = db.Column(db.Date, nullable=False)
    due_date = db.Column(db.Date)
    subtotal = db.Column(db.Numeric(12, 2), default=0)
    total_discount = db.Column(db.Numeric(12, 2), default=0)
    total_taxable = db.Column(db.Numeric(12, 2), default=0)
    total_cgst = db.Column(db.Numeric(12, 2), default=0)
    total_sgst = db.Column(db.Numeric(12, 2), default=0)
    total_igst = db.Column(db.Numeric(12, 2), default=0)
    round_off = db.Column(db.Numeric(12, 2), default=0)
    grand_total = db.Column(db.Numeric(12, 2), default=0)
    amount_paid = db.Column(db.Numeric(12, 2), default=0)
    balance_due = db.Column(db.Numeric(12, 2), default=0)
    payment_method = db.Column(db.String(30))
    status = db.Column(db.String(20), default='draft')
    payment_status = db.Column(db.String(20), default='due')
    is_intra_state = db.Column(db.Boolean, default=True)
    notes = db.Column(db.Text)
    terms = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    items = db.relationship('InvoiceItem', backref='invoice', lazy=True, cascade='all, delete-orphan')
    payments = db.relationship('Payment', backref='invoice', lazy=True, cascade='all, delete-orphan')
    creator = db.relationship('User', backref='created_invoices')
    quotation_id = db.Column(db.Integer, db.ForeignKey('quotations.id'), unique=True, nullable=True)
    quotation = db.relationship('Quotation', foreign_keys=[quotation_id],
                                backref=db.backref('converted_invoices', lazy=True))

    @property
    def balance(self):
        """Outstanding balance (template convenience)."""
        return (self.grand_total or 0) - (self.amount_paid or 0)


class InvoiceItem(db.Model):
    __tablename__ = 'invoice_items'
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoices.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    product_name = db.Column(db.String(120))
    hsn = db.Column(db.String(10))
    qty = db.Column(db.Integer, default=1)
    unit = db.Column(db.String(20), default='pcs')
    price = db.Column(db.Numeric(12, 2), default=0)
    discount = db.Column(db.Numeric(5, 2), default=0)
    gst_rate = db.Column(db.Numeric(5, 2), default=18)
    taxable_value = db.Column(db.Numeric(12, 2), default=0)
    cgst = db.Column(db.Numeric(12, 2), default=0)
    sgst = db.Column(db.Numeric(12, 2), default=0)
    igst = db.Column(db.Numeric(12, 2), default=0)
    total = db.Column(db.Numeric(12, 2), default=0)


class Payment(db.Model):
    __tablename__ = 'payments'
    id = db.Column(db.Integer, primary_key=True)
    invoice_id = db.Column(db.Integer, db.ForeignKey('invoices.id'), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customers.id'))
    payment_date = db.Column(db.Date, default=date.today)
    amount = db.Column(db.Numeric(12, 2), nullable=False)
    payment_method = db.Column(db.String(30))
    reference_number = db.Column(db.String(50))
    transaction_id = db.Column(db.String(80))
    utr = db.Column(db.String(50))
    remarks = db.Column(db.Text)
    notes = db.Column(db.Text)
    received_by = db.Column(db.String(80))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    customer = db.relationship('Customer', backref='payments')


class Quotation(db.Model):
    __tablename__ = 'quotations'
    id = db.Column(db.Integer, primary_key=True)
    quotation_number = db.Column(db.String(30), unique=True, nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customers.id'))
    customer_name = db.Column(db.String(120))
    customer_mobile = db.Column(db.String(20))
    customer_email = db.Column(db.String(120))
    customer_address = db.Column(db.Text)
    customer_gstin = db.Column(db.String(15))
    customer_state = db.Column(db.String(50))
    customer_state_code = db.Column(db.Integer)
    quotation_date = db.Column(db.Date, nullable=False)
    valid_until = db.Column(db.Date)
    subtotal = db.Column(db.Numeric(12, 2), default=0)
    total_discount = db.Column(db.Numeric(12, 2), default=0)
    total_taxable = db.Column(db.Numeric(12, 2), default=0)
    total_cgst = db.Column(db.Numeric(12, 2), default=0)
    total_sgst = db.Column(db.Numeric(12, 2), default=0)
    total_igst = db.Column(db.Numeric(12, 2), default=0)
    round_off = db.Column(db.Numeric(12, 2), default=0)
    grand_total = db.Column(db.Numeric(12, 2), default=0)
    is_intra_state = db.Column(db.Boolean, default=True)
    notes = db.Column(db.Text)
    terms = db.Column(db.Text)
    status = db.Column(db.String(20), default='draft')
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    items = db.relationship('QuotationItem', backref='quotation', lazy=True, cascade='all, delete-orphan')
    creator = db.relationship('User', backref='created_quotations')


class QuotationItem(db.Model):
    __tablename__ = 'quotation_items'
    id = db.Column(db.Integer, primary_key=True)
    quotation_id = db.Column(db.Integer, db.ForeignKey('quotations.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    product_name = db.Column(db.String(120))
    hsn = db.Column(db.String(10))
    qty = db.Column(db.Integer, default=1)
    unit = db.Column(db.String(20), default='pcs')
    price = db.Column(db.Numeric(12, 2), default=0)
    discount = db.Column(db.Numeric(5, 2), default=0)
    gst_rate = db.Column(db.Numeric(5, 2), default=18)
    taxable_value = db.Column(db.Numeric(12, 2), default=0)
    cgst = db.Column(db.Numeric(12, 2), default=0)
    sgst = db.Column(db.Numeric(12, 2), default=0)
    igst = db.Column(db.Numeric(12, 2), default=0)
    total = db.Column(db.Numeric(12, 2), default=0)


class PurchaseOrder(db.Model):
    __tablename__ = 'purchase_orders'
    id = db.Column(db.Integer, primary_key=True)
    po_number = db.Column(db.String(30), unique=True, nullable=False)
    supplier_id = db.Column(db.Integer, db.ForeignKey('suppliers.id'))
    supplier_name = db.Column(db.String(120))
    order_date = db.Column(db.Date, nullable=False)
    expected_date = db.Column(db.Date)
    subtotal = db.Column(db.Numeric(12, 2), default=0)
    total_tax = db.Column(db.Numeric(12, 2), default=0)
    grand_total = db.Column(db.Numeric(12, 2), default=0)
    status = db.Column(db.String(20), default='draft')
    notes = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    items = db.relationship('PurchaseItem', backref='purchase_order', lazy=True, cascade='all, delete-orphan')
    supplier = db.relationship('Supplier', backref='purchase_orders')
    creator = db.relationship('User', backref='created_purchases')


class PurchaseItem(db.Model):
    __tablename__ = 'purchase_items'
    id = db.Column(db.Integer, primary_key=True)
    purchase_order_id = db.Column(db.Integer, db.ForeignKey('purchase_orders.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    product_name = db.Column(db.String(120))
    hsn = db.Column(db.String(10))
    qty = db.Column(db.Integer, default=1)
    unit = db.Column(db.String(20), default='pcs')
    price = db.Column(db.Numeric(12, 2), default=0)
    gst_rate = db.Column(db.Numeric(5, 2), default=18)
    total = db.Column(db.Numeric(12, 2), default=0)


class Settings(db.Model):
    __tablename__ = 'settings'
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class GSTMaster(db.Model):
    __tablename__ = 'gst_master'
    id = db.Column(db.Integer, primary_key=True)
    state_code = db.Column(db.Integer, nullable=False)
    state_name = db.Column(db.String(50), nullable=False)
    cgst_rate = db.Column(db.Numeric(5, 2), default=0)
    sgst_rate = db.Column(db.Numeric(5, 2), default=0)
    igst_rate = db.Column(db.Numeric(5, 2), default=0)
    is_active = db.Column(db.Boolean, default=True)


class AuditLog(db.Model):
    __tablename__ = 'audit_logs'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    action = db.Column(db.String(100), nullable=False)
    entity_type = db.Column(db.String(50))
    entity_id = db.Column(db.Integer)
    details = db.Column(db.Text)
    ip_address = db.Column(db.String(45))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user = db.relationship('User', backref='audit_logs')


class Notification(db.Model):
    __tablename__ = 'notifications'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text)
    is_read = db.Column(db.Boolean, default=False)
    notification_type = db.Column(db.String(50), default='info')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user = db.relationship('User', backref='notifications')


class InvoiceSequence(db.Model):
    """Database-backed counters for atomic, never-recycled document numbering.

    One row per numbering group (e.g. ``INV-16082026-``). Values are bumped
    with an atomic UPDATE so concurrent requests always receive unique numbers,
    and a row's counter never decreases — cancelled documents keep their number
    forever.
    """
    __tablename__ = 'invoice_sequences'
    id = db.Column(db.Integer, primary_key=True)
    seq_key = db.Column(db.String(40), unique=True, nullable=False, index=True)
    prefix = db.Column(db.String(30), nullable=False)
    period = db.Column(db.String(20), nullable=False)
    last_value = db.Column(db.Integer, nullable=False, default=0)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


############################################################
# HELPER FUNCTIONS
############################################################


def log_audit(user_id, action, entity_type=None, entity_id=None, details=None, ip_address=None):
    db.session.add(AuditLog(user_id=user_id, action=action, entity_type=entity_type, entity_id=entity_id, details=details, ip_address=ip_address))
    db.session.commit()

def get_setting(key, default=None):
    s = Settings.query.filter_by(key=key).first()
    return s.value if s else default

def set_setting(key, value):
    s = Settings.query.filter_by(key=key).first()
    if s: s.value = str(value)
    else: s = Settings(key=key, value=str(value)); db.session.add(s)
    db.session.commit()


PAYMENT_METHODS = ['cash', 'upi', 'card', 'split']


def _payment_method_label(method):
    return {
        'cash': 'Cash', 'upi': 'UPI', 'card': 'Card',
        'split': 'Split Payment',
    }.get((method or '').lower(), (method or 'Cash').replace('_', ' ').title())


def _payment_status_label(ps):
    return {'due': 'Due', 'partial': 'Partially Paid', 'paid': 'Paid',
            'cancelled': 'Cancelled', 'pending': 'Due'}.get((ps or '').lower(), (ps or 'Due').capitalize())


def _payment_balance(inv):
    return max(Decimal('0'), (inv.grand_total or Decimal('0')) - (inv.amount_paid or Decimal('0')))


def _set_payment_state(inv, amount_paid):
    """Keep Invoice.amount_paid / balance_due / payment_status in sync."""
    amount_paid = Decimal(str(amount_paid or '0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    if amount_paid < 0:
        amount_paid = Decimal('0')
    inv.amount_paid = amount_paid
    balance = (inv.grand_total or Decimal('0')) - amount_paid
    if balance < Decimal('0'):
        balance = Decimal('0')
    inv.balance_due = balance
    if inv.status == 'cancelled':
        inv.payment_status = 'cancelled'
    elif amount_paid <= Decimal('0'):
        inv.payment_status = 'due'
    elif balance <= Decimal('0.005'):
        inv.payment_status = 'paid'
    else:
        inv.payment_status = 'partial'


def _payment_display_ref(p):
    """Single human-readable reference for a payment."""
    for f in ('reference_number', 'utr', 'transaction_id'):
        v = getattr(p, f, None)
        if v and str(v).strip():
            return str(v).strip()
    return ''


def _payments_newest(inv):
    """Payments newest-first, NULL dates pushed to the end."""
    try:
        return sorted(list(inv.payments or []),
                      key=lambda p: (p.payment_date or date.min, p.id or 0), reverse=True)
    except TypeError:
        return list(inv.payments or [])


def _public_settings():
    _hidden = {'smtp_server', 'smtp_port', 'smtp_email', 'smtp_password',
               'mail_server', 'mail_port', 'mail_username', 'mail_password',
               'company_pan', 'company_qr'}
    return {s.key: s.value for s in Settings.query.all() if s.key not in _hidden}


class BackupService:
    @staticmethod
    def create_backup(app):
        backup_dir = app.config.get('BACKUP_FOLDER', 'backups/database')
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = os.path.join(backup_dir, f"gv_powers_erp_{timestamp}.sql")
        db_url = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if 'sqlite' in db_url:
            db_path = db_url.replace('sqlite:///', '')
            if not os.path.isabs(db_path):
                db_path = os.path.join(app.config.get('BASE_DIR', '.'), db_path)
            if not os.path.exists(db_path):
                raise Exception(f"SQLite database file not found: {db_path}")
            shutil.copy2(db_path, filepath + '.db')
            return filepath + '.db'
        result = subprocess.run(['pg_dump', db_url, '-f', filepath], capture_output=True, text=True, timeout=300)
        if result.returncode == 0: return filepath
        raise Exception(f"Backup failed: {result.stderr}")

    @staticmethod
    def list_backups(app):
        backup_dir = app.config.get('BACKUP_FOLDER', 'backups/database')
        if not os.path.exists(backup_dir): return []
        return [{'filename': f, 'filepath': os.path.join(backup_dir, f), 'size': os.path.getsize(os.path.join(backup_dir, f)), 'created': datetime.fromtimestamp(os.path.getmtime(os.path.join(backup_dir, f)))} for f in sorted(os.listdir(backup_dir), reverse=True)]


def _load_company_settings(app):
    """Refresh in-memory COMPANY config from the fixed GV POWERS identity.

    Company information is FIXED for this installation and is never editable from
    the admin panel. The config dict is always rebuilt from the fixed defaults so
    stale/removed database values can never leak back into the admin panel, PDFs
    or reports.
    """
    try:
        app.config['COMPANY'] = dict(_FIXED_COMPANY)
    except Exception:
        pass


def _ensure_default_settings():
    """Clean up stale bank-related settings if they exist."""
    changed = False
    removed = Settings.query.filter(Settings.key.in_([
        'company_pan', 'bank_name', 'bank_branch', 'bank_account',
        'bank_ifsc', 'upi_id', 'company_qr',
    ])).all()
    for row in removed:
        db.session.delete(row)
        changed = True
    if changed:
        db.session.commit()


def _seed_database():
    """No default data seeded. Start with a clean database."""


############################################################
# GST ENGINE
############################################################


class GSTService:
    @staticmethod
    def calculate_gst(amount: Decimal, gst_rate: Decimal, is_intra_state: bool = True) -> Dict[str, Decimal]:
        gst_amount = (amount * gst_rate / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if is_intra_state:
            half = (gst_amount / Decimal('2')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            return {'cgst': half, 'sgst': half, 'igst': Decimal('0.00'), 'total_tax': gst_amount}
        return {'cgst': Decimal('0.00'), 'sgst': Decimal('0.00'), 'igst': gst_amount, 'total_tax': gst_amount}

    @staticmethod
    def determine_tax_type(supplier_state_code: int, customer_state_code: int) -> bool:
        return supplier_state_code == customer_state_code

    @staticmethod
    def get_gst_summary(invoice):
        summary = {'cgst_total': Decimal('0'), 'sgst_total': Decimal('0'), 'igst_total': Decimal('0'), 'taxable_total': Decimal('0'), 'hsn_wise': {}}
        for item in invoice.items:
            hsn = item.hsn or 'N/A'
            if hsn not in summary['hsn_wise']:
                summary['hsn_wise'][hsn] = {'taxable': Decimal('0'), 'cgst': Decimal('0'), 'sgst': Decimal('0'), 'igst': Decimal('0'), 'rate': item.gst_rate}
            for k in ['taxable', 'cgst', 'sgst', 'igst']:
                val = getattr(item, 'taxable_value' if k == 'taxable' else k)
                summary['hsn_wise'][hsn][k] += val
                summary[f'{k}_total'] += val
            summary['taxable_total'] += item.taxable_value
        summary['total_tax'] = summary['cgst_total'] + summary['sgst_total'] + summary['igst_total']
        return summary


############################################################
# PDF GENERATOR — PROFESSIONAL COMMERCIAL INVOICE
############################################################

# --- Color Palette ---
_PDF_NAVY       = rl_colors.HexColor("#0D1B2A")
_PDF_BLUE       = rl_colors.HexColor("#1B4F8A")
_PDF_LIGHT      = rl_colors.HexColor("#F4F6F9")
_PDF_BORDER     = rl_colors.HexColor("#D0D5DD")
_PDF_DARK       = rl_colors.HexColor("#1A1A2E")
_PDF_MUTED      = rl_colors.HexColor("#5A6577")
_PDF_WATERMARK  = rl_colors.HexColor("#E8EBF0")
_PDF_WHITE      = rl_colors.white
_PDF_GREEN      = rl_colors.HexColor("#1A7F4B")
_PDF_RED        = rl_colors.HexColor("#B42318")

# --- Layout ---
_PDF_M          = 36
_PDF_TOP        = 36
_PDF_BOTTOM     = 44
_PDF_PAGE_W, _PDF_PAGE_H = A4
_PDF_USABLE_W   = _PDF_PAGE_W - 2 * _PDF_M

# --- Body font sizes (compact, professional) ---
_INV_BODY       = 8.5
_INV_BODY_LEAD  = 10.5
_INV_SMALL      = 8
_INV_SMALL_LEAD = 10
_INV_TINY       = 7
_INV_TINY_LEAD  = 9


def _pdf_inr(value, signed=False):
    """Rs. + Indian thousands separator (e.g. Rs. 59,000.00)."""
    if value is None:
        value = Decimal('0')
    amt = format_indian_currency(value, symbol=False)
    if signed and Decimal(str(value)) > 0:
        amt = "+" + amt
    return "Rs. " + amt


def _pdf_date(v):
    if not v:
        return ''
    return v.strftime('%d-%m-%Y')


def _pdf_styles():
    s = getSampleStyleSheet()
    styles_kw = {
        'SectionTitle': dict(fontSize=10, fontName=_PDF_FONT_BOLD, textColor=_PDF_BLUE, leading=13),
        'TableCell':    dict(fontSize=_INV_BODY, fontName=_PDF_FONT, textColor=_PDF_DARK, leading=_INV_BODY_LEAD),
        'TableCellBold':dict(fontSize=_INV_BODY, fontName=_PDF_FONT_BOLD, textColor=_PDF_DARK, leading=_INV_BODY_LEAD),
        'TableHeader':  dict(fontSize=_INV_SMALL, fontName=_PDF_FONT_BOLD, textColor=_PDF_WHITE, leading=_INV_SMALL_LEAD),
        'FinePrint':    dict(fontSize=_INV_TINY, fontName=_PDF_FONT, textColor=_PDF_MUTED, leading=_INV_TINY_LEAD),
    }
    for name, kw in styles_kw.items():
        s.add(ParagraphStyle(name, parent=s['Normal'], **kw))
    return s


class _InvCanvas(canvas.Canvas):
    """A4 canvas that draws the invoice footer (with Page X of Y) on save."""

    def __init__(self, *a, **k):
        canvas.Canvas.__init__(self, *a, **k)
        self._saved = []
        self._invoice_footer = None

    def showPage(self):
        self._saved.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        footer = self._invoice_footer
        if len(self._code):
            self._saved.append(dict(self.__dict__))
        n = len(self._saved)
        for pst in self._saved:
            count = getattr(self, '_annotationCount', 0)
            self.__dict__.update(pst)
            self._annotationCount = count
            if footer:
                footer(self, self._pageNumber, n)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)


def _pdf_watermark(c, text, pw, ph):
    c.saveState()
    c.setFillColor(_PDF_WATERMARK)
    c.setFont(_PDF_FONT_BOLD, 40)
    c.translate(pw / 2, ph / 2)
    c.rotate(45)
    c.drawCentredString(0, 0, text)
    c.restoreState()


def _pdf_footer(c, co, pg, total, pw):
    c.saveState()
    c.setStrokeColor(_PDF_BORDER)
    c.setLineWidth(0.4)
    c.line(_PDF_M, _PDF_BOTTOM + 4, pw - _PDF_M, _PDF_BOTTOM + 4)
    name = (co.get('name') or 'GV Powers').upper()
    line = f"{name} | GSTIN: {co.get('gstin', '')} | {co.get('website', '')}"
    c.setFont(_PDF_FONT_BOLD, 7)
    c.setFillColor(_PDF_MUTED)
    c.drawCentredString(pw / 2, _PDF_BOTTOM - 6, line)
    c.setFont(_PDF_FONT, _INV_TINY)
    c.drawCentredString(pw / 2, _PDF_BOTTOM - 17, f"This is a computer-generated invoice. | Page {pg} of {total}")
    c.restoreState()


def _pdf_header(c, co, label, inv, pw):
    """Professional three-zone header: LEFT logo+details | CENTER tagline | RIGHT invoice meta."""
    xs = _PDF_M
    rx = pw - _PDF_M
    yt = A4[1] - _PDF_TOP

    # --- LEFT ZONE: Logo + Company Info ---
    logo_path = os.path.join(BASE_DIR, "static", "img", "logo", "img.png")
    logo_sz = 58
    logo_y = yt - logo_sz
    if os.path.exists(logo_path):
        try:
            c.drawImage(logo_path, xs, logo_y, width=logo_sz, height=logo_sz,
                        preserveAspectRatio=True, mask='auto')
        except Exception:
            logo_path = None
    if not logo_path or not os.path.exists(logo_path):
        c.setFillColor(_PDF_NAVY)
        c.roundRect(xs, logo_y, logo_sz, logo_sz, 6, fill=1, stroke=0)
        c.setFillColor(_PDF_WHITE)
        c.setFont(_PDF_FONT_BOLD, 20)
        c.drawCentredString(xs + logo_sz / 2, logo_y + logo_sz / 2 - 6, "GV")

    lx = xs + logo_sz + 12
    ty = yt - 1
    c.setFont(_PDF_FONT_BOLD, 16)
    c.setFillColor(_PDF_NAVY)
    c.drawString(lx, ty, co.get("name", "GV Powers"))
    ty -= 14
    c.setFont(_PDF_FONT_OBLIQUE, 8.5)
    c.setFillColor(_PDF_BLUE)
    c.drawString(lx, ty, co.get("tagline", ""))
    ty -= 12
    c.setFont(_PDF_FONT, _INV_SMALL)
    c.setFillColor(_PDF_MUTED)
    c.drawString(lx, ty, co.get("services", ""))
    ty -= 12
    addr_lines = _pdf_wrap(co.get("address", ""), 52)[:2]
    for al in addr_lines:
        c.setFont(_PDF_FONT, _INV_SMALL)
        c.setFillColor(_PDF_MUTED)
        c.drawString(lx, ty, al)
        ty -= 11
    c.setFont(_PDF_FONT, _INV_SMALL)
    c.setFillColor(_PDF_MUTED)
    c.drawString(lx, ty, f"Phone: {co.get('phone', '')}  |  Mobile: {co.get('mobile', '')}  |  Email: {co.get('email', '')}")
    ty -= 11
    c.setFont(_PDF_FONT_BOLD, _INV_SMALL)
    c.setFillColor(_PDF_DARK)
    c.drawString(lx, ty, f"GSTIN: {co.get('gstin', '')}  |  Website: {co.get('website', '')}")

    # --- RIGHT ZONE: TAX INVOICE + Copy + Meta ---
    c.setFont(_PDF_FONT_BOLD, 20)
    c.setFillColor(_PDF_NAVY)
    c.drawRightString(rx, yt - 1, "TAX INVOICE")
    c.setFont(_PDF_FONT_BOLD, 8.5)
    c.setFillColor(_PDF_BLUE)
    c.drawRightString(rx, yt - 15, label)
    meta_y = yt - 30
    c.setFont(_PDF_FONT, _INV_BODY)
    c.setFillColor(_PDF_DARK)
    c.drawRightString(rx, meta_y, f"Invoice #:  {inv.invoice_number}")
    meta_y -= 13
    c.drawRightString(rx, meta_y, f"Date:  {_pdf_date(inv.invoice_date)}")
    if inv.due_date:
        meta_y -= 13
        c.drawRightString(rx, meta_y, f"Due Date:  {_pdf_date(inv.due_date)}")

    # --- Separator lines ---
    sep_y = yt - 82
    c.setStrokeColor(_PDF_NAVY)
    c.setLineWidth(1.0)
    c.line(xs, sep_y, rx, sep_y)
    c.setStrokeColor(_PDF_BORDER)
    c.setLineWidth(0.4)
    c.line(xs, sep_y - 3, rx, sep_y - 3)

    return sep_y - 10


def _pdf_cust_info(c, inv, y, pw):
    """Two-column customer info: BILL TO (left) + INVOICE DETAILS (right)."""
    xs = _PDF_M
    avail = pw - 2 * _PDF_M
    gap = 14
    bw = (avail - gap) / 2
    bh = 88

    csc = inv.customer_state_code
    state_str = (inv.customer_state or 'N/A') + (f" ({csc})" if csc else "")

    # --- BILL TO (left) ---
    bx = xs
    c.setStrokeColor(_PDF_BORDER)
    c.setLineWidth(0.4)
    c.setFillColor(_PDF_LIGHT)
    c.roundRect(bx, y - bh, bw, bh, 3, fill=1, stroke=1)
    c.setFont(_PDF_FONT_BOLD, 8.5)
    c.setFillColor(_PDF_BLUE)
    c.drawString(bx + 10, y - 13, "BILL TO")
    dy = y - 26
    c.setFont(_PDF_FONT_BOLD, 11)
    c.setFillColor(_PDF_DARK)
    c.drawString(bx + 10, dy, (inv.customer_name or 'N/A')[:42])
    dy -= 14
    c.setFont(_PDF_FONT, _INV_BODY)
    c.setFillColor(_PDF_MUTED)
    for ln in _pdf_wrap(inv.customer_address or '', 42)[:2]:
        c.drawString(bx + 10, dy, ln)
        dy -= 11
    if inv.customer_mobile:
        c.drawString(bx + 10, dy, f"Mobile: {inv.customer_mobile}")
        dy -= 11
    if inv.customer_email:
        c.drawString(bx + 10, dy, f"Email: {inv.customer_email}")
        dy -= 11
    if inv.customer_gstin:
        c.drawString(bx + 10, dy, f"GSTIN: {inv.customer_gstin}")
        dy -= 11
    c.drawString(bx + 10, dy, f"State: {state_str}")

    # --- INVOICE DETAILS (right) ---
    ix = xs + bw + gap
    c.setStrokeColor(_PDF_BORDER)
    c.setLineWidth(0.4)
    c.setFillColor(_PDF_LIGHT)
    c.roundRect(ix, y - bh, bw, bh, 3, fill=1, stroke=1)
    c.setFont(_PDF_FONT_BOLD, 8.5)
    c.setFillColor(_PDF_BLUE)
    c.drawString(ix + 10, y - 13, "INVOICE DETAILS")
    rows = [('Invoice #', inv.invoice_number),
            ('Date', _pdf_date(inv.invoice_date))]
    if inv.due_date:
        rows.append(('Due Date', _pdf_date(inv.due_date)))
    rows.append(('Place of Supply', state_str))
    dy = y - 28
    for lbl, val in rows:
        c.setFont(_PDF_FONT, _INV_BODY)
        c.setFillColor(_PDF_MUTED)
        c.drawString(ix + 10, dy, lbl)
        c.setFont(_PDF_FONT_BOLD, _INV_BODY)
        c.setFillColor(_PDF_DARK)
        c.drawRightString(ix + bw - 10, dy, str(val)[:36])
        dy -= 14
    return y - bh - 10


def _pdf_product_table(c, inv, styles, y, pw, new_page):
    """13-column product table with navy header, alternating rows, proper splits."""
    xs = _PDF_M
    avail = pw - 2 * _PDF_M

    # Column widths: IDX PROD HSN QTY UNIT RATE DISC GST TAX CGST SGST IGST AMT
    cw_raw = [14, 80, 32, 22, 26, 52, 26, 24, 52, 44, 44, 44, 56]
    scale = avail / sum(cw_raw)
    cw = [w * scale for w in cw_raw]

    hdrs = ["#", "PRODUCT", "HSN", "QTY", "UNIT", "RATE", "DISC.", "GST", "TAXABLE",
            "CGST", "SGST", "IGST", "AMOUNT"]
    data = [[Paragraph(h, styles['TableHeader']) for h in hdrs]]

    for idx, it in enumerate(inv.items, 1):
        data.append([
            Paragraph(str(idx), styles['TableCell']),
            Paragraph(str(it.product_name)[:50], styles['TableCell']),
            Paragraph(str(it.hsn or "-"), styles['TableCell']),
            Paragraph(str(it.qty), styles['TableCell']),
            Paragraph(str(it.unit or ''), styles['TableCell']),
            Paragraph(_pdf_inr(it.price), styles['TableCell']),
            Paragraph(f"{it.discount}%", styles['TableCell']),
            Paragraph(f"{it.gst_rate}%", styles['TableCell']),
            Paragraph(_pdf_inr(it.taxable_value), styles['TableCell']),
            Paragraph(_pdf_inr(it.cgst), styles['TableCell']),
            Paragraph(_pdf_inr(it.sgst), styles['TableCell']),
            Paragraph(_pdf_inr(it.igst), styles['TableCell']),
            Paragraph(f"<b>{_pdf_inr(it.total)}</b>", styles['TableCellBold']),
        ])

    t = Table(data, colWidths=cw, repeatRows=1)

    cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), _PDF_NAVY),
        ('TEXTCOLOR',  (0, 0), (-1, 0), _PDF_WHITE),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
        ('GRID',       (0, 0), (-1, -1), 0.3, _PDF_BORDER),
        ('LINEBELOW',  (0, 0), (-1, 0), 0.8, _PDF_NAVY),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (4, -1), 'CENTER'),
        ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
    ]
    for i in range(1, len(data)):
        if i % 2 == 0:
            cmds.append(('BACKGROUND', (0, i), (-1, i), _PDF_ROW_ALT))
    t.setStyle(TableStyle(cmds))

    # Paginate
    remain = t
    while remain is not None:
        if y - (_PDF_BOTTOM + 8) > 0:
            parts = remain.split(avail, y - (_PDF_BOTTOM + 8))
        else:
            parts = []
        head = parts[0] if parts else None
        tail = parts[1] if len(parts) > 1 else None
        if head is None:
            y = new_page()
            continue
        tw, th = head.wrap(avail, 0)
        if y - th < _PDF_BOTTOM + 8:
            y = new_page()
            remain = head
            continue
        head.drawOn(c, xs, y - th)
        y -= th
        remain = tail
    return y


def _pdf_wrap(text, width_chars):
    words = text.split()
    out, line = [], ""
    for w in words:
        if len(line) + len(w) + 1 <= width_chars:
            line = (line + " " + w).strip()
        else:
            if line:
                out.append(line)
            line = w
    if line:
        out.append(line)
    return out or [""]


def _pdf_gst_table(c, inv, styles, y, pw):
    """Compact HSN / GST SUMMARY table."""
    xs = _PDF_M
    avail = pw - 2 * _PDF_M

    hdrs = ["HSN Code", "Taxable Value", "GST Rate", "CGST", "SGST", "IGST", "Total Tax"]
    data = [[Paragraph(h, styles['TableHeader']) for h in hdrs]]

    hsn = {}
    for it in inv.items:
        h = it.hsn or "N/A"
        if h not in hsn:
            hsn[h] = {"taxable": 0, "rate": it.gst_rate, "cgst": 0, "sgst": 0, "igst": 0}
        hsn[h]["taxable"] += it.taxable_value
        hsn[h]["cgst"] += it.cgst
        hsn[h]["sgst"] += it.sgst
        hsn[h]["igst"] += it.igst

    for h, d in hsn.items():
        data.append([
            Paragraph(h, styles['TableCell']),
            Paragraph(_pdf_inr(d['taxable']), styles['TableCell']),
            Paragraph(f"{d['rate']:.0f}%", styles['TableCell']),
            Paragraph(_pdf_inr(d['cgst']), styles['TableCell']),
            Paragraph(_pdf_inr(d['sgst']), styles['TableCell']),
            Paragraph(_pdf_inr(d['igst']), styles['TableCell']),
            Paragraph(f"<b>{_pdf_inr(d['cgst'] + d['sgst'] + d['igst'])}</b>", styles['TableCellBold']),
        ])

    tt = Decimal(str(inv.total_cgst)) + Decimal(str(inv.total_sgst)) + Decimal(str(inv.total_igst))
    data.append([
        Paragraph("<b>Total</b>", styles['TableCellBold']),
        Paragraph(f"<b>{_pdf_inr(inv.total_taxable)}</b>", styles['TableCellBold']),
        Paragraph("", styles['TableCell']),
        Paragraph(f"<b>{_pdf_inr(inv.total_cgst)}</b>", styles['TableCellBold']),
        Paragraph(f"<b>{_pdf_inr(inv.total_sgst)}</b>", styles['TableCellBold']),
        Paragraph(f"<b>{_pdf_inr(inv.total_igst)}</b>", styles['TableCellBold']),
        Paragraph(f"<b>{_pdf_inr(tt)}</b>", styles['TableCellBold']),
    ])

    cw = [avail * 0.16, avail * 0.16, avail * 0.12, avail * 0.14, avail * 0.14, avail * 0.14, avail * 0.14]
    t = Table(data, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), _PDF_LIGHT),
        ('TEXTCOLOR',  (0, 0), (-1, 0), _PDF_NAVY),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('TOPPADDING',    (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING',   (0, 0), (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
        ('GRID',       (0, 0), (-1, -1), 0.3, _PDF_BORDER),
        ('LINEBELOW',  (0, -1), (-1, -1), 0.6, _PDF_BLUE),
        ('BACKGROUND', (0, -1), (-1, -1), _PDF_LIGHT),
    ]))

    c.setFont(_PDF_FONT_BOLD, 9)
    c.setFillColor(_PDF_BLUE)
    c.drawString(xs, y - 11, "HSN / GST SUMMARY")
    y -= 16
    tw, th = t.wrap(avail, 0)
    t.drawOn(c, xs, y - th)
    return y - th - 8


def _pdf_terms(c, inv, co, styles, y, pw):
    xs = _PDF_M
    terms = (inv.terms or '').strip()
    if not terms:
        try:
            row = Settings.query.filter_by(key='invoice_terms').first()
            terms = (row.value or '').strip() if row else ''
        except Exception:
            terms = ''
    if not terms:
        terms = ("Payment within 15 days.  Goods once sold will not be returned or exchanged.  "
                 "E&OE (Errors and Omissions Excepted).")
    c.setFont(_PDF_FONT_BOLD, 9)
    c.setFillColor(_PDF_BLUE)
    c.drawString(xs, y - 10, "TERMS & CONDITIONS")
    c.setFont(_PDF_FONT, _INV_BODY)
    c.setFillColor(_PDF_MUTED)
    ty = y - 21
    for ln in _pdf_wrap(terms, 118)[:4]:
        c.drawString(xs, ty, ln)
        ty -= 11
    return ty - 4


def _pdf_summary(c, inv, styles, y, pw):
    """Right-aligned totals summary with grand total and amount in words."""
    xs = pw - _PDF_M - 250
    bw = 250

    lines = [("Subtotal", _pdf_inr(inv.subtotal))]
    if inv.total_discount > 0:
        lines.append(("Discount", "-" + _pdf_inr(inv.total_discount)))
    lines.append(("Taxable Amount", _pdf_inr(inv.total_taxable)))
    if inv.is_intra_state:
        lines.extend([("CGST", _pdf_inr(inv.total_cgst)), ("SGST", _pdf_inr(inv.total_sgst))])
    else:
        lines.append(("IGST", _pdf_inr(inv.total_igst)))
    if inv.round_off != 0:
        ro = Decimal(str(inv.round_off))
        ro_txt = _pdf_inr(abs(ro)) if ro > 0 else "-" + _pdf_inr(abs(ro))
        lines.append(("Round Off", ro_txt))

    aw = amount_to_words(inv.grand_total)
    aw_lines = _pdf_wrap(aw, 34)
    th = 48 + len(lines) * 14 + len(aw_lines) * 11

    c.setStrokeColor(_PDF_BORDER)
    c.setLineWidth(0.4)
    c.setFillColor(_PDF_LIGHT)
    c.roundRect(xs, y - th, bw, th, 3, fill=1, stroke=1)

    ty = y - 13
    for lbl, val in lines:
        c.setFont(_PDF_FONT, _INV_BODY)
        c.setFillColor(_PDF_MUTED)
        c.drawString(xs + 10, ty, lbl)
        c.setFont(_PDF_FONT, _INV_BODY)
        c.setFillColor(_PDF_DARK)
        c.drawRightString(xs + bw - 10, ty, str(val))
        ty -= 14

    c.setStrokeColor(_PDF_BLUE)
    c.setLineWidth(0.8)
    c.line(xs + 10, ty + 2, xs + bw - 10, ty + 2)
    ty -= 7

    c.setFont(_PDF_FONT_BOLD, 12)
    c.setFillColor(_PDF_NAVY)
    c.drawString(xs + 10, ty, "GRAND TOTAL")
    c.setFont(_PDF_FONT_BOLD, 13)
    c.setFillColor(_PDF_NAVY)
    c.drawRightString(xs + bw - 10, ty, _pdf_inr(inv.grand_total))
    ty -= 17

    c.setFont(_PDF_FONT_OBLIQUE, 7.5)
    c.setFillColor(_PDF_MUTED)
    c.drawString(xs + 10, ty, "Amount in Words")
    ty -= 10
    c.setFont(_PDF_FONT_BOLD, _INV_BODY)
    c.setFillColor(_PDF_DARK)
    for seg in aw_lines:
        c.drawString(xs + 10, ty, seg)
        ty -= 11
    return y - th - 10


def _pdf_payment(c, inv, styles, y, pw):
    """Left-aligned payment information box."""
    xs = _PDF_M
    avail = pw - 2 * _PDF_M
    totals_w = 250
    gap = 14
    bw = avail - totals_w - gap
    bh = 88

    bal = inv.balance_due if inv.balance_due is not None else (
        Decimal(str(inv.grand_total)) - Decimal(str(inv.amount_paid)))

    c.setStrokeColor(_PDF_BORDER)
    c.setLineWidth(0.4)
    c.setFillColor(_PDF_LIGHT)
    c.roundRect(xs, y - bh, bw, bh, 3, fill=1, stroke=1)

    c.setFont(_PDF_FONT_BOLD, 8.5)
    c.setFillColor(_PDF_BLUE)
    c.drawString(xs + 10, y - 13, "PAYMENT INFORMATION")

    rows = [
        ("Payment Method", _payment_method_label(inv.payment_method) if inv.payment_method else 'N/A'),
        ("Amount Paid", _pdf_inr(inv.amount_paid)),
        ("Balance Due", _pdf_inr(bal)),
    ]
    dy = y - 28
    for lbl, val in rows:
        c.setFont(_PDF_FONT, _INV_BODY)
        c.setFillColor(_PDF_MUTED)
        c.drawString(xs + 10, dy, lbl)
        if lbl == "Balance Due":
            col = _PDF_GREEN if not bal else _PDF_RED
        else:
            col = _PDF_DARK
        c.setFont(_PDF_FONT_BOLD, _INV_BODY)
        c.setFillColor(col)
        c.drawRightString(xs + bw - 10, dy, str(val))
        dy -= 17
    return y - bh - 10


def _pdf_decl(c, inv, styles, y, pw, co=None):
    """Declaration + compact two-column signature area."""
    co = co or {}
    xs = _PDF_M
    avail = pw - 2 * _PDF_M

    # Declaration
    c.setFont(_PDF_FONT_BOLD, 9)
    c.setFillColor(_PDF_BLUE)
    c.drawString(xs, y - 10, "DECLARATION")
    c.setFont(_PDF_FONT, _INV_BODY)
    c.setFillColor(_PDF_MUTED)
    for i, line in enumerate(["This is a computer-generated invoice.",
                              "Subject to local jurisdiction.",
                              "E&OE (Errors and Omissions Excepted)."]):
        c.drawString(xs, y - 21 - i * 11, line)

    # Signature area
    sy = y - 55
    sig_h = 60
    cw2 = (avail - 24) / 2

    # Left: Authorized Signature
    lx = xs + 8
    sig_line_y = sy - sig_h + 16
    c.setStrokeColor(_PDF_MUTED)
    c.setLineWidth(0.4)
    c.line(lx, sig_line_y, lx + cw2, sig_line_y)
    c.setFont(_PDF_FONT, _INV_SMALL)
    c.setFillColor(_PDF_MUTED)
    c.drawString(lx, sig_line_y - 10, "Authorized Signature")
    c.setFont(_PDF_FONT_BOLD, _INV_BODY)
    c.setFillColor(_PDF_NAVY)
    c.drawString(lx, sig_line_y - 22, co.get("name", "GV Powers"))
    c.setFont(_PDF_FONT, _INV_SMALL)
    c.setFillColor(_PDF_MUTED)
    c.drawString(lx, sig_line_y - 34, "Date: _______________")

    # Right: Customer Signature
    sx2 = xs + cw2 + 24
    c.setStrokeColor(_PDF_MUTED)
    c.setLineWidth(0.4)
    c.line(sx2, sig_line_y, sx2 + cw2, sig_line_y)
    c.setFont(_PDF_FONT, _INV_SMALL)
    c.setFillColor(_PDF_MUTED)
    c.drawString(sx2, sig_line_y - 10, "Customer Signature")
    c.setFont(_PDF_FONT, _INV_SMALL)
    c.setFillColor(_PDF_MUTED)
    c.drawString(sx2, sig_line_y - 34, "Date: _______________")

    return sy - sig_h - 6


def _pdf_render(inv, co, label):
    buf = io.BytesIO()
    pw, ph = A4
    styles = _pdf_styles()
    c = _InvCanvas(buf, pagesize=A4)
    c._invoice_footer = lambda cv, pg, total: _pdf_footer(cv, co, pg, total, pw)
    y = 0

    def new_page(first=False):
        nonlocal y
        if not first:
            c.showPage()
        _pdf_watermark(c, label, pw, ph)
        if inv.status == 'cancelled':
            _pdf_watermark(c, 'CANCELLED', pw, ph)
        y = _pdf_header(c, co, label, inv, pw)
        return y

    y = new_page(first=True)
    y = _pdf_cust_info(c, inv, y, pw)
    y = _pdf_product_table(c, inv, styles, y, pw, new_page)
    if y - 120 < _PDF_BOTTOM + 8:
        y = new_page()
    y = _pdf_gst_table(c, inv, styles, y, pw)
    if y - 140 < _PDF_BOTTOM + 8:
        y = new_page()
    y_pay = _pdf_payment(c, inv, styles, y, pw)
    y_sum = _pdf_summary(c, inv, styles, y, pw)
    y = min(y_pay, y_sum)
    if y - 70 < _PDF_BOTTOM + 8:
        y = new_page()
    y = _pdf_terms(c, inv, co, styles, y, pw)
    if y - 160 < _PDF_BOTTOM + 8:
        y = new_page()
    y = _pdf_decl(c, inv, styles, y, pw, co)
    c.save()
    buf.seek(0)
    return buf


_COPY_LABELS = {"owner": "ADMIN / OWNER COPY", "customer": "CUSTOMER COPY", "gst": "GST TAX COPY"}


def generate_owner_copy(inv, co=None):
    return _pdf_render(inv, co or {}, "ADMIN / OWNER COPY")


def generate_customer_copy(inv, co=None):
    return _pdf_render(inv, co or {}, "CUSTOMER COPY")


def generate_gst_copy(inv, co=None):
    return _pdf_render(inv, co or {}, "GST TAX COPY")


def generate_invoice_pdf(inv, copy_type="customer", co=None):
    return _pdf_render(inv, co or {}, _COPY_LABELS.get(copy_type, "CUSTOMER COPY"))


############################################################
# EMAIL SERVICE
############################################################


def build_invoice_email(inv, co, pdf_buf, recipient=None):
    co = co or {}
    company_name = co.get('name', 'GV Powers')
    company_email = co.get('email', '')
    company_phone = co.get('phone', '')
    company_mobile = co.get('mobile', '')
    company_address = co.get('address', '')

    msg = MIMEMultipart('related')
    msg['Subject'] = f"Invoice {inv.invoice_number} from {company_name}"
    msg['To'] = recipient or inv.customer_email or ''

    logo_path = os.path.join(BASE_DIR, "static", "img", "logo", "img.png")

    items_html = ''
    for item in inv.items:
        items_html += f'''
            <tr>
                <td style="padding:8px 10px;border-bottom:1px solid #eee;font-size:13px;">{item.product_name}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #eee;font-size:13px;text-align:center;">{item.qty}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #eee;font-size:13px;text-align:right;">Rs. {item.price:,.2f}</td>
                <td style="padding:8px 10px;border-bottom:1px solid #eee;font-size:13px;text-align:right;">Rs. {item.total:,.2f}</td>
            </tr>'''

    payment_status_badge = {
        'paid': '<span style="background:#d4edda;color:#155724;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;">✓ Paid</span>',
        'partial': '<span style="background:#fff3cd;color:#856404;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;">Partially Paid</span>',
        'pending': '<span style="background:#f8d7da;color:#721c24;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;">Due</span>',
        'due': '<span style="background:#f8d7da;color:#721c24;padding:4px 12px;border-radius:20px;font-size:12px;font-weight:600;">Due</span>',
    }.get(inv.payment_status or 'due', '')

    html = f'''<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#f4f6f9;font-family:Arial,Helvetica,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="max-width:600px;margin:0 auto;">
<tr><td style="padding:30px 0 10px 0;text-align:center;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
        <tr><td style="background:#1a3a5c;padding:30px 40px;text-align:center;">
            <img src="cid:logo" alt="{company_name}" style="width:60px;height:60px;border-radius:8px;margin-bottom:10px;">
            <h1 style="color:#ffffff;font-size:22px;margin:8px 0 4px 0;font-weight:700;">{company_name}</h1>
            <p style="color:#8ba4c4;font-size:13px;margin:0;">{co.get('tagline', '')}</p>
        </td></tr>
        <tr><td style="padding:30px 40px 10px 40px;">
            <p style="font-size:15px;color:#333;margin:0 0 4px 0;">Dear <strong>{inv.customer_name}</strong>,</p>
            <p style="font-size:14px;color:#666;margin:0 0 20px 0;line-height:1.6;">Thank you for choosing GV Powers & Energy Solutions. Your invoice is attached below. We truly appreciate your business and look forward to serving you again!</p>
        </td></tr>
        <tr><td style="padding:0 40px;">
            <table width="100%" cellpadding="0" cellspacing="0" style="background:#f8fafc;border-radius:8px;padding:16px 20px;margin-bottom:20px;">
                <tr>
                    <td style="font-size:12px;color:#888;text-transform:uppercase;letter-spacing:1px;">Invoice No</td>
                    <td style="font-size:12px;color:#888;text-transform:uppercase;letter-spacing:1px;">Date</td>
                    <td style="font-size:12px;color:#888;text-transform:uppercase;letter-spacing:1px;">Status</td>
                </tr>
                <tr>
                    <td style="font-size:16px;font-weight:700;color:#1a3a5c;padding-top:4px;">{inv.invoice_number}</td>
                    <td style="font-size:16px;font-weight:600;color:#333;padding-top:4px;">{inv.invoice_date.strftime('%d-%b-%Y') if inv.invoice_date else ''}</td>
                    <td style="padding-top:4px;">{payment_status_badge}</td>
                </tr>
            </table>
        </td></tr>
        <tr><td style="padding:0 40px;">
            <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:20px;">
                <tr><td style="font-size:13px;color:#888;padding-bottom:4px;">Due Date</td>
                <td style="font-size:13px;color:#333;font-weight:600;text-align:right;padding-bottom:4px;">{inv.due_date.strftime('%d-%b-%Y') if inv.due_date else 'N/A'}</td></tr>
                <tr><td style="font-size:13px;color:#888;padding-bottom:4px;">Amount Paid</td>
                <td style="font-size:13px;color:#333;text-align:right;padding-bottom:4px;">Rs. {inv.amount_paid:,.2f}</td></tr>
                <tr><td style="font-size:13px;color:#888;">Balance Due</td>
                <td style="font-size:13px;color:#d32f2f;font-weight:700;text-align:right;">Rs. {(inv.grand_total - inv.amount_paid):,.2f}</td></tr>
            </table>
        </td></tr>
        <tr><td style="padding:0 40px;">
            <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;margin-bottom:20px;">
                <thead>
                    <tr style="background:#1a3a5c;">
                        <th style="padding:10px;color:#fff;font-size:12px;text-align:left;">Item</th>
                        <th style="padding:10px;color:#fff;font-size:12px;text-align:center;">Qty</th>
                        <th style="padding:10px;color:#fff;font-size:12px;text-align:right;">Rate</th>
                        <th style="padding:10px;color:#fff;font-size:12px;text-align:right;">Amount</th>
                    </tr>
                </thead>
                <tbody>{items_html}</tbody>
            </table>
        </td></tr>
        <tr><td style="padding:0 40px 20px 40px;">
            <table width="100%" cellpadding="0" cellspacing="0" style="border-top:2px solid #1a3a5c;padding-top:12px;">
                <tr><td style="font-size:15px;color:#333;font-weight:700;">Total Amount</td>
                <td style="font-size:20px;color:#1a3a5c;font-weight:800;text-align:right;">Rs. {inv.grand_total:,.2f}</td></tr>
            </table>
        </td></tr>
        <tr><td style="padding:0 40px 8px 40px;">
            <p style="font-size:13px;color:#666;margin:0;">Amount in words: <strong style="color:#1a3a5c;">{amount_to_words(inv.grand_total)}</strong></p>
        </td></tr>
        <tr><td style="padding:0 40px 30px 40px;border-bottom:1px solid #eee;">
            <p style="font-size:12px;color:#888;font-style:italic;margin:0;">Invoice attached as PDF. Please retain for your records.</p>
        </td></tr>
        <tr><td style="background:#f8fafc;padding:20px 40px 24px 40px;">
            <table width="100%" cellpadding="0" cellspacing="0">
                <tr>
                    <td style="font-size:12px;color:#888;padding-bottom:4px;">📞 {company_phone} {(' | ' + company_mobile) if company_mobile else ''}</td>
                    <td style="font-size:12px;color:#888;text-align:right;padding-bottom:4px;">✉️ {company_email}</td>
                </tr>
                <tr><td colspan="2" style="font-size:11px;color:#aaa;padding-top:4px;">{company_address}</td></tr>
                <tr><td colspan="2" style="font-size:11px;color:#aaa;padding-top:12px;border-top:1px solid #ddd;margin-top:8px;padding-top:8px;">GSTIN: {co.get('gstin', 'N/A')} | Website: {co.get('website', '')}</td></tr>
            </table>
        </td></tr>
    </table>
    <p style="font-size:11px;color:#aaa;margin-top:16px;">This is an auto-generated email from {company_name}. Please do not reply.</p>
</td></tr></table>
</body></html>'''

    html_part = MIMEText(html, 'html')
    msg.attach(html_part)

    if os.path.exists(logo_path):
        try:
            with open(logo_path, 'rb') as f:
                logo_img = MIMEImage(f.read())
                logo_img.add_header('Content-ID', '<logo>')
                logo_img.add_header('Content-Disposition', 'inline', filename='logo.png')
                msg.attach(logo_img)
        except Exception:
            pass

    pdf_buf.seek(0)
    pdf_att = MIMEBase('application', 'pdf')
    pdf_att.set_payload(pdf_buf.read())
    email_encoders.encode_base64(pdf_att)
    pdf_att.add_header('Content-Disposition', 'attachment', filename=f"{inv.invoice_number}.pdf")
    msg.attach(pdf_att)

    return msg


def send_invoice_email(inv, co=None, recipient=None):
    co = co or {}
    # SMTP is configured ONLY via environment variables (see .env). Never read from
    # the database or the Admin Panel — these credentials are backend-only.
    smtp_server = os.getenv('MAIL_SERVER', '').strip() or os.getenv('SMTP_SERVER', '').strip()
    smtp_port = int(os.getenv('MAIL_PORT', '') or os.getenv('SMTP_PORT', '465'))
    smtp_email = os.getenv('MAIL_USERNAME', '').strip() or os.getenv('SMTP_EMAIL', '').strip()
    smtp_password = os.getenv('MAIL_PASSWORD', '').strip() or os.getenv('SMTP_PASSWORD', '').strip()

    if not smtp_password or not smtp_email:
        return False, "Outbound email is not configured. Add MAIL_PASSWORD/MAIL_USERNAME to .env."

    recipient = recipient or inv.customer_email
    if not recipient:
        return False, "No customer email address on this invoice."

    pdf_buf = generate_invoice_pdf(inv, 'customer', co)
    msg = build_invoice_email(inv, co, pdf_buf, recipient=recipient)
    msg['From'] = f"{co.get('name', 'GV Powers')} <{smtp_email}>"

    try:
        use_ssl = os.getenv('MAIL_USE_SSL', 'true').lower() in ('1', 'true', 'yes', 'on')
        use_tls = os.getenv('MAIL_USE_TLS', '').lower() in ('1', 'true', 'yes', 'on')
        if use_ssl:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        else:
            server = smtplib.SMTP(smtp_server, smtp_port)
            if use_tls:
                server.starttls()
        with server:
            server.login(smtp_email, smtp_password)
            server.sendmail(smtp_email, [recipient], msg.as_string())
        return True, f"Invoice {inv.invoice_number} sent to {recipient}"
    except smtplib.SMTPAuthenticationError:
        return False, "SMTP authentication failed. Check MAIL_USERNAME/MAIL_PASSWORD in .env."
    except smtplib.SMTPException as e:
        return False, f"SMTP error: {str(e)}"
    except Exception as e:
        return False, f"Failed to send: {str(e)}"


############################################################
# LOGIN MANAGER
############################################################


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated: flash('Please log in.', 'warning'); return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated: flash('Please log in.', 'warning'); return redirect(url_for('login'))
        if not current_user.is_admin: flash('Admin access required.', 'danger'); return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


# Endpoint function names a non-admin (sales) user is allowed to reach.
# Everything else is treated as admin-only for non-admin roles.
SALES_ENDPOINTS = {
    'index', 'login', 'logout', 'profile', 'dashboard',
    'invoice_history', 'new_invoice', 'view_invoice', 'invoice_preview',
    'download_invoice_pdf', 'download_all_pdfs', 'print_invoice_pdf',
    'email_invoice', 'edit_invoice', 'cancel_invoice', 'global_search',
    'search_customers',
    'api_product_search', 'create_customer', 'get_notifications',
    'mark_notification_read', 'not_found', 'forbidden', 'server_error',
    'add_payment',
}


def sales_role_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated: flash('Please log in.', 'warning'); return redirect(url_for('login'))
        if current_user.role != 'sales': admin_required(f)
        return f(*args, **kwargs)
    return decorated


@app.before_request
def enforce_role_permissions():
    if not current_user.is_authenticated:
        return
    if current_user.is_admin:
        return
    ep = request.endpoint or ''
    if ep in SALES_ENDPOINTS or ep.startswith('static'):
        return
    flash('You do not have permission to access this page.', 'danger')
    return redirect(url_for('dashboard'))


############################################################
# ADMIN ROUTES
############################################################

@app.route('/')
def index():
    if not current_user.is_authenticated: return redirect(url_for('login'))
    return redirect(url_for('dashboard'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username', '').strip()).first()
        if user and user.check_password(request.form.get('password', '').strip()) and user.is_active:
            login_user(user); session.permanent = True
            user.last_login = datetime.utcnow(); db.session.commit()
            log_audit(user.id, 'login', details=f'User {user.username} logged in', ip_address=request.remote_addr)
            flash(f'Welcome back, {user.full_name}!', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid credentials or account disabled.', 'danger')
    return render_template('auth/login.html')


@app.route('/logout', methods=['POST'])
def logout():
    uid = current_user.id if current_user.is_authenticated else None
    if uid: log_audit(uid, 'logout', ip_address=request.remote_addr)
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ----- Dashboard & Profile -----


@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'sales':
        return _sales_dashboard()
    _run_low_stock_check()
    today = date.today()

    rng = request.args.get('range', 'month')
    if rng == '7d':
        from_dt, to_dt, range_label = today - timedelta(days=6), today, 'Last 7 Days'
    elif rng == '30d':
        from_dt, to_dt, range_label = today - timedelta(days=29), today, 'Last 30 Days'
    elif rng == 'month':
        from_dt, to_dt, range_label = today.replace(day=1), today, 'This Month'
    else:
        fy_start = date(today.year - 1, 4, 1) if today.month < 4 else date(today.year, 4, 1)
        from_dt, to_dt, range_label = fy_start, today, 'Financial Year'

    range_sales = db.session.query(func.coalesce(func.sum(Invoice.grand_total), 0)).filter(
        Invoice.invoice_date >= from_dt, Invoice.invoice_date <= to_dt,
        Invoice.status != 'cancelled').scalar() or 0

    range_collections = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).join(
        Invoice, Invoice.id == Payment.invoice_id
    ).filter(Payment.payment_date >= from_dt, Payment.payment_date <= to_dt,
             Invoice.status != 'cancelled').scalar() or 0

    range_invoices_count = Invoice.query.filter(
        Invoice.invoice_date >= from_dt, Invoice.invoice_date <= to_dt,
        Invoice.status != 'cancelled').count()

    opening_balance = 0.0
    if from_dt > date.min:
        before_inv = db.session.query(func.coalesce(func.sum(Invoice.grand_total), 0)).filter(
            Invoice.invoice_date < from_dt, Invoice.status != 'cancelled').scalar() or 0
        before_paid = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).join(
            Invoice, Invoice.id == Payment.invoice_id
        ).filter(Invoice.invoice_date < from_dt, Invoice.status != 'cancelled').scalar() or 0
        opening_balance = float(before_inv) - float(before_paid)

    outstanding_amount = db.session.query(func.coalesce(func.sum(Invoice.balance_due), 0)).filter(
        Invoice.status != 'cancelled', Invoice.balance_due > 0
    ).scalar() or 0

    pending_invoices_count = Invoice.query.filter(
        Invoice.status != 'cancelled', Invoice.amount_paid < Invoice.grand_total
    ).count()
    paid_invoices_count = Invoice.query.filter(
        Invoice.status != 'cancelled', Invoice.amount_paid >= Invoice.grand_total
    ).count()

    top_products = (
        db.session.query(InvoiceItem.product_name, func.sum(InvoiceItem.qty).label('total_qty'))
        .group_by(InvoiceItem.product_name)
        .order_by(func.sum(InvoiceItem.qty).desc())
        .limit(5)
        .all()
    )

    if (to_dt - from_dt).days <= 62:
        daily_rows = (db.session.query(func.date(Invoice.invoice_date).label('d'),
                                       func.coalesce(func.sum(Invoice.grand_total), 0).label('t'))
                      .filter(Invoice.invoice_date >= from_dt, Invoice.invoice_date <= to_dt,
                              Invoice.status != 'cancelled')
                      .group_by(text('d')).order_by(text('d')).all())
        row_map = {r[0]: float(r[1]) for r in daily_rows}
        chart_data = [{'label': d.strftime('%d %b'), 'value': row_map.get(d.isoformat(), 0.0)}
                      for d in (from_dt + timedelta(days=i) for i in range((to_dt - from_dt).days + 1))]
    else:
        monthly_rows = (db.session.query(extract('year', Invoice.invoice_date).label('y'),
                                         extract('month', Invoice.invoice_date).label('m'),
                                         func.coalesce(func.sum(Invoice.grand_total), 0).label('t'))
                        .filter(Invoice.invoice_date >= from_dt, Invoice.invoice_date <= to_dt,
                                Invoice.status != 'cancelled')
                        .group_by(text('y'), text('m')).order_by(text('y, m')).all())
        chart_data = [{'label': '%s-%02d' % (int(r[0]), int(r[1])), 'value': float(r[2])} for r in monthly_rows]

    outstanding_customers = (
        db.session.query(Customer.id, Customer.name,
                         func.coalesce(func.sum(Invoice.balance_due), 0).label('due'))
        .join(Invoice, Invoice.customer_id == Customer.id)
        .filter(Invoice.status != 'cancelled', Invoice.balance_due > 0)
        .group_by(Customer.id, Customer.name)
        .order_by(desc('due')).limit(10).all()
    )

    activity = []
    for inv in Invoice.query.order_by(desc(Invoice.created_at)).limit(5).all():
        activity.append({'type': 'invoice', 'title': 'Invoice created', 'ref': inv.invoice_number,
                         'amount': inv.grand_total, 'date': inv.created_at or datetime.utcnow(),
                         'link': url_for('invoice_preview', iid=inv.id)})
    for q in Quotation.query.order_by(desc(Quotation.created_at)).limit(5).all():
        activity.append({'type': 'quotation', 'title': 'Quotation %s' % q.status, 'ref': q.quotation_number,
                         'amount': q.grand_total, 'date': q.created_at or datetime.utcnow(),
                         'link': url_for('view_quotation', qid=q.id)})
    for p in Payment.query.order_by(desc(Payment.created_at)).limit(5).all():
        activity.append({'type': 'payment', 'title': 'Payment received', 'ref': _payment_display_ref(p) or ('#%s' % p.id),
                         'amount': p.amount, 'date': p.created_at or datetime.utcnow(),
                         'link': url_for('invoice_preview', iid=p.invoice_id)})
    activity.sort(key=lambda a: a['date'], reverse=True)
    activity = activity[:8]

    return render_template('admin/dashboard.html',
        total_invoices=Invoice.query.count(), total_customers=Customer.query.count(),
        total_products=Product.query.count(), total_revenue=db.session.query(func.sum(Invoice.grand_total)).scalar() or 0,
        pending_invoices=pending_invoices_count,
        paid_invoices=paid_invoices_count,
        outstanding_amount=outstanding_amount,
        range_sales=range_sales, range_collections=range_collections,
        range_invoices_count=range_invoices_count, opening_balance=opening_balance,
        from_date=from_dt, to_date=to_dt, range_label=range_label, selected_range=rng,
        low_stock_alerts=Product.query.filter(Product.stock_quantity <= Product.min_stock).count(),
        low_stock_products=Product.query.filter(Product.stock_quantity <= Product.min_stock).order_by(Product.stock_quantity.asc()).limit(10).all(),
        recent_invoices=Invoice.query.order_by(desc(Invoice.created_at)).limit(5).all(),
        top_products=top_products, chart_data=chart_data,
        outstanding_customers=outstanding_customers, activity=activity,
        now_date=today.strftime('%d %B %Y'))


def _sales_dashboard():
    today_start = date.today()
    base = Invoice.query.filter_by(created_by=current_user.id)
    recent = base.order_by(desc(Invoice.created_at)).limit(10).all()
    today_invoices = base.filter(func.date(Invoice.invoice_date) == today_start).count()
    today_amount = db.session.query(func.coalesce(func.sum(Invoice.grand_total), 0)).filter(
        Invoice.created_by == current_user.id,
        func.date(Invoice.invoice_date) == today_start,
    ).scalar() or 0
    recent_customers = Customer.query.order_by(desc(Customer.created_at)).limit(5).all()
    return render_template('sales/dashboard.html',
        recent_invoices=recent, today_invoices=today_invoices, today_amount=today_amount,
        recent_customers=recent_customers,
        now_date=today_start.strftime('%d %B %Y'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    user = db.session.get(User, current_user.id)
    if request.method == 'POST':
        user.full_name = request.form.get('full_name', user.full_name); user.email = request.form.get('email', user.email)
        pw = request.form.get('new_password', '').strip()
        if pw: user.set_password(pw)
        db.session.commit(); flash('Profile updated.', 'success'); return redirect(url_for('profile'))
    return render_template('admin/profile.html', user=user)


@app.route('/users')
@admin_required
def users_list():
    return render_template('admin/users.html', users=User.query.order_by(User.created_at.desc()).all())


@app.route('/users/add', methods=['POST'])
@admin_required
def create_user():
    u = User(username=request.form['username'].strip(), email=request.form['email'].strip(), full_name=request.form['full_name'].strip(), role=request.form.get('role', 'staff'), is_active=True)
    u.set_password(request.form['password']); db.session.add(u); db.session.commit()
    flash('User created.', 'success'); return redirect(url_for('users_list'))


@app.route('/users/<int:user_id>/edit', methods=['POST'])
@admin_required
def edit_user(user_id):
    u = db.session.get(User, user_id)
    if u:
        u.username = request.form.get('username', u.username).strip(); u.email = request.form.get('email', u.email).strip()
        u.full_name = request.form.get('full_name', u.full_name).strip(); u.role = request.form.get('role', u.role)
        pw = request.form.get('password', '').strip()
        if pw: u.set_password(pw)
        db.session.commit(); flash('User updated.', 'success')
    return redirect(url_for('users_list'))


@app.route('/users/<int:user_id>/toggle', methods=['POST'])
@admin_required
def toggle_user(user_id):
    u = db.session.get(User, user_id)
    if u: u.is_active = not u.is_active; db.session.commit(); flash(f'User {u.username} {"activated" if u.is_active else "deactivated"}.', 'success')
    return redirect(url_for('users_list'))


@app.route('/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    u = db.session.get(User, user_id)
    if u and u.role != 'admin':
        db.session.delete(u); db.session.commit(); flash('User deleted.', 'success')
    elif u and u.role == 'admin':
        flash('Cannot delete admin user.', 'danger')
    return redirect(url_for('users_list'))


def _fmt_bytes(n):
    if not n: return '—'
    for unit in ['B', 'KB', 'MB', 'GB']:
        if n < 1024:
            return f'{n:.1f} {unit}'
        n /= 1024
    return f'{n:.1f} TB'


@app.route('/settings', methods=['GET', 'POST'])
@admin_required
def settings_page():
    if request.method == 'POST':
        # Server-side re-validation of key identity fields (never fail silently).
        # Use the required official GSTIN format (15 chars), per requirements.
        _gst_fmt = re.compile(r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[0-9A-Z]{1}Z[0-9A-Z]{1}$')
        gstin = (request.form.get('company_gstin') or '').strip().upper()
        if gstin and not _gst_fmt.match(gstin):
            msg = 'Invalid GSTIN — must be 15 characters in the official format.'
            current_app.logger.warning('Settings save rejected: %s (gstin=%r)', msg, gstin)
            log_audit(current_user.id, 'settings_save_failed', 'settings', details=msg)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'error': msg}), 200
            flash(msg, 'danger')
            return redirect(url_for('settings_page'))
        try:
            # SMTP / email credentials are backend-only (from .env) and must never
            # be writable via the Settings endpoint. Ignore any such keys defensively.
            # Company identity is FIXED for this installation and is never editable
            # from the admin panel, so those keys are ignored as well.
            _forbidden = {'smtp_server', 'smtp_port', 'smtp_email', 'smtp_password', 'mail_server', 'mail_port', 'mail_username', 'mail_password'}
            _fixed_company_keys = {'company_name', 'company_gstin', 'company_phone', 'company_mobile',
                                   'company_website', 'company_email', 'company_state', 'company_state_code',
                                   'company_address', 'company_city', 'company_pincode', 'company_country'}
            for key, val in request.form.items():
                if key.startswith('file_'):
                    continue
                if key in _forbidden or key in _fixed_company_keys:
                    continue
                if key == 'theme':
                    current_user.theme = val.strip() if isinstance(val, str) else val
                    set_setting('theme', val)
                    continue  # theme also drives /settings/theme
                set_setting(key, val or '')
            db.session.commit()
            log_audit(current_user.id, 'settings_updated', 'settings', details='Company settings updated')
            _load_company_settings(current_app)
            current_app.logger.info('Settings saved successfully by user %s', current_user.username)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True, 'message': 'Settings saved successfully.', 'settings': _public_settings()})
            flash('Settings saved successfully.', 'success')
            return redirect(url_for('settings_page'))
        except Exception:
            db.session.rollback()
            log.exception('Settings save failed')
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'error': 'Database error.'}), 200
            flash('Database error while saving settings.', 'danger')
            return redirect(url_for('settings_page'))
    _load_company_settings(current_app)
    email_configured = bool((os.getenv('MAIL_PASSWORD') or os.getenv('SMTP_PASSWORD') or '').strip()
                            and (os.getenv('MAIL_USERNAME') or os.getenv('SMTP_EMAIL') or '').strip())
    import flask as _flask, platform as _platform, socket as _socket
    _db_path = current_app.config.get('SQLALCHEMY_DATABASE_URI', 'sqlite:///gv_powers.db')
    sys_info = {
        'erp_version': '1.0.0',
        'flask_version': _flask.__version__,
        'python_version': sys.version.split()[0],
        'os_name': f'{_platform.system()} {_platform.release()}',
        'db_status': 'Connected',
        'db_size': _fmt_bytes(os.path.getsize(os.path.join(current_app.root_path, 'gv_powers.db'))) if os.path.exists(os.path.join(current_app.root_path, 'gv_powers.db')) else '—',
        'server_time': datetime.now().strftime('%d %b %Y %H:%M:%S'),
        'cloudflare': os.getenv('CLOUDFLARE_STATUS') or 'N/A',
        'environment': os.getenv('FLASK_ENV', 'production'),
        'uptime': '—',
    }
    _backups = BackupService.list_backups(current_app._get_current_object())
    _last = _backups[0] if _backups else None
    db_size = _fmt_bytes(os.path.getsize(os.path.join(current_app.root_path, 'gv_powers.db'))) if os.path.exists(os.path.join(current_app.root_path, 'gv_powers.db')) else '—'
    sys_info['db_size'] = db_size
    return render_template('admin/settings.html', settings=_public_settings(), company=current_app.config.get('COMPANY', {}), email_configured=email_configured, sys_info=sys_info, db_size=db_size, last_backup=_last['filepath'] if _last else None, backups=_backups)


@app.route('/audit-logs')
@admin_required
def audit_logs():
    page = request.args.get('page', 1, type=int)
    action_filter = request.args.get('action', '').strip()
    q = AuditLog.query
    if action_filter:
        q = q.filter(AuditLog.action.ilike(f'%{action_filter}%'))
    return render_template('admin/audit_logs.html', logs=q.order_by(desc(AuditLog.created_at)).paginate(page=page, per_page=50), action_filter=action_filter)


@app.route('/backup')
@admin_required
def backup_page():
    return render_template('admin/backup.html', backups=BackupService.list_backups(current_app._get_current_object()))


@app.route('/backup/create', methods=['POST'])
@admin_required
def create_backup():
    try:
        fp = BackupService.create_backup(current_app._get_current_object())
        log_audit(current_user.id, 'backup_created', details=f'Backup: {fp}'); flash('Backup created.', 'success')
    except Exception as e: flash(f'Backup failed: {str(e)}', 'danger')
    return redirect(url_for('backup_page'))


@app.route('/backup/download/<filename>')
@admin_required
def download_backup(filename):
    backup_dir = current_app.config.get('BACKUP_FOLDER', 'backups/database')
    filepath = os.path.join(backup_dir, secure_filename(filename))
    if not os.path.exists(filepath): abort(404)
    return send_file(filepath, as_attachment=True)


@app.route('/backup/restore', methods=['POST'])
@admin_required
def restore_backup():
    flash('Restore functionality is under development.', 'info')
    return redirect(url_for('backup_page'))


# ----- Customers -----

@app.route('/customers')
@login_required
def customers_list():
    q = Customer.query; search = request.args.get('q', '').strip()
    if search: q = q.filter(or_(Customer.name.ilike(f'%{search}%'), Customer.mobile.ilike(f'%{search}%'), Customer.email.ilike(f'%{search}%'), Customer.gstin.ilike(f'%{search}%')))
    return render_template('customers/customers.html', customers=q.order_by(desc(Customer.created_at)).all(), search=search)


@app.route('/customers/add', methods=['POST'])
@login_required
def create_customer():
    gstin = request.form.get('gstin', '').strip()
    if gstin and not validate_gstIN(gstin): flash('Invalid GSTIN.', 'danger'); return redirect(url_for('customers_list'))
    c = Customer(name=request.form['name'].strip(), mobile=request.form.get('mobile', '').strip(), email=request.form.get('email', '').strip(), address=request.form.get('address', '').strip(), gstin=gstin, state=request.form.get('state', ''), state_code=int(request.form.get('state_code', 29)))
    db.session.add(c); db.session.commit()
    log_audit(current_user.id, 'customer_created', 'customer', c.id, f'Created: {c.name}')
    flash('Customer added.', 'success')
    if current_user.role == 'sales':
        return redirect(url_for('new_invoice'))
    return redirect(url_for('customers_list'))


@app.route('/customers/<int:cid>')
@login_required
def customer_profile(cid):
    c = db.session.get(Customer, cid)
    if not c: abort(404)
    invs = Invoice.query.filter_by(customer_id=cid).order_by(desc(Invoice.invoice_date)).all()
    active = [i for i in invs if i.status != 'cancelled']
    total_paid = sum((i.amount_paid or 0) for i in active)
    outstanding = sum((i.balance_due if i.balance_due is not None else i.balance) for i in active)
    payments = (db.session.query(Payment)
                .join(Invoice, Invoice.id == Payment.invoice_id)
                .filter(Invoice.customer_id == cid, Invoice.status != 'cancelled')
                .order_by(desc(Payment.payment_date), desc(Payment.id)).limit(50).all())
    return render_template('customers/customer_profile.html', customer=c, invoices=invs,
                           total_paid=total_paid, outstanding=outstanding, payments=payments)


def _customer_ledger_data(cid, from_date=None, to_date=None):
    """Build the account statement for a customer from real records.

    Ledger entries are computed strictly from Invoice (debit) and Payment
    (credit) records and are never editable. ``opening`` carries the balance
    from before ``from_date`` (zero when no start date is given). Cancelled
    invoices are excluded.
    """
    from_dt = from_date or date.min
    to_dt = to_date or date.today()
    opening = Decimal('0')
    if from_date is not None:
        inv_before = db.session.query(func.coalesce(func.sum(Invoice.grand_total), 0)).filter(
            Invoice.customer_id == cid, Invoice.status != 'cancelled',
            Invoice.invoice_date < from_dt).scalar() or 0
        paid_before = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).join(
            Invoice, Invoice.id == Payment.invoice_id).filter(
            Invoice.customer_id == cid, Invoice.status != 'cancelled',
            Payment.payment_date < from_dt).scalar() or 0
        opening = (Decimal(inv_before) - Decimal(paid_before)).quantize(Decimal('0.01'))

    invoices = (Invoice.query.filter(Invoice.customer_id == cid,
                                     Invoice.status != 'cancelled',
                                     Invoice.invoice_date >= from_dt,
                                     Invoice.invoice_date <= to_dt)
                .order_by(Invoice.invoice_date, Invoice.id).all())
    payments = (db.session.query(Payment).join(Invoice, Invoice.id == Payment.invoice_id)
                .filter(Invoice.customer_id == cid, Invoice.status != 'cancelled',
                        Payment.payment_date >= from_dt, Payment.payment_date <= to_dt)
                .order_by(Payment.payment_date, Payment.id).all())

    entries = []
    _q2 = Decimal('0.01')
    for inv in invoices:
        entries.append({
            'date': inv.invoice_date, 'sort': (inv.invoice_date, 0, inv.id),
            'type': 'invoice', 'reference': inv.invoice_number,
            'description': 'Invoice %s' % inv.invoice_number,
            'debit': Decimal(inv.grand_total or 0).quantize(_q2), 'credit': Decimal('0'),
            'link': url_for('view_invoice', iid=inv.id),
        })
    for p in payments:
        entries.append({
            'date': p.payment_date, 'sort': (p.payment_date, 1, p.id),
            'type': 'payment',
            'reference': _payment_display_ref(p) or ('#%s' % p.id),
            'description': 'Payment received%s' % (' on %s' % p.invoice.invoice_number if p.invoice else ''),
            'debit': Decimal('0'), 'credit': Decimal(p.amount or 0).quantize(_q2),
            'link': url_for('view_invoice', iid=p.invoice_id),
        })
    entries.sort(key=lambda e: e['sort'])

    balance = opening
    for e in entries:
        balance += e['debit'] - e['credit']
        e['balance'] = balance.quantize(_q2)

    total_debit = sum((e['debit'] for e in entries), Decimal('0')).quantize(_q2)
    total_credit = sum((e['credit'] for e in entries), Decimal('0')).quantize(_q2)
    return {
        'opening': opening, 'entries': entries,
        'total_debit': total_debit, 'total_credit': total_credit,
        'closing': (opening + total_debit - total_credit).quantize(_q2),
        'from_date': from_dt, 'to_date': to_dt,
    }


def _parse_ledger_dates():
    from_str = request.args.get('from', '').strip()
    to_str = request.args.get('to', '').strip()
    from_dt = datetime.strptime(from_str, '%Y-%m-%d').date() if from_str else None
    to_dt = datetime.strptime(to_str, '%Y-%m-%d').date() if to_str else date.today()
    return from_dt, to_dt


@app.route('/customers/<int:cid>/ledger')
@login_required
def customer_ledger(cid):
    c = db.session.get(Customer, cid)
    if not c:
        abort(404)
    from_dt, to_dt = _parse_ledger_dates()
    if from_dt and to_dt and from_dt > to_dt:
        flash('From date cannot be after the To date.', 'danger')
        from_dt, to_dt = None, date.today()
    data = _customer_ledger_data(cid, from_dt, to_dt)
    return render_template('customers/customer_ledger.html', customer=c, data=data,
                           from_date=from_dt, to_date=to_dt, today=date.today())


@app.route('/customers/<int:cid>/ledger/export/<fmt>')
@login_required
def customer_ledger_export(cid, fmt):
    c = db.session.get(Customer, cid)
    if not c or fmt not in ('pdf', 'excel', 'csv'):
        abort(404)
    from_dt, to_dt = _parse_ledger_dates()
    data = _customer_ledger_data(cid, from_dt, to_dt)
    fname = 'ledger_%s_%s' % (c.name.replace(' ', '_').lower(), date.today().strftime('%Y%m%d'))
    headers = ['Date', 'Type', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    rows = []
    for e in data['entries']:
        rows.append([e['date'].isoformat(), e['type'].title(), e['reference'], e['description'],
                     float(e['debit']), float(e['credit']), float(e['balance'])])
    rows.append(['', 'Opening balance', '', '', float(data['opening']), '', ''])
    rows.sort(key=lambda r: r[0])
    rows.append(['', 'Total', '', '', float(data['total_debit']), float(data['total_credit']), ''])
    rows.append(['', 'Closing balance', '', '', '', '', float(data['closing'])])

    if fmt == 'csv':
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(headers)
        w.writerows(rows)
        buf = io.BytesIO(out.getvalue().encode('utf-8-sig'))
        return send_file(buf, mimetype='text/csv', as_attachment=True, download_name=fname + '.csv')

    if fmt == 'excel':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ledger'
        ws.append(['Customer Ledger', c.name])
        ws.append(headers)
        for r in rows:
            ws.append(r)
        _style_excel(ws)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=fname + '.xlsx')

    pdf_buf = _build_report_pdf('Ledger - %s' % c.name, 'Account statement', headers, rows, {})
    return send_file(pdf_buf, mimetype='application/pdf', as_attachment=True, download_name=fname + '.pdf')


@app.route('/api/v1/customers/<int:cid>/ledger')
@login_required
def api_customer_ledger(cid):
    c = db.session.get(Customer, cid)
    if not c:
        return jsonify({'success': False, 'error': 'Customer not found'}), 404
    from_dt, to_dt = _parse_ledger_dates()
    data = _customer_ledger_data(cid, from_dt, to_dt)
    return jsonify({
        'success': True,
        'customer': {'id': c.id, 'name': c.name},
        'opening': str(data['opening']), 'closing': str(data['closing']),
        'total_debit': str(data['total_debit']), 'total_credit': str(data['total_credit']),
        'entries': [{'date': e['date'].isoformat(), 'type': e['type'], 'reference': e['reference'],
                     'description': e['description'], 'debit': str(e['debit']),
                     'credit': str(e['credit']), 'balance': str(e['balance'])} for e in data['entries']],
    })


@app.route('/customers/<int:cid>/edit', methods=['POST'])
@login_required
def edit_customer(cid):
    c = db.session.get(Customer, cid)
    if not c: abort(404)
    c.name = request.form['name'].strip(); c.mobile = request.form.get('mobile', '').strip(); c.email = request.form.get('email', '').strip()
    c.address = request.form.get('address', '').strip(); c.gstin = request.form.get('gstin', '').strip()
    c.state = request.form.get('state', ''); c.state_code = int(request.form.get('state_code', 29))
    db.session.commit(); log_audit(current_user.id, 'customer_updated', 'customer', c.id)
    flash('Customer updated.', 'success'); return redirect(url_for('customer_profile', cid=c.id))


@app.route('/customers/<int:cid>/delete', methods=['POST'])
@login_required
def delete_customer(cid):
    c = db.session.get(Customer, cid)
    if c: db.session.delete(c); db.session.commit(); log_audit(current_user.id, 'customer_deleted', 'customer', cid); flash('Customer deleted.', 'success')
    return redirect(url_for('customers_list'))


# ----- Products & Stock Management -----

def _save_product_image(file):
    if not file or not getattr(file, 'filename', ''):
        return None
    folder = os.path.join(BASE_DIR, 'static', 'uploads', 'products')
    os.makedirs(folder, exist_ok=True)
    fname = secure_filename(file.filename or '')
    if not fname:
        return None
    name, ext = os.path.splitext(fname)
    fname = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{name}{ext}".lower()
    file.save(os.path.join(folder, fname))
    return 'uploads/products/' + fname


def _stock_product_or_404(pid):
    p = db.session.get(Product, pid)
    if not p:
        abort(404)
    return p


def _record_movement(pid, movement_type, quantity, reference_type=None, reference_id=None, notes=None):
    db.session.add(StockMovement(product_id=pid, movement_type=movement_type, quantity=quantity,
                                 reference_type=reference_type, reference_id=reference_id, notes=notes,
                                 user_id=current_user.id))


def _low_stock_suggested_qty(p):
    """Suggested purchase quantity for a low-stock product."""
    if not p or not (p.min_stock or 0):
        return 1
    return max(int(p.min_stock) * 2 - int(p.stock_quantity or 0), 1)


def _run_low_stock_check():
    """Scan products at/below minimum stock and notify admins/managers.

    Notifications are emitted once per low-stock episode (deduplicated through
    ``low_stock_alert_active`` + ``last_low_stock_notification_at``) and are
    auto-resolved when stock recovers above the minimum. Call this after any
    stock-affecting operation (invoice, conversion, PO receive) and on
    dashboard/report views so alerts stay current.
    """
    try:
        low = Product.query.filter(Product.is_active == True,
                                   Product.min_stock > 0,
                                   Product.stock_quantity <= Product.min_stock,
                                   Product.stock_quantity > 0).all()
        changed = False
        for p in low:
            if p.low_stock_alert_active:
                continue
            if p.last_low_stock_notification_at and \
                    (datetime.utcnow() - p.last_low_stock_notification_at) < timedelta(hours=24):
                continue
            p.low_stock_alert_active = True
            p.last_low_stock_notification_at = datetime.utcnow()
            suggested = _low_stock_suggested_qty(p)
            msg = f'{p.name} has only {p.stock_quantity} unit(s) left (min: {p.min_stock}). Suggested purchase: {suggested}.'
            for u in User.query.filter(User.role.in_(['admin', 'manager'])).all():
                db.session.add(Notification(user_id=u.id, title='Low stock alert', message=msg,
                                            notification_type='warning'))
            log_audit(current_user.id if hasattr(current_user, 'id') else None,
                      'low_stock_generated', 'product', p.id, msg)
            changed = True
        resolved = Product.query.filter(Product.low_stock_alert_active == True,
                                        Product.stock_quantity > Product.min_stock).all()
        for p in resolved:
            p.low_stock_alert_active = False
            log_audit(current_user.id if hasattr(current_user, 'id') else None,
                      'low_stock_resolved', 'product', p.id,
                      f'Stock recovered to {p.stock_quantity} (> min {p.min_stock})')
            changed = True
        if changed:
            db.session.commit()
    except Exception:
        db.session.rollback()


@app.route('/products')
@login_required
def products_list():
    q = Product.query; search = request.args.get('q', '').strip(); cat = request.args.get('category', type=int)
    stock_filter = request.args.get('stock', '').strip()
    if search: q = q.filter(or_(Product.name.ilike(f'%{search}%'), Product.sku.ilike(f'%{search}%'), Product.barcode.ilike(f'%{search}%'), Product.hsn.ilike(f'%{search}%'), Product.brand.ilike(f'%{search}%')))
    if cat: q = q.filter_by(category_id=cat)
    if stock_filter == 'in_stock': q = q.filter(Product.stock_quantity > 0)
    elif stock_filter == 'low_stock': q = q.filter(Product.stock_quantity > 0, Product.stock_quantity <= Product.min_stock)
    elif stock_filter == 'out_of_stock': q = q.filter(Product.stock_quantity <= 0)
    return render_template('products/products.html', products=q.order_by(Product.name).all(), categories=Category.query.order_by(Category.name).all(), search=search, selected_category=cat, stock_filter=stock_filter, suppliers=Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all())


@app.route('/products/<int:pid>')
@login_required
def product_profile(pid):
    p = _stock_product_or_404(pid)
    movements = StockMovement.query.filter_by(product_id=pid).order_by(desc(StockMovement.created_at)).limit(100).all()
    balance = p.current_stock
    for m in movements:
        m.balance = balance
        balance -= m.quantity or 0
    return render_template('products/product_profile.html', product=p,
        movements=movements, date=date,
        category=db.session.get(Category, p.category_id) if p.category_id else None)


@app.route('/products/add', methods=['POST'])
@login_required
def create_product():
    stock = to_int(request.form.get('stock_quantity'), 0) or to_int(request.form.get('opening_stock'), 0) or 0
    p = Product(
        name=request.form['name'].strip(), sku=request.form.get('sku', '').strip() or None,
        barcode=request.form.get('barcode', '').strip() or None, hsn=request.form.get('hsn', '').strip(),
        brand=request.form.get('brand', '').strip(), category_id=to_int(request.form.get('category_id')),
        supplier_id=to_int(request.form.get('supplier_id')), description=request.form.get('description', '').strip(),
        unit=request.form.get('unit', 'pcs'), purchase_price=Decimal(request.form.get('purchase_price', '0') or '0'),
        selling_price=Decimal(request.form.get('selling_price', '0') or '0'), gst_rate=Decimal(request.form.get('gst_rate', '18') or '18'),
        opening_stock=stock, stock_quantity=stock, min_stock=to_int(request.form.get('min_stock'), 0),
        max_stock=to_int(request.form.get('max_stock'), 500), location=request.form.get('location', '').strip(),
        status=request.form.get('status', 'active'), warehouse=request.form.get('warehouse', '').strip(),
        warranty=request.form.get('warranty', '').strip(), image=_save_product_image(request.files.get('image')),
        is_active=request.form.get('is_active', 'on') == 'on')
    db.session.add(p)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        bc = request.form.get('barcode', '').strip()
        sku = request.form.get('sku', '').strip()
        if bc:
            dup = Product.query.filter_by(barcode=bc).first()
            if dup:
                flash(f'Barcode "{bc}" already exists on product "{dup.name}".', 'danger')
                return redirect(url_for('products_list'))
        if sku:
            dup = Product.query.filter_by(sku=sku).first()
            if dup:
                flash(f'SKU "{sku}" already exists on product "{dup.name}".', 'danger')
                return redirect(url_for('products_list'))
        flash('A product with that SKU or barcode already exists.', 'danger')
        return redirect(url_for('products_list'))
    if stock > 0:
        _record_movement(p.id, 'opening', stock, reference_type='opening_stock', notes='Opening stock on creation')
        db.session.commit()
    log_audit(current_user.id, 'product_created', 'product', p.id)
    flash('Product added.', 'success'); return redirect(url_for('products_list'))


@app.route('/products/<int:pid>/edit', methods=['POST'])
@login_required
def edit_product(pid):
    p = _stock_product_or_404(pid)
    p.name = request.form['name'].strip(); p.sku = request.form.get('sku', '').strip() or None
    p.barcode = request.form.get('barcode', '').strip() or None; p.hsn = request.form.get('hsn', '').strip()
    p.brand = request.form.get('brand', '').strip(); p.category_id = to_int(request.form.get('category_id'))
    p.supplier_id = to_int(request.form.get('supplier_id')); p.description = request.form.get('description', '').strip()
    p.unit = request.form.get('unit', 'pcs'); p.purchase_price = Decimal(request.form.get('purchase_price', '0') or '0')
    p.selling_price = Decimal(request.form.get('selling_price', '0') or '0'); p.gst_rate = Decimal(request.form.get('gst_rate', '18') or '18')
    p.min_stock = to_int(request.form.get('min_stock'), 0); p.max_stock = to_int(request.form.get('max_stock'), 500)
    p.location = request.form.get('location', '').strip(); p.status = request.form.get('status', p.status or 'active')
    p.warehouse = request.form.get('warehouse', '').strip(); p.warranty = request.form.get('warranty', '').strip()
    p.is_active = request.form.get('is_active', 'on') == 'on'
    img = _save_product_image(request.files.get('image'))
    if img: p.image = img
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        bc = request.form.get('barcode', '').strip()
        sku = request.form.get('sku', '').strip()
        if bc:
            dup = Product.query.filter(Product.barcode == bc, Product.id != pid).first()
            if dup:
                flash(f'Barcode "{bc}" already exists on product "{dup.name}".', 'danger')
                return redirect(url_for('product_profile', pid=p.id))
        if sku:
            dup = Product.query.filter(Product.sku == sku, Product.id != pid).first()
            if dup:
                flash(f'SKU "{sku}" already exists on product "{dup.name}".', 'danger')
                return redirect(url_for('product_profile', pid=p.id))
        flash('A product with that SKU or barcode already exists.', 'danger')
        return redirect(url_for('product_profile', pid=p.id))
    log_audit(current_user.id, 'product_updated', 'product', p.id); flash('Product updated.', 'success')
    return redirect(url_for('product_profile', pid=p.id))


@app.route('/products/<int:pid>/stock/add', methods=['POST'])
@login_required
def add_stock(pid):
    p = _stock_product_or_404(pid)
    try:
        qty = int(request.form.get('quantity', '0') or '0')
    except (TypeError, ValueError):
        qty = 0
    if qty <= 0:
        flash('Enter a valid quantity to add.', 'danger')
        return redirect(url_for('product_profile', pid=p.id))
    purchase_cost = Decimal(request.form.get('purchase_cost', '0') or '0')
    if purchase_cost > 0:
        p.purchase_price = purchase_cost
    supplier = request.form.get('supplier', '').strip()
    purchase_date = request.form.get('purchase_date', '').strip()
    notes = request.form.get('remarks', '').strip()
    if supplier: notes = (f'Supplier: {supplier}. ' + notes).strip()
    p.stock_quantity = (p.stock_quantity or 0) + qty
    pd = None
    if purchase_date:
        try: pd = datetime.strptime(purchase_date, '%Y-%m-%d')
        except ValueError: pd = None
    p.last_purchase = pd or datetime.utcnow()
    _record_movement(p.id, 'purchase', qty, reference_type='manual', notes=notes or 'Stock added manually')
    db.session.commit()
    log_audit(current_user.id, 'stock_added', 'product', p.id, f'+{qty} units, New: {p.current_stock}')
    flash(f'{qty} units added. New stock: {p.current_stock}', 'success')
    return redirect(url_for('product_profile', pid=p.id))


@app.route('/products/<int:pid>/stock/remove', methods=['POST'])
@login_required
def remove_stock(pid):
    p = _stock_product_or_404(pid)
    try:
        qty = int(request.form.get('quantity', '0') or '0')
    except (TypeError, ValueError):
        qty = 0
    if qty <= 0:
        flash('Enter a valid quantity to remove.', 'danger')
        return redirect(url_for('product_profile', pid=p.id))
    if qty > p.current_stock:
        flash(f'Cannot remove {qty} units. Only {p.current_stock} units available.', 'danger')
        return redirect(url_for('product_profile', pid=p.id))
    reason = request.form.get('reason', '').strip() or 'adjustment'
    remarks = request.form.get('remarks', '').strip()
    notes = f'{reason}{": " + remarks if remarks else ""}'
    p.stock_quantity = p.current_stock - qty
    _record_movement(p.id, 'adjustment', -qty, reference_type='manual', notes=notes)
    db.session.commit()
    log_audit(current_user.id, 'stock_removed', 'product', p.id, f'-{qty} units ({reason}), New: {p.current_stock}')
    flash(f'{qty} units removed. New stock: {p.current_stock}', 'success')
    return redirect(url_for('product_profile', pid=p.id))


@app.route('/products/<int:pid>/stock/delete', methods=['POST'])
@admin_required
def delete_stock(pid):
    p = _stock_product_or_404(pid)
    confirm = (request.form.get('confirm_delete', '') or '').strip()
    if confirm != 'DELETE':
        flash('You must type DELETE to confirm stock deletion.', 'danger')
        return redirect(url_for('product_profile', pid=p.id))
    old = p.current_stock or 0
    if old <= 0:
        flash('Stock is already empty.', 'warning')
        return redirect(url_for('product_profile', pid=p.id))
    p.stock_quantity = 0
    _record_movement(p.id, 'reset', -old, reference_type='manual', notes='Stock deleted by admin')
    db.session.commit()
    log_audit(current_user.id, 'stock_deleted', 'product', p.id, f'Removed {old} units, reset to 0')
    flash('Stock deleted. Quantity reset to 0.', 'success')
    return redirect(url_for('product_profile', pid=p.id))


@app.route('/products/<int:pid>/delete', methods=['POST'])
@login_required
def delete_product(pid):
    p = db.session.get(Product, pid)
    if p:
        StockMovement.query.filter_by(product_id=pid).delete(synchronize_session=False)
        db.session.delete(p); db.session.commit(); log_audit(current_user.id, 'product_deleted', 'product', pid); flash('Product deleted.', 'success')
    return redirect(url_for('products_list'))


@app.route('/products/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_products():
    ids = request.form.getlist('product_ids')
    if ids:
        StockMovement.query.filter(StockMovement.product_id.in_(ids)).delete(synchronize_session=False)
        Product.query.filter(Product.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit(); log_audit(current_user.id, 'products_bulk_deleted', details=f'Deleted {len(ids)}'); flash(f'{len(ids)} products deleted.', 'success')
    return redirect(url_for('products_list'))



@app.route('/categories', methods=['POST'])
@login_required
def create_category():
    name = request.form.get('name', '').strip()
    if name: db.session.add(Category(name=name, description=request.form.get('description', '').strip())); db.session.commit(); flash('Category added.', 'success')
    return redirect(url_for('products_list'))


# ----- Suppliers & Purchases -----

@app.route('/suppliers')
@login_required
def suppliers_list():
    q = Supplier.query; search = request.args.get('q', '').strip()
    if search: q = q.filter(or_(Supplier.name.ilike(f'%{search}%'), Supplier.mobile.ilike(f'%{search}%'), Supplier.gstin.ilike(f'%{search}%')))
    return render_template('suppliers/suppliers.html', suppliers=q.order_by(desc(Supplier.created_at)).all(), search=search)


@app.route('/suppliers/add', methods=['POST'])
@login_required
def create_supplier():
    s = Supplier(name=request.form['name'].strip(), contact_person=request.form.get('contact_person', '').strip(), mobile=request.form.get('mobile', '').strip(), email=request.form.get('email', '').strip(), address=request.form.get('address', '').strip(), gstin=request.form.get('gstin', '').strip(), state=request.form.get('state', ''), state_code=int(request.form.get('state_code', 29)))
    db.session.add(s); db.session.commit(); log_audit(current_user.id, 'supplier_created', 'supplier', s.id)
    flash('Supplier added.', 'success'); return redirect(url_for('suppliers_list'))


@app.route('/suppliers/<int:sid>')
@login_required
def supplier_profile(sid):
    s = db.session.get(Supplier, sid)
    if not s: abort(404)
    return render_template('suppliers/supplier_profile.html', supplier=s,
        purchase_orders=PurchaseOrder.query.filter_by(supplier_id=sid).order_by(desc(PurchaseOrder.created_at)).all())


@app.route('/suppliers/<int:sid>/edit', methods=['POST'])
@login_required
def edit_supplier(sid):
    s = db.session.get(Supplier, sid)
    if not s: abort(404)
    s.name = request.form['name'].strip(); s.contact_person = request.form.get('contact_person', '').strip(); s.mobile = request.form.get('mobile', '').strip()
    s.email = request.form.get('email', '').strip(); s.address = request.form.get('address', '').strip(); s.gstin = request.form.get('gstin', '').strip()
    s.state = request.form.get('state', ''); s.state_code = int(request.form.get('state_code', 29))
    db.session.commit(); log_audit(current_user.id, 'supplier_updated', 'supplier', s.id)
    flash('Supplier updated.', 'success'); return redirect(url_for('supplier_profile', sid=s.id))


@app.route('/suppliers/<int:sid>/delete', methods=['POST'])
@login_required
def delete_supplier(sid):
    s = db.session.get(Supplier, sid)
    if s: db.session.delete(s); db.session.commit(); log_audit(current_user.id, 'supplier_deleted', 'supplier', sid); flash('Supplier deleted.', 'success')
    return redirect(url_for('suppliers_list'))


@app.route('/purchase-orders')
@login_required
def purchase_orders():
    return render_template('suppliers/purchase_orders.html', orders=PurchaseOrder.query.order_by(desc(PurchaseOrder.created_at)).all())


@app.route('/purchase-orders/add', methods=['GET', 'POST'])
@login_required
def create_purchase_order():
    if request.method == 'POST':
        po = PurchaseOrder(po_number=generate_purchase_order_number(), supplier_id=int(request.form['supplier_id']), supplier_name=request.form.get('supplier_name', ''), order_date=datetime.strptime(request.form['order_date'], '%Y-%m-%d').date(), expected_date=datetime.strptime(request.form['expected_date'], '%Y-%m-%d').date() if request.form.get('expected_date') else None, notes=request.form.get('notes', ''), status='draft', created_by=current_user.id)
        db.session.add(po); db.session.flush()
        pids = request.form.getlist('product_id[]'); qtys = request.form.getlist('quantity[]'); prices = request.form.getlist('price[]')
        sub = Decimal('0')
        for i in range(len(pids)):
            pid = to_int(pids[i])
            if not pid: continue
            qty = to_int(qtys[i], 1) or 1
            price = Decimal(str(prices[i] or '0'))
            prod = db.session.get(Product, pid)
            total = (price * qty).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            db.session.add(PurchaseItem(purchase_order_id=po.id, product_id=pid,
                product_name=prod.name if prod else 'Product #%s' % pid, hsn=prod.hsn if prod else '',
                qty=qty, unit=prod.unit if prod else 'pcs', price=price, gst_rate=prod.gst_rate if prod else Decimal('18'), total=total))
            sub += total
        po.subtotal = sub; po.total_tax = Decimal('0'); po.grand_total = sub
        db.session.commit(); _run_low_stock_check(); log_audit(current_user.id, 'purchase_order_created', 'purchase_order', po.id)
        flash(f'PO {po.po_number} created.', 'success'); return redirect(url_for('view_purchase_order', oid=po.id))
    prefill = None
    pid = to_int(request.args.get('product_id'))
    if pid:
        prod = db.session.get(Product, pid)
        prefill = {
            'product_id': pid,
            'suggested_qty': max(to_int(request.args.get('suggested_qty')) or 1, 1),
            'price': str(prod.purchase_price) if prod and prod.purchase_price else '0',
            'order_date': date.today().isoformat(),
        }
    return render_template('suppliers/purchase_orders.html', orders=PurchaseOrder.query.order_by(desc(PurchaseOrder.created_at)).all(), suppliers=Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all(), products=Product.query.filter_by(is_active=True).order_by(Product.name).all(), show_add=True, prefill=prefill)


@app.route('/purchase-orders/<int:oid>')
@login_required
def view_purchase_order(oid):
    o = db.session.get(PurchaseOrder, oid)
    if not o: abort(404)
    return render_template('suppliers/purchase_order_detail.html', order=o)


@app.route('/purchase-orders/<int:oid>/delete', methods=['POST'])
@login_required
def delete_purchase_order(oid):
    o = db.session.get(PurchaseOrder, oid)
    if not o:
        flash('Purchase order not found.', 'danger')
        return redirect(url_for('purchase_orders'))
    if o.status == 'received':
        for item in o.items:
            if item.product_id:
                prod = db.session.get(Product, item.product_id)
                if prod:
                    prod.stock_quantity = max(0, (prod.stock_quantity or 0) - item.qty)
                    _record_movement(item.product_id, 'return', -item.qty,
                                     reference_type='purchase_order', reference_id=o.id,
                                     notes=f'Stock reverted on PO deletion: {o.order_number}')
    num = o.order_number
    db.session.delete(o)
    db.session.commit()
    log_audit(current_user.id, 'purchase_order_deleted', 'purchase_order', oid, f'Deleted {num}')
    flash(f'Purchase order {num} deleted.', 'success')
    return redirect(url_for('purchase_orders'))


@app.route('/purchase-orders/<int:oid>/status', methods=['POST'])
@login_required
def update_po_status(oid):
    o = db.session.get(PurchaseOrder, oid)
    if o:
        new_status = request.form.get('status', o.status)
        was_received = o.status == 'received'
        o.status = new_status
        if new_status == 'received' and not was_received:
            for it in o.items:
                if not it.product_id or not (it.qty or 0): continue
                prod = db.session.get(Product, it.product_id)
                if not prod: continue
                prod.stock_quantity = (prod.stock_quantity or 0) + it.qty
                prod.last_purchase = datetime.utcnow()
                _record_movement(prod.id, 'purchase', it.qty, reference_type='purchase_order', reference_id=o.id, notes=f'Received PO {o.po_number}')
            db.session.commit()
            _run_low_stock_check()
            log_audit(current_user.id, 'po_received_stock', 'purchase_order', oid, f'Stock added for {o.po_number}')
        db.session.commit()
        log_audit(current_user.id, 'po_status_updated', 'purchase_order', oid, f'Status: {o.status}')
        flash(f'Status updated to {o.status}.', 'success')
    return redirect(url_for('view_purchase_order', oid=oid))


@app.route('/purchases')
@login_required
def purchases_list():
    return redirect(url_for('purchase_orders'))


############################################################
# SALES ROUTES
############################################################

# ----- Invoices -----

@app.route('/invoices')
@login_required
def invoice_history():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip()
    payment_status = request.args.get('payment_status', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    sales_person_id = request.args.get('sales_person', '').strip()

    base_query = Invoice.query
    if not current_user.is_admin:
        base_query = base_query.filter_by(created_by=current_user.id)

    if status:
        base_query = base_query.filter(Invoice.status == status)
    if search:
        base_query = base_query.filter(
            or_(
                Invoice.invoice_number.ilike(f'%{search}%'),
                Invoice.customer_name.ilike(f'%{search}%'),
                Invoice.customer_mobile.ilike(f'%{search}%'),
                Invoice.customer_gstin.ilike(f'%{search}%'),
            )
        )
    if date_from:
        try:
            base_query = base_query.filter(Invoice.invoice_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to:
        try:
            base_query = base_query.filter(Invoice.invoice_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass
    if sales_person_id:
        try:
            base_query = base_query.filter(Invoice.created_by == int(sales_person_id))
        except ValueError:
            pass

    ZERO = Decimal('0')
    if payment_status == 'paid':
        base_query = base_query.filter(Invoice.status != 'cancelled', Invoice.amount_paid >= Invoice.grand_total)
    elif payment_status in ('pending', 'due'):
        base_query = base_query.filter(Invoice.status != 'cancelled', Invoice.amount_paid < ZERO + Decimal('0.01'))
    elif payment_status == 'partial':
        base_query = base_query.filter(Invoice.status != 'cancelled', Invoice.amount_paid >= ZERO + Decimal('0.01'), Invoice.amount_paid < Invoice.grand_total)

    stats_query = base_query
    total_invoices = stats_query.count()
    revenue_row = stats_query.with_entities(db.func.coalesce(db.func.sum(Invoice.grand_total), ZERO)).first()
    total_revenue = revenue_row[0] if revenue_row else ZERO
    paid_count = stats_query.filter(
        Invoice.status != 'cancelled',
        Invoice.amount_paid >= Invoice.grand_total
    ).count()
    pending_row = stats_query.filter(
        Invoice.status != 'cancelled',
        Invoice.grand_total > Invoice.amount_paid
    ).with_entities(db.func.coalesce(db.func.sum(Invoice.grand_total - Invoice.amount_paid), ZERO)).first()
    total_pending = pending_row[0] if pending_row else ZERO

    sales_users = User.query.filter(User.role.in_(['admin', 'sales']), User.is_active == True).order_by(User.full_name).all()
    invoices = base_query.order_by(desc(Invoice.created_at)).paginate(page=page, per_page=25)

    return render_template(
        'sales/invoice_history.html',
        invoices=invoices, search=search, status=status,
        payment_status=payment_status, date_from=date_from,
        date_to=date_to, sales_person_id=sales_person_id,
        sales_users=sales_users,
        stats={'total': total_invoices, 'revenue': total_revenue, 'pending': total_pending, 'paid_count': paid_count},
        today=date.today(),
    )


def _norm_mobile(s):
    return (s or '').strip().replace(' ', '').replace('-', '')


def _resolve_customer(d):
    """Return the customer a new invoice should attach to.

    Duplicate-detection priority (never create duplicates):
        1. Mobile number (exact)
        2. GSTIN (exact)
        3. Email (exact)
        4. Name + Mobile (exact)
    If an explicit customer_id was supplied it wins. When nothing matches,
    a new Customer record is created automatically from the submitted details.
    """
    cid = to_int(d.get('customer_id'))
    if cid:
        c = db.session.get(Customer, cid)
        if c:
            return c
    mobile = _norm_mobile(d.get('customer_mobile'))
    gstin = (d.get('customer_gstin') or '').strip().upper()
    email = (d.get('customer_email') or '').strip().lower()
    name = (d.get('customer_name') or '').strip()

    if mobile:
        c = Customer.query.filter(Customer.mobile.isnot(None),
                                  func.replace(func.lower(Customer.mobile), ' ', '') == mobile).first()
        if c:
            return c
    if gstin and len(gstin) >= 4:
        c = Customer.query.filter(Customer.gstin.isnot(None), func.upper(Customer.gstin) == gstin).first()
        if c:
            return c
    if email and '@' in email:
        c = Customer.query.filter(Customer.email.isnot(None), func.lower(Customer.email) == email).first()
        if c:
            return c
    if name and mobile:
        c = Customer.query.filter(func.lower(Customer.name) == name.lower(), Customer.mobile.isnot(None),
                                  func.replace(func.lower(Customer.mobile), ' ', '') == mobile).first()
        if c:
            return c
    return _create_customer(d, name)


def _create_customer(d, name=''):
    comp_sc = current_app.config.get('COMPANY_STATE_CODE', 29)
    n = name or (d.get('customer_name') or '').strip() or 'Walk-in Customer'
    sc = int(d.get('customer_state_code', comp_sc) or comp_sc)
    c = Customer(
        name=n,
        mobile=(d.get('customer_mobile') or '').strip() or None,
        email=(d.get('customer_email') or '').strip() or None,
        address=(d.get('customer_address') or '').strip() or None,
        gstin=(d.get('customer_gstin') or '').strip().upper() or None,
        state=(d.get('customer_state') or '').strip() or None,
        state_code=sc,
    )
    db.session.add(c)
    db.session.flush()
    return c


@app.route('/invoices/create', methods=['GET', 'POST'])
@login_required
def new_invoice():
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form
        items_data = data.get('items', [])
        if not items_data:
            pn = data.getlist('product_name[]') if hasattr(data, 'getlist') else data.get('product_name', [])
            if isinstance(pn, str): pn = [pn]
            pids = data.getlist('product_id[]') if hasattr(data, 'getlist') else data.get('product_id', [])
            qtys = data.getlist('qty[]') if hasattr(data, 'getlist') else data.get('qty', [])
            rates = data.getlist('rate[]') if hasattr(data, 'getlist') else data.get('rate', [])
            discs = data.getlist('discount[]') if hasattr(data, 'getlist') else data.get('discount', [])
            gst_r = data.getlist('gst_rate[]') if hasattr(data, 'getlist') else data.get('gst_rate', [])
            hsns = data.getlist('hsn[]') if hasattr(data, 'getlist') else data.get('hsn', [])
            items_data = []
            for i in range(len(pn) if isinstance(pn, (list, tuple)) else 0):
                if not pn[i].strip(): continue
                items_data.append({'product_name': pn[i], 'product_id': int(pids[i]) if pids and i < len(pids) and pids[i] else None, 'qty': int(qtys[i]) if qtys and i < len(qtys) else 1, 'price': Decimal(rates[i]) if rates and i < len(rates) else 0, 'discount': Decimal(discs[i]) if discs and i < len(discs) else 0, 'gst_rate': Decimal(gst_r[i]) if gst_r and i < len(gst_r) else 18, 'hsn': hsns[i] if hsns and i < len(hsns) else ''})
        allow_oos = str(data.get('allow_out_of_stock', '') or '').lower() in ('1', 'true', 'yes', 'on')
        stock_errors = []
        for item in items_data:
            if not item.get('product_name', '').strip(): continue
            pid = to_int(item.get('product_id'))
            if not pid: continue
            try:
                qty = int(item.get('qty', 1))
            except (TypeError, ValueError):
                qty = 1
            prod = db.session.get(Product, pid)
            if not prod: continue
            if qty <= 0:
                stock_errors.append(f'{item["product_name"]}: quantity must be greater than zero.')
            elif not allow_oos and qty > prod.current_stock:
                stock_errors.append(f'{item["product_name"]}: insufficient stock. Only {prod.current_stock} unit(s) available.')
        if stock_errors:
            msg = ' '.join(stock_errors)
            if request.is_json:
                return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('new_invoice'))
        inv_num = None
        try:
            # Allocate the number inside the save transaction: it is only
            # consumed when the invoice commits successfully. On any failure the
            # whole transaction (counter bump + invoice) rolls back together.
            inv_num = generate_invoice_number(commit=False)
            cust = _resolve_customer(data)
            inv_date_str = data.get('invoice_date', '') if isinstance(data, dict) else data.get('invoice_date', '')
            inv_date = datetime.strptime(inv_date_str, '%Y-%m-%d').date() if inv_date_str else date.today()
            due_date_str = data.get('due_date', '')
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else None
            comp_sc = current_app.config.get('COMPANY_STATE_CODE', 29)
            cust_sc = int(data.get('customer_state_code', comp_sc))
            intra = comp_sc == cust_sc
            inv = Invoice(invoice_number=inv_num, customer_id=cust.id, customer_name=cust.name, customer_mobile=cust.mobile or '', customer_email=cust.email or '', customer_address=cust.address or '', customer_gstin=cust.gstin or '', customer_state=cust.state or '', customer_state_code=cust.state_code or cust_sc, invoice_date=inv_date, due_date=due_date, payment_method=data.get('payment_method', ''), is_intra_state=intra, notes=data.get('notes', ''), terms=data.get('terms', ''), created_by=current_user.id)
            db.session.add(inv); db.session.flush()
            sub = Decimal('0'); td = Decimal('0'); tt = Decimal('0'); tc = Decimal('0'); ts = Decimal('0'); ti = Decimal('0')
            for item in items_data:
                if not item.get('product_name', '').strip(): continue
                qty = int(item.get('qty', 1)); rate = Decimal(str(item.get('price', '0'))); dp = Decimal(str(item.get('discount', '0'))); gr = Decimal(str(item.get('gst_rate', '18'))); pid = to_int(item.get('product_id'))
                ls = rate*qty; da = (ls*dp/Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP); tax = ls-da
                gst = GSTService.calculate_gst(tax, gr, intra); lt = tax + gst['total_tax']
                db.session.add(InvoiceItem(invoice_id=inv.id, product_id=pid, product_name=item['product_name'].strip(), hsn=item.get('hsn', ''), qty=qty, price=rate, discount=dp, gst_rate=gr, taxable_value=tax, cgst=gst['cgst'], sgst=gst['sgst'], igst=gst['igst'], total=lt))
                sub += ls; td += da; tt += tax; tc += gst['cgst']; ts += gst['sgst']; ti += gst['igst']
                if pid:
                    prod = db.session.get(Product, pid)
                    if prod:
                        prod.stock_quantity = (prod.stock_quantity or 0) - qty
                        prod.last_sale = datetime.utcnow()
                        _record_movement(pid, 'sale', -qty, reference_type='invoice', reference_id=inv.id, notes=f'Invoice {inv_num}')
            gt = tt+tc+ts+ti; rg = gt.quantize(Decimal('1'), rounding=ROUND_HALF_UP); ro = rg-gt
            inv.subtotal=sub; inv.total_discount=td; inv.total_taxable=tt; inv.total_cgst=tc; inv.total_sgst=ts; inv.total_igst=ti; inv.round_off=ro; inv.grand_total=rg
            ap = Decimal(str(data.get('amount_paid', '0') or '0'))
            inv.amount_paid = ap
            inv.balance_due = rg - ap
            inv.payment_status = 'paid' if ap >= rg else 'partial' if ap > 0 else 'due'
            inv.status = 'completed'
            if inv.customer_id:
                c = db.session.get(Customer, inv.customer_id)
                if c: c.total_purchases = (c.total_purchases or Decimal('0'))+rg; c.invoice_count = (c.invoice_count or 0)+1
            db.session.commit()
        except Exception:
            db.session.rollback()
            if request.is_json:
                return jsonify({'success': False, 'error': 'Invoice could not be saved. No changes were made.'}), 400
            flash('Invoice could not be saved. No changes were made.', 'danger')
            return redirect(url_for('new_invoice'))
        log_audit(current_user.id, 'invoice_created', 'invoice', inv.id, f'{inv_num}, Total: {rg}')

        flash(f'Invoice {inv_num} created.', 'success')
        resp = jsonify({'success': True, 'redirect': url_for('view_invoice', iid=inv.id)}) if request.is_json else redirect(url_for('view_invoice', iid=inv.id))

        @after_this_request
        def _post_invoice_tasks(response):
            if inv.customer_email:
                try:
                    co = current_app.config.get('COMPANY', {})
                    ok, emsg = send_invoice_email(inv, co)
                    if ok:
                        log_audit(current_user.id, 'email_invoice', 'invoice', inv.id, f'Auto-sent to {inv.customer_email}')
                    else:
                        log_audit(current_user.id, 'email_failed', 'invoice', inv.id, emsg)
                except Exception as e:
                    log_audit(current_user.id, 'email_failed', 'invoice', inv.id, str(e))
            _run_low_stock_check()
            return response

        return resp
    inv_num = peek_next_invoice_number()
    return render_template('billing/new_invoice.html', customers=Customer.query.order_by(Customer.name).all(), products=Product.query.filter_by(is_active=True).order_by(Product.name).all(), company=current_app.config.get('COMPANY', {}), today=date.today().isoformat(), invoice_number=inv_num)


def _inv_accessible(inv):
    """Admins may access any invoice; sales may only access invoices they created."""
    if current_user.is_admin:
        return True
    return inv.created_by is not None and inv.created_by == current_user.id


def _jinv_editable(inv):
    """Invoices may be edited only while DRAFT or UNPAID (due) and not cancelled."""
    return inv.status != 'cancelled' and (inv.status == 'draft' or inv.payment_status in ('pending', 'due'))


def _inv_editable_for_user(inv):
    """Highest legitimate edit permission for the current user.

    Admins may edit any non-cancelled invoice (including completed/paid ones,
    preserving payment history); other roles keep the draft/unpaid-only rule.
    Cancelled invoices are always immutable.
    """
    if inv.status == 'cancelled':
        return False
    if current_user.is_admin:
        return True
    return inv.status == 'draft' or inv.payment_status in ('pending', 'due')


@app.route('/invoices/<int:iid>')
@login_required
def view_invoice(iid):
    inv = db.session.get(Invoice, iid)
    if not inv: abort(404)
    if not _inv_accessible(inv): abort(403)
    return render_template('billing/invoice_preview.html', invoice=inv, amount_words=amount_to_words(inv.grand_total), company=current_app.config.get('COMPANY', {}))


@app.route('/invoices/<int:iid>/preview')
@login_required
def invoice_preview(iid):
    inv = db.session.get(Invoice, iid)
    if not inv: abort(404)
    if not _inv_accessible(inv): abort(403)
    return render_template('billing/invoice_preview.html', invoice=inv, amount_words=amount_to_words(inv.grand_total), company=current_app.config.get('COMPANY', {}))


@app.route('/invoices/<int:iid>/edit', methods=['GET', 'POST'])
@login_required
def edit_invoice(iid):
    inv = db.session.get(Invoice, iid)
    if not inv: abort(404)
    if not _inv_accessible(inv):
        if request.is_json: return jsonify({'success': False, 'error': 'You cannot edit another user\'s invoice.'}), 403
        abort(403)
    if not _inv_editable_for_user(inv):
        if request.method == 'POST':
            if request.is_json: return jsonify({'success': False, 'error': 'This invoice cannot be edited because it has already been finalized.'}), 400
            flash('This invoice cannot be edited because it has already been finalized.', 'danger')
            return redirect(url_for('view_invoice', iid=inv.id))
        flash('This invoice cannot be edited because it has already been finalized.', 'danger')
        return redirect(url_for('view_invoice', iid=inv.id))
    if request.method == 'POST':
        old_gt = inv.grand_total or Decimal('0')
        old_amount_paid = inv.amount_paid or Decimal('0')
        old_customer_id = inv.customer_id
        old_items = [(it.product_id, it.qty) for it in inv.items if it.product_id]

        data = request.get_json(silent=True) or request.form
        items_data = data.get('items', [])
        if not items_data:
            pn = data.getlist('product_name[]') if hasattr(data, 'getlist') else data.get('product_name', [])
            if isinstance(pn, str): pn = [pn]
            pids = data.getlist('product_id[]') if hasattr(data, 'getlist') else data.get('product_id', [])
            qtys = data.getlist('qty[]') if hasattr(data, 'getlist') else data.get('qty', [])
            rates = data.getlist('rate[]') if hasattr(data, 'getlist') else data.get('rate', [])
            discs = data.getlist('discount[]') if hasattr(data, 'getlist') else data.get('discount', [])
            gst_r = data.getlist('gst_rate[]') if hasattr(data, 'getlist') else data.get('gst_rate', [])
            hsns = data.getlist('hsn[]') if hasattr(data, 'getlist') else data.get('hsn', [])
            units = data.getlist('unit[]') if hasattr(data, 'getlist') else data.get('unit', [])
            items_data = []
            for i in range(len(pn) if isinstance(pn, (list, tuple)) else 0):
                if not pn[i].strip(): continue
                items_data.append({'product_name': pn[i], 'product_id': int(pids[i]) if pids and i < len(pids) and pids[i] else None, 'qty': int(qtys[i]) if qtys and i < len(qtys) else 1, 'price': Decimal(rates[i]) if rates and i < len(rates) else 0, 'discount': Decimal(discs[i]) if discs and i < len(discs) else 0, 'gst_rate': Decimal(gst_r[i]) if gst_r and i < len(gst_r) else 18, 'hsn': hsns[i] if hsns and i < len(hsns) else '', 'unit': units[i] if units and i < len(units) else 'pcs'})

        valid_items = [it for it in items_data if it.get('product_name', '').strip()]
        if not valid_items:
            if request.is_json: return jsonify({'success': False, 'error': 'Invoice must have at least one item.'}), 400
            flash('Invoice must have at least one item.', 'danger')
            return redirect(url_for('edit_invoice', iid=inv.id))

        # Inventory safety: account for old quantity -> new quantity (net delta).
        old_map = {}
        for pid, qty in old_items:
            old_map[pid] = old_map.get(pid, 0) + qty
        new_map = {}
        for item in valid_items:
            pid = to_int(item.get('product_id'))
            try:
                qty = int(item.get('qty', 1))
            except (TypeError, ValueError):
                qty = 1
            if not pid or qty <= 0:
                continue
            new_map[pid] = new_map.get(pid, 0) + qty

        allow_oos = str(data.get('allow_out_of_stock', '') or '').lower() in ('1', 'true', 'yes', 'on')
        stock_errors = []
        for item in valid_items:
            pid = to_int(item.get('product_id'))
            if not pid: continue
            try:
                qty = int(item.get('qty', 1))
            except (TypeError, ValueError):
                qty = 1
            if qty <= 0:
                prod = db.session.get(Product, pid)
                stock_errors.append(f'{prod.name if prod else "Item"}: quantity must be greater than zero.')
        for pid in sorted(set(old_map) | set(new_map)):
            delta = new_map.get(pid, 0) - old_map.get(pid, 0)
            if delta <= 0:
                continue
            prod = db.session.get(Product, pid)
            if prod and not allow_oos and delta > prod.current_stock:
                stock_errors.append(f'{prod.name}: insufficient stock. Only {prod.current_stock} unit(s) available.')
        if stock_errors:
            msg = ' '.join(stock_errors)
            if request.is_json: return jsonify({'success': False, 'error': msg}), 400
            flash(msg, 'danger')
            return redirect(url_for('edit_invoice', iid=inv.id))

        # Recalculate every financial value on the SERVER (browser totals ignored).
        comp_sc = current_app.config.get('COMPANY_STATE_CODE', 29)
        try:
            cust_sc = int(data.get('customer_state_code') or inv.customer_state_code or comp_sc)
        except (TypeError, ValueError):
            cust_sc = comp_sc
        intra = comp_sc == cust_sc
        sub = Decimal('0'); td = Decimal('0'); tt = Decimal('0')
        tc = Decimal('0'); ts = Decimal('0'); ti = Decimal('0')
        new_line_items = []
        for item in valid_items:
            qty = int(item.get('qty', 1)); rate = Decimal(str(item.get('price', '0')))
            dp = Decimal(str(item.get('discount', '0'))); gr = Decimal(str(item.get('gst_rate', '18')))
            pid = to_int(item.get('product_id'))
            ls = rate * qty
            da = (ls * dp / Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            tax = ls - da
            gst = GSTService.calculate_gst(tax, gr, intra)
            lt = tax + gst['total_tax']
            new_line_items.append({
                'product_id': pid, 'product_name': item['product_name'].strip(),
                'hsn': item.get('hsn', ''), 'qty': qty, 'unit': item.get('unit', 'pcs'),
                'price': rate, 'discount': dp, 'gst_rate': gr, 'taxable_value': tax,
                'cgst': gst['cgst'], 'sgst': gst['sgst'], 'igst': gst['igst'], 'total': lt,
            })
            sub += ls; td += da; tt += tax; tc += gst['cgst']; ts += gst['sgst']; ti += gst['igst']
        gt = tt + tc + ts + ti
        rg = gt.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        ro = rg - gt

        # Payment: read from submitted data and sync payment record.
        submitted_paid = Decimal(str(data.get('amount_paid') or '0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if submitted_paid < Decimal('0'):
            submitted_paid = Decimal('0')
        if submitted_paid > rg + Decimal('0.005'):
            err = (f'Amount paid ({format_indian_currency(submitted_paid)}) exceeds the invoice total '
                   f'({format_indian_currency(rg)}). Please correct the paid amount.')
            if request.is_json: return jsonify({'success': False, 'error': err}), 400
            flash(err, 'danger')
            return redirect(url_for('edit_invoice', iid=inv.id))
        submitted_ref = (data.get('payment_reference') or '').strip()

        # ---- all validations passed: apply changes ----
        cid = to_int(data.get('customer_id'))
        if cid:
            c = db.session.get(Customer, cid)
            if c:
                inv.customer_id = c.id
                inv.customer_name = c.name
                inv.customer_mobile = c.mobile or ''
                inv.customer_email = c.email or ''
                inv.customer_address = c.address or ''
                inv.customer_gstin = c.gstin or ''
                inv.customer_state = c.state or ''
                inv.customer_state_code = c.state_code or cust_sc
        else:
            inv.customer_name = (data.get('customer_name') or inv.customer_name or '').strip()
            inv.customer_mobile = (data.get('customer_mobile') or inv.customer_mobile or '').strip()
            inv.customer_email = (data.get('customer_email') or inv.customer_email or '').strip()
            inv.customer_address = (data.get('customer_address') or inv.customer_address or '').strip()
            inv.customer_gstin = (data.get('customer_gstin') or inv.customer_gstin or '').strip()
            inv.customer_state = (data.get('customer_state') or inv.customer_state or '').strip()
            inv.customer_state_code = cust_sc
        inv.invoice_date = datetime.strptime(data.get('invoice_date', '') or inv.invoice_date.strftime('%Y-%m-%d'), '%Y-%m-%d').date()
        inv.due_date = (datetime.strptime(data.get('due_date', ''), '%Y-%m-%d').date() if data.get('due_date')
                        else None if 'due_date' in data else inv.due_date)
        inv.payment_method = (data.get('payment_method') or inv.payment_method or '').strip()
        inv.notes = (data.get('notes') if 'notes' in data else inv.notes) or ''
        inv.terms = (data.get('terms') if 'terms' in data else inv.terms) or ''
        inv.updated_at = datetime.utcnow()

        # inventory movements: old -> new
        for pid in sorted(set(old_map) | set(new_map)):
            delta = new_map.get(pid, 0) - old_map.get(pid, 0)
            if delta == 0: continue
            prod = db.session.get(Product, pid)
            if not prod: continue
            if delta > 0:
                prod.stock_quantity = (prod.stock_quantity or 0) - delta
                prod.last_sale = datetime.utcnow()
                _record_movement(pid, 'sale', -delta, reference_type='invoice', reference_id=inv.id,
                                 notes=f'Edited {inv.invoice_number}')
            else:
                prod.stock_quantity = (prod.stock_quantity or 0) - delta
                _record_movement(pid, 'return', -delta, reference_type='invoice', reference_id=inv.id,
                                 notes=f'Edited {inv.invoice_number}')

        for it in inv.items:
            db.session.delete(it)
        for li in new_line_items:
            db.session.add(InvoiceItem(invoice_id=inv.id, **li))
        inv.subtotal = sub; inv.total_discount = td; inv.total_taxable = tt
        inv.total_cgst = tc; inv.total_sgst = ts; inv.total_igst = ti
        inv.round_off = ro; inv.grand_total = rg; inv.is_intra_state = intra
        _set_payment_state(inv, submitted_paid)

        # keep customer purchase aggregates accurate
        if inv.customer_id and inv.customer_id == old_customer_id:
            client = db.session.get(Customer, inv.customer_id)
            if client:
                client.total_purchases = max(Decimal('0'), (client.total_purchases or Decimal('0')) - old_gt + rg)
        elif inv.customer_id:
            if old_customer_id:
                old_c = db.session.get(Customer, old_customer_id)
                if old_c:
                    old_c.total_purchases = max(Decimal('0'), (old_c.total_purchases or Decimal('0')) - old_gt)
                    old_c.invoice_count = max(0, (old_c.invoice_count or 0) - 1)
            new_c = db.session.get(Customer, inv.customer_id)
            if new_c:
                new_c.total_purchases = (new_c.total_purchases or Decimal('0')) + rg
                new_c.invoice_count = (new_c.invoice_count or 0) + 1

        db.session.commit()

        # Sync payment record: create/update to match the edited amount_paid.
        old_total_paid = Decimal(str(old_amount_paid or '0')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        if submitted_paid != old_total_paid:
            existing_payments = _payments_newest(inv)
            if submitted_paid > Decimal('0'):
                if existing_payments:
                    # Update the most recent payment to reflect the new total paid.
                    mp = existing_payments[0]
                    mp.amount = submitted_paid
                    mp.payment_method = inv.payment_method or 'cash'
                    if submitted_ref:
                        mp.reference_number = submitted_ref
                    mp.updated_at = datetime.utcnow()
                else:
                    # No previous payments — create a new one.
                    new_pay = Payment(
                        invoice_id=inv.id,
                        customer_id=inv.customer_id,
                        payment_date=inv.invoice_date or date.today(),
                        amount=submitted_paid,
                        payment_method=inv.payment_method or 'cash',
                        reference_number=submitted_ref or None,
                        received_by=current_user.full_name if hasattr(current_user, 'full_name') else current_user.username,
                    )
                    db.session.add(new_pay)
                db.session.commit()

        _run_low_stock_check()
        log_audit(current_user.id, 'invoice_updated', 'invoice', inv.id,
                  f'Edited {inv.invoice_number} (Total: {old_gt} -> {rg}, items: {len(old_items)} -> {len(valid_items)})')
        if request.is_json:
            return jsonify({'success': True, 'message': 'Invoice updated successfully.',
                            'redirect': url_for('view_invoice', iid=inv.id)})
        flash('Invoice updated.', 'success')
        return redirect(url_for('view_invoice', iid=inv.id))

    invoice_data = {
        'customer_id': inv.customer_id,
        'customer_name': inv.customer_name or '',
        'customer_mobile': inv.customer_mobile or '',
        'customer_email': inv.customer_email or '',
        'customer_address': inv.customer_address or '',
        'customer_gstin': inv.customer_gstin or '',
        'customer_state': inv.customer_state or '',
        'customer_state_code': inv.customer_state_code or '',
        'invoice_date': inv.invoice_date.isoformat() if inv.invoice_date else date.today().isoformat(),
        'due_date': inv.due_date.isoformat() if inv.due_date else '',
        'payment_method': inv.payment_method or 'cash',
        'amount_paid': float(inv.amount_paid or 0),
        'payment_reference': _payment_display_ref(inv.payments[0]) if inv.payments else '',
        'notes': inv.notes or '',
        'terms': inv.terms or '',
        'items': [{
            'product_id': it.product_id, 'product_name': it.product_name,
            'hsn': it.hsn or '', 'qty': it.qty, 'unit': it.unit or 'pcs',
            'price': float(it.price or 0), 'discount': float(it.discount or 0),
            'gst_rate': float(it.gst_rate or 0),
        } for it in inv.items],
    }
    return render_template('billing/new_invoice.html', invoice=inv, invoice_number=inv.invoice_number,
                           invoice_data=invoice_data, customers=Customer.query.order_by(Customer.name).all(),
                           products=Product.query.filter_by(is_active=True).order_by(Product.name).all(),
                           edit_mode=True, company=current_app.config.get('COMPANY', {}),
                           today=date.today().isoformat())


@app.route('/invoices/<int:iid>/delete', methods=['POST'])
@login_required
def delete_invoice(iid):
    inv = db.session.get(Invoice, iid)
    if not inv:
        if request.is_json: return jsonify({'success': False, 'error': 'Invoice not found.'})
        abort(404)
    if not current_user.is_admin:
        if request.is_json: return jsonify({'success': False, 'error': 'Only administrators can delete invoices.'})
        abort(403)
    data = request.get_json(silent=True) or request.form
    pw = (data.get('password') or '').strip()
    if not current_user.check_password(pw):
        if request.is_json: return jsonify({'success': False, 'error': 'Incorrect administrator password.'})
        flash('Incorrect administrator password.', 'danger'); return redirect(url_for('invoice_history'))
    if inv.status == 'cancelled':
        if request.is_json: return jsonify({'success': False, 'error': 'Cancelled invoices cannot be deleted.'})
        flash('Cancelled invoices cannot be deleted.', 'danger'); return redirect(url_for('invoice_history'))
    num = inv.invoice_number
    client = None
    if inv.customer_id:
        client = db.session.get(Customer, inv.customer_id)
    for item in inv.items:
        if item.product_id:
            prod = db.session.get(Product, item.product_id)
            if prod:
                prod.stock_quantity = (prod.stock_quantity or 0) + item.qty
                _record_movement(item.product_id, 'return', item.qty,
                                 reference_type='invoice_delete', reference_id=inv.id,
                                 notes=f'Deleted {num}')
    if client:
        client.total_purchases = max(Decimal('0'), (client.total_purchases or Decimal('0')) - (inv.grand_total or 0))
        client.invoice_count = max(0, (client.invoice_count or 0) - 1)
    db.session.delete(inv)
    db.session.commit()
    log_audit(current_user.id, 'invoice_deleted', 'invoice', iid, f'Deleted {num} by admin')
    if request.is_json: return jsonify({'success': True, 'message': f'Invoice {num} deleted.'})
    flash(f'Invoice {num} deleted.', 'success')
    return redirect(url_for('invoice_history'))


@app.route('/invoices/<int:inv_id>/pdf/<copy_type>')
@login_required
def download_invoice_pdf(inv_id, copy_type):
    inv = db.session.get(Invoice, inv_id)
    if not inv: abort(404)
    if not _inv_accessible(inv): abort(403)
    buf = generate_invoice_pdf(inv, copy_type=copy_type, co=current_app.config.get('COMPANY', {}))
    return send_file(buf, mimetype='application/pdf', download_name=f"{inv.invoice_number}_{copy_type}.pdf", as_attachment=True)


@app.route('/invoices/<int:inv_id>/pdf/all')
@login_required
def download_all_pdfs(inv_id):
    import zipfile
    inv = db.session.get(Invoice, inv_id)
    if not inv: abort(404)
    if not _inv_accessible(inv): abort(403)
    co = current_app.config.get('COMPANY', {})
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for ct in ['customer', 'owner', 'gst']:
            pdf_buf = generate_invoice_pdf(inv, copy_type=ct, co=co)
            zf.writestr(f"{inv.invoice_number}_{ct}.pdf", pdf_buf.read())
    buf.seek(0)
    return send_file(buf, mimetype='application/zip', download_name=f"{inv.invoice_number}_all_copies.zip", as_attachment=True)


@app.route('/invoices/<int:inv_id>/print/<copy_type>')
@login_required
def print_invoice_pdf(inv_id, copy_type):
    inv = db.session.get(Invoice, inv_id)
    if not inv: abort(404)
    if not _inv_accessible(inv): abort(403)
    return render_template('billing/print_invoice.html', invoice=inv, copy_type=copy_type, copy_label=_COPY_LABELS.get(copy_type, 'CUSTOMER COPY'), amount_words=amount_to_words(inv.grand_total), company=current_app.config.get('COMPANY', {}))


@app.route('/invoices/<int:iid>/email', methods=['POST'])
@login_required
def email_invoice(iid):
    inv = db.session.get(Invoice, iid)
    if not inv: abort(404)
    if not _inv_accessible(inv):
        if request.is_json: return jsonify({'success': False, 'error': 'You cannot share another user\'s invoice.'})
        abort(403)

    data = request.get_json(silent=True) or {}
    recipient = data.get('email', '').strip() or inv.customer_email
    if not recipient:
        if request.is_json: return jsonify({'success': False, 'error': 'No email address provided.'})
        flash('No email address.', 'danger'); return redirect(url_for('view_invoice', iid=iid))

    co = current_app.config.get('COMPANY', {})
    ok, msg = send_invoice_email(inv, co, recipient=recipient)

    if ok:
        log_audit(current_user.id, 'email_invoice', 'invoice', inv.id, f'Sent to {recipient}')
        if request.is_json: return jsonify({'success': True, 'message': msg})
        flash(msg, 'success')
    else:
        if request.is_json: return jsonify({'success': False, 'error': msg})
        flash(msg, 'danger')
    return redirect(url_for('view_invoice', iid=iid))


def _is_ajax():
    """True when the request expects a JSON response (AJAX or fetch API)."""
    return bool(request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest')


def _payment_payload(data):
    """Extract a payment field dict from a JSON or form payload (shared by add/edit)."""
    def _get(*keys, default=''):
        for k in keys:
            if data.get(k):
                return str(data[k]).strip()
        return default
    method = _get('payment_method', 'method').lower()
    if method not in PAYMENT_METHODS:
        method = 'cash'
    date_str = _get('payment_date', 'date')
    pdate = date.today()
    if date_str:
        try:
            pdate = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            pdate = date.today()
    return {
        'amount': Decimal(str(data.get('amount', '0') or '0')),
        'payment_method': method,
        'payment_date': pdate,
        'reference_number': _get('reference_number', 'reference', 'payment_reference'),
        'transaction_id': _get('transaction_id'),
        'utr': _get('utr'),
        'remarks': _get('remarks', 'notes'),
        'received_by': _get('received_by', 'received'),
    }


def _payment_json(p, max_amount=None):
    """Serialize a Payment for the JSON APIs."""
    return {
        'id': p.id,
        'invoice_id': p.invoice_id,
        'amount': float(p.amount or 0),
        'payment_method': p.payment_method or 'cash',
        'method_label': _payment_method_label(p.payment_method or 'cash'),
        'payment_date': p.payment_date.strftime('%Y-%m-%d') if p.payment_date else None,
        'reference_number': p.reference_number or '',
        'transaction_id': p.transaction_id or '',
        'utr': p.utr or '',
        'remarks': p.remarks or '',
        'received_by': p.received_by or '',
        'display_ref': _payment_display_ref(p),
        'max_amount': float(max_amount) if max_amount is not None else float(p.amount or 0),
    }


@app.route('/api/invoices/<int:iid>/payment', methods=['POST'])
@login_required
def add_payment(iid):
    inv = db.session.get(Invoice, iid)
    if not inv:
        if request.is_json: return jsonify({'success': False, 'error': 'Invoice not found.'})
        abort(404)
    if not _inv_accessible(inv):
        if request.is_json: return jsonify({'success': False, 'error': 'You cannot record a payment on another user\'s invoice.'})
        abort(403)
    if inv.status == 'cancelled':
        if request.is_json: return jsonify({'success': False, 'error': 'Cancelled invoices cannot receive payments.'})
        flash('Cancelled invoices cannot receive payments.', 'danger')
        return redirect(url_for('view_invoice', iid=iid))

    data = request.get_json(silent=True) or request.form
    pay = _payment_payload(data)
    balance = _payment_balance(inv)

    if pay['amount'] <= 0:
        if request.is_json: return jsonify({'success': False, 'error': 'Payment amount must be greater than zero.'})
        flash('Payment amount must be greater than zero.', 'danger')
        return redirect(url_for('view_invoice', iid=iid))
    if pay['amount'] > balance + Decimal('0.005'):
        if request.is_json: return jsonify({'success': False, 'error': f'Payment exceeds outstanding balance of {format_indian_currency(balance)}.'})
        flash(f'Payment exceeds outstanding balance of {format_indian_currency(balance)}.', 'danger')
        return redirect(url_for('view_invoice', iid=iid))

    p = Payment(invoice_id=inv.id, customer_id=inv.customer_id, amount=pay['amount'],
                payment_method=pay['payment_method'], payment_date=pay['payment_date'],
                reference_number=pay['reference_number'], transaction_id=pay['transaction_id'],
                utr=pay['utr'],
                remarks=pay['remarks'], notes=pay['remarks'], received_by=pay['received_by'] or current_user.full_name)
    db.session.add(p)
    new_paid = (inv.amount_paid or Decimal('0')) + pay['amount']
    _set_payment_state(inv, new_paid)
    inv.payment_method = pay['payment_method']
    db.session.flush()
    db.session.commit()
    log_audit(current_user.id, 'Payment Added', 'payment', p.id,
              f'{inv.invoice_number} | {format_indian_currency(pay["amount"])} via {_payment_method_label(pay["payment_method"])}',
              ip_address=request.remote_addr)

    flash(f'Payment of {format_indian_currency(pay["amount"])} recorded.', 'success')
    if request.is_json:
        resp = jsonify({'success': True, 'message': 'Payment recorded successfully.',
                        'amount_paid': float(inv.amount_paid), 'balance': float(inv.balance_due),
                        'balance_due': float(inv.balance_due),
                        'payment_status': inv.payment_status,
                        'status_label': _payment_status_label(inv.payment_status),
                        'payment': _payment_json(p)})
    else:
        resp = redirect(url_for('view_invoice', iid=iid))

    if inv.payment_status == 'paid' and inv.customer_email:
        @after_this_request
        def _post_payment_email(response):
            try:
                co = current_app.config.get('COMPANY', {})
                ok, emsg = send_invoice_email(inv, co)
                if ok:
                    log_audit(current_user.id, 'email_invoice', 'invoice', inv.id, f'Payment receipt auto-sent to {inv.customer_email}')
                else:
                    log_audit(current_user.id, 'email_failed', 'invoice', inv.id, emsg)
            except Exception as e:
                log_audit(current_user.id, 'email_failed', 'invoice', inv.id, str(e))
            return response

    return resp


@app.route('/api/payments/<int:pid>/edit', methods=['POST'])
@login_required
def edit_payment(pid):
    if not current_user.is_admin:
        if request.is_json: return jsonify({'success': False, 'error': 'Only administrators can edit payments.'})
        abort(403)
    p = db.session.get(Payment, pid)
    if not p:
        if request.is_json: return jsonify({'success': False, 'error': 'Payment not found.'})
        abort(404)
    inv = db.session.get(Invoice, p.invoice_id)
    if not inv:
        if request.is_json: return jsonify({'success': False, 'error': 'Invoice not found.'})
        abort(404)

    data = request.get_json(silent=True) or request.form
    pay = _payment_payload(data)
    if pay['amount'] <= 0:
        if request.is_json: return jsonify({'success': False, 'error': 'Payment amount must be greater than zero.'})
        flash('Payment amount must be greater than zero.', 'danger')
        return redirect(url_for('view_invoice', iid=inv.id))

    other_paid = (inv.amount_paid or Decimal('0')) - (p.amount or Decimal('0'))
    max_allowed = (inv.grand_total or Decimal('0')) - other_paid
    if pay['amount'] > max_allowed + Decimal('0.005'):
        if request.is_json: return jsonify({'success': False, 'error': f'Payment exceeds outstanding balance of {format_indian_currency(max_allowed)}.'})
        flash(f'Payment exceeds outstanding balance of {format_indian_currency(max_allowed)}.', 'danger')
        return redirect(url_for('view_invoice', iid=inv.id))

    old_amt = p.amount
    p.amount = pay['amount']; p.payment_method = pay['payment_method']; p.payment_date = pay['payment_date']
    p.reference_number = pay['reference_number']; p.transaction_id = pay['transaction_id']; p.utr = pay['utr']
    p.remarks = pay['remarks']; p.notes = pay['remarks']
    p.received_by = pay['received_by'] or current_user.full_name
    p.updated_at = datetime.utcnow()
    _set_payment_state(inv, other_paid + pay['amount'])
    inv.payment_method = pay['payment_method']
    db.session.commit()
    log_audit(current_user.id, 'Payment Edited', 'payment', p.id,
              f'{inv.invoice_number} | {format_indian_currency(old_amt)} -> {format_indian_currency(pay["amount"])}',
              ip_address=request.remote_addr)
    if request.is_json:
        return jsonify({'success': True, 'message': 'Payment updated successfully.',
                        'amount_paid': float(inv.amount_paid), 'balance': float(inv.balance_due),
                        'balance_due': float(inv.balance_due),
                        'payment_status': inv.payment_status,
                        'status_label': _payment_status_label(inv.payment_status),
                        'payment': _payment_json(p, max_amount=max_allowed)})
    flash('Payment updated.', 'success')
    return redirect(url_for('view_invoice', iid=inv.id))


@app.route('/api/payments/<int:pid>/delete', methods=['POST'])
@login_required
def delete_payment(pid):
    if not current_user.is_admin:
        if _is_ajax(): return jsonify({'success': False, 'error': 'Only administrators can delete payments.'})
        abort(403)
    p = db.session.get(Payment, pid)
    if not p:
        if _is_ajax(): return jsonify({'success': False, 'error': 'Payment not found.'})
        abort(404)
    inv = db.session.get(Invoice, p.invoice_id)
    num = inv.invoice_number if inv else '?'
    amt = p.amount
    if inv:
        new_paid = max(Decimal('0'), (inv.amount_paid or Decimal('0')) - (p.amount or Decimal('0')))
        _set_payment_state(inv, new_paid)
    db.session.delete(p)
    db.session.commit()
    log_audit(current_user.id, 'Payment Deleted', 'payment', pid,
              f'{num} | {format_indian_currency(amt)}',
              ip_address=request.remote_addr)
    if _is_ajax():
        return jsonify({'success': True, 'message': 'Payment deleted successfully.',
                        'amount_paid': float(inv.amount_paid) if inv else 0,
                        'balance': float(inv.balance_due) if inv else 0,
                        'balance_due': float(inv.balance_due) if inv else 0,
                        'payment_status': inv.payment_status if inv else 'due',
                        'status_label': _payment_status_label(inv.payment_status) if inv else 'Due'})
    flash('Payment deleted.', 'success')
    return redirect(url_for('view_invoice', iid=inv.id if inv else 0))


@app.route('/api/payments/<int:pid>/get', methods=['GET'])
@login_required
def get_payment(pid):
    """Fetch a single payment (admin-only) for the edit form."""
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Only administrators can edit payments.'}), 403
    p = db.session.get(Payment, pid)
    if not p:
        return jsonify({'success': False, 'error': 'Payment not found.'}), 404
    inv = db.session.get(Invoice, p.invoice_id)
    other_paid = (inv.amount_paid - p.amount) if inv else Decimal('0')
    max_allowed = (inv.grand_total - other_paid) if inv else p.amount
    return jsonify({'success': True, 'payment': _payment_json(p, max_amount=max_allowed)})


@app.route('/api/invoices/<int:iid>/cancel', methods=['POST'])
@login_required
def cancel_invoice(iid):
    inv = db.session.get(Invoice, iid)
    if not inv:
        if request.is_json: return jsonify({'success': False, 'error': 'Invoice not found.'})
        abort(404)
    if not _inv_accessible(inv):
        if request.is_json: return jsonify({'success': False, 'error': 'You cannot cancel another user\'s invoice.'})
        abort(403)
    if inv.status == 'cancelled':
        return jsonify({'success': False, 'error': 'Invoice already cancelled.'})
    data = request.get_json(silent=True) or request.form
    reason = (data.get('reason') or '').strip()
    inv.status = 'cancelled'
    inv.payment_status = 'cancelled'
    inv.balance_due = inv.grand_total
    for item in inv.items:
        if item.product_id:
            prod = db.session.get(Product, item.product_id)
            if prod:
                prod.stock_quantity = (prod.stock_quantity or 0) + item.qty
                _record_movement(item.product_id, 'return', item.qty,
                                 reference_type='invoice_cancel', reference_id=inv.id,
                                 notes=f'Cancelled {inv.invoice_number}')
    if inv.customer_id:
        client = db.session.get(Customer, inv.customer_id)
        if client:
            client.total_purchases = max(Decimal('0'), (client.total_purchases or Decimal('0')) - (inv.grand_total or 0))
            client.invoice_count = max(0, (client.invoice_count or 0) - 1)
    db.session.commit()
    log_audit(current_user.id, 'invoice_cancelled', 'invoice', iid, f'Cancelled {inv.invoice_number}. Reason: {reason or "unspecified"}')
    if request.is_json: return jsonify({'success': True, 'message': f'Invoice {inv.invoice_number} cancelled.'})
    flash(f'Invoice {inv.invoice_number} cancelled.', 'success')
    return redirect(url_for('view_invoice', iid=iid))


# ----- Quotations -----

@app.route('/quotations')
@login_required
def quotations_list():
    _expire_overdue_quotations()
    return render_template('quotations/quotations.html', quotations=Quotation.query.order_by(desc(Quotation.quotation_date)).all())


@app.route('/quotations/create', methods=['GET', 'POST'])
@login_required
def new_quotation():
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form
        qn = generate_quotation_number(); cust = _resolve_customer(data)
        comp_sc = current_app.config.get('COMPANY_STATE_CODE', 29); cust_sc = int(data.get('customer_state_code', comp_sc)); intra = comp_sc == cust_sc
        qd_str = data.get('quotation_date', '')
        qd = datetime.strptime(qd_str, '%Y-%m-%d').date() if qd_str else date.today()
        vu_str = data.get('valid_until', '')
        vu = datetime.strptime(vu_str, '%Y-%m-%d').date() if vu_str else None
        qt = Quotation(quotation_number=qn, customer_id=cust.id, customer_name=cust.name, customer_mobile=cust.mobile or '', customer_email=cust.email or '', customer_address=cust.address or '', customer_gstin=cust.gstin or '', customer_state=cust.state or '', customer_state_code=cust.state_code or cust_sc, quotation_date=qd, valid_until=vu, notes=data.get('notes', ''), terms=data.get('terms', ''), created_by=current_user.id, is_intra_state=intra)
        db.session.add(qt); db.session.flush()
        items_data = data.get('items', [])
        if not items_data:
            pns = data.getlist('product_name[]') if hasattr(data, 'getlist') else data.get('product_name', [])
            if isinstance(pns, str): pns = [pns]
            pids = data.getlist('product_id[]') if hasattr(data, 'getlist') else data.get('product_id', [])
            qtys = data.getlist('qty[]') if hasattr(data, 'getlist') else data.get('qty', [])
            rates = data.getlist('rate[]') if hasattr(data, 'getlist') else data.get('rate', [])
            discs = data.getlist('discount[]') if hasattr(data, 'getlist') else data.get('discount', [])
            grs = data.getlist('gst_rate[]') if hasattr(data, 'getlist') else data.get('gst_rate', [])
            hsns = data.getlist('hsn[]') if hasattr(data, 'getlist') else data.get('hsn', [])
            items_data = []
            for i in range(len(pns) if isinstance(pns, (list, tuple)) else 0):
                if not pns[i].strip(): continue
                items_data.append({'product_name': pns[i], 'product_id': int(pids[i]) if pids and i < len(pids) and pids[i] else None, 'qty': int(qtys[i]) if qtys and i < len(qtys) else 1, 'price': Decimal(rates[i]) if rates and i < len(rates) else 0, 'discount': Decimal(discs[i]) if discs and i < len(discs) else 0, 'gst_rate': Decimal(grs[i]) if grs and i < len(grs) else 18, 'hsn': hsns[i] if hsns and i < len(hsns) else ''})
        sub=Decimal('0'); td=Decimal('0'); tt=Decimal('0'); tc=Decimal('0'); ts=Decimal('0'); ti=Decimal('0')
        for item in items_data:
            if not item.get('product_name', '').strip(): continue
            qty=int(item.get('qty', 1)); rate=Decimal(str(item.get('price', '0'))); dp=Decimal(str(item.get('discount', '0'))); gr=Decimal(str(item.get('gst_rate', '18'))); pid=to_int(item.get('product_id'))
            ls=rate*qty; da=(ls*dp/Decimal('100')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP); tax=ls-da
            gst=GSTService.calculate_gst(tax, gr, intra); lt=tax+gst['total_tax']
            db.session.add(QuotationItem(quotation_id=qt.id, product_id=pid, product_name=item['product_name'].strip(), hsn=item.get('hsn', ''), qty=qty, price=rate, discount=dp, gst_rate=gr, taxable_value=tax, cgst=gst['cgst'], sgst=gst['sgst'], igst=gst['igst'], total=lt))
            sub+=ls; td+=da; tt+=tax; tc+=gst['cgst']; ts+=gst['sgst']; ti+=gst['igst']
        gt=tt+tc+ts+ti; rg=gt.quantize(Decimal('1'), rounding=ROUND_HALF_UP); ro=rg-gt; qt.subtotal=sub; qt.total_discount=td; qt.total_taxable=tt; qt.total_cgst=tc; qt.total_sgst=ts; qt.total_igst=ti; qt.round_off=ro; qt.grand_total=rg
        db.session.commit(); log_audit(current_user.id, 'quotation_created', 'quotation', qt.id); flash(f'Quotation {qn} created.', 'success')
        if request.is_json: return jsonify({'success': True, 'redirect': url_for('view_quotation', qid=qt.id)})
        return redirect(url_for('view_quotation', qid=qt.id))
    qn = generate_quotation_number()
    return render_template('quotations/new_quotation.html', customers=Customer.query.order_by(Customer.name).all(), products=Product.query.filter_by(is_active=True).order_by(Product.name).all(), company=current_app.config.get('COMPANY', {}), today=date.today().isoformat(), quotation_number=qn)


@app.route('/quotations/<int:qid>')
@login_required
def view_quotation(qid):
    _expire_overdue_quotations()
    q = db.session.get(Quotation, qid)
    if not q: abort(404)
    return render_template('quotations/quotation_preview.html', quotation=q, amount_words=amount_to_words(q.grand_total), company=current_app.config.get('COMPANY', {}))


@app.route('/quotations/<int:qid>/preview')
@login_required
def quotation_preview(qid):
    q = db.session.get(Quotation, qid)
    if not q: abort(404)
    return render_template('quotations/quotation_preview.html', quotation=q, amount_words=amount_to_words(q.grand_total), company=current_app.config.get('COMPANY', {}))


@app.route('/quotations/<int:qid>/convert', methods=['POST'])
@login_required
def convert_quotation_to_invoice(qid):
    qt = db.session.get(Quotation, qid)
    if not qt:
        abort(404)
    if qt.status != 'accepted':
        flash('Only ACCEPTED quotations can be converted to an invoice.', 'danger')
        return redirect(url_for('view_quotation', qid=qid))
    if qt.converted_invoices:
        flash('This quotation has already been converted to an invoice.', 'danger')
        return redirect(url_for('view_quotation', qid=qid))
    for qi in qt.items:
        if qi.product_id:
            prod = db.session.get(Product, qi.product_id)
            if prod and (prod.stock_quantity or 0) < qi.qty:
                flash(f'Insufficient stock for {qi.product_name}. Only {prod.stock_quantity} unit(s) available.', 'danger')
                return redirect(url_for('view_quotation', qid=qid))
    try:
        inv_num = generate_invoice_number(commit=False)
        gt_rounded = qt.grand_total.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        round_off = gt_rounded - qt.grand_total
        inv = Invoice(invoice_number=inv_num, quotation_id=qt.id, customer_id=qt.customer_id, customer_name=qt.customer_name, customer_mobile=qt.customer_mobile, customer_email=qt.customer_email, customer_address=qt.customer_address, customer_gstin=qt.customer_gstin, customer_state=qt.customer_state, customer_state_code=qt.customer_state_code, invoice_date=date.today(), is_intra_state=qt.is_intra_state, subtotal=qt.subtotal, total_discount=qt.total_discount, total_taxable=qt.total_taxable, total_cgst=qt.total_cgst, total_sgst=qt.total_sgst, total_igst=qt.total_igst, round_off=round_off, grand_total=gt_rounded, notes=qt.notes, terms=qt.terms, status='completed', payment_status='due', amount_paid=Decimal('0'), balance_due=gt_rounded, created_by=current_user.id)
        db.session.add(inv); db.session.flush()
        for qi in qt.items:
            db.session.add(InvoiceItem(invoice_id=inv.id, product_id=qi.product_id, product_name=qi.product_name, hsn=qi.hsn, qty=qi.qty, unit=qi.unit, price=qi.price, discount=qi.discount, gst_rate=qi.gst_rate, taxable_value=qi.taxable_value, cgst=qi.cgst, sgst=qi.sgst, igst=qi.igst, total=qi.total))
            if qi.product_id:
                prod = db.session.get(Product, qi.product_id)
                if prod:
                    prod.stock_quantity = max(0, (prod.stock_quantity or 0) - qi.qty)
                    _record_movement(prod.id, 'sale', -qi.qty, 'invoice', inv.id, f'From quotation {qt.quotation_number}')
        if inv.customer_id:
            c = db.session.get(Customer, inv.customer_id)
            if c:
                c.total_purchases = (c.total_purchases or Decimal('0')) + qt.grand_total
                c.invoice_count = (c.invoice_count or 0) + 1
        qt.status = 'converted'
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        flash('This quotation has already been converted. Please refresh.', 'danger')
        return redirect(url_for('view_quotation', qid=qid))
    except Exception as e:
        db.session.rollback()
        log_audit(current_user.id, 'quotation_conversion_failed', 'quotation', qid, str(e))
        flash(f'Conversion failed: {e}', 'danger')
        return redirect(url_for('view_quotation', qid=qid))
    log_audit(current_user.id, 'quotation_converted', 'quotation', qid, f'To: {inv_num}')
    for u in User.query.filter(User.role.in_(['admin', 'manager'])).all():
        db.session.add(Notification(user_id=u.id, title='Quotation converted',
                                    message=f'{qt.quotation_number} was converted to Invoice {inv_num}.',
                                    notification_type='success'))
    db.session.commit()
    _run_low_stock_check()
    flash(f'Converted to Invoice {inv_num}.', 'success')
    return redirect(url_for('view_invoice', iid=inv.id))


def _expire_overdue_quotations():
    """Auto-expire quotations whose validity window has passed."""
    try:
        qts = Quotation.query.filter(Quotation.status.in_(['draft', 'sent']),
                                     Quotation.valid_until.isnot(None),
                                     Quotation.valid_until < date.today()).all()
        for qt in qts:
            qt.status = 'expired'
            log_audit(current_user.id, 'quotation_expired', 'quotation', qt.id, 'Auto-expired past valid_until')
        if qts:
            db.session.commit()
    except Exception:
        db.session.rollback()


@app.route('/quotations/<int:qid>/status', methods=['POST'])
@login_required
def update_quotation_status(qid):
    qt = db.session.get(Quotation, qid)
    if not qt:
        abort(404)
    if qt.status == 'converted':
        flash('Converted quotations cannot change status.', 'danger')
        return redirect(url_for('view_quotation', qid=qid))
    new = request.form.get('status') or (request.get_json(silent=True) or {}).get('status', '')
    if new not in ('draft', 'sent', 'accepted', 'rejected', 'expired'):
        flash('Invalid status.', 'danger')
        return redirect(url_for('view_quotation', qid=qid))
    old = qt.status
    qt.status = new
    db.session.commit()
    log_audit(current_user.id, 'quotation_status_changed', 'quotation', qid, f'{old} -> {new}')
    flash(f'Quotation marked as {new}.', 'success')
    return redirect(url_for('view_quotation', qid=qid))


# ------------------------------------------------------------------
# SHARED REPORT HELPERS
# ------------------------------------------------------------------

_ZERO = Decimal('0')


def _period_dates(period):
    t = date.today()
    if period == 'today': return t, t
    if period == 'yesterday': return t - timedelta(days=1), t - timedelta(days=1)
    if period == 'week': return t - timedelta(days=t.weekday()), t
    if period == 'month': return date(t.year, t.month, 1), t
    if period == 'year': return date(t.year, 1, 1), t
    return None, None


def _report_dates(args, default_period='month'):
    sd = ed = None
    period = (args.get('period') or '').strip()
    if period:
        sd, ed = _period_dates(period)
    s = (args.get('start_date') or '').strip()
    e = (args.get('end_date') or '').strip()
    if s:
        try: sd = datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError: pass
    if e:
        try: ed = datetime.strptime(e, '%Y-%m-%d').date()
        except ValueError: pass
    if sd is None or ed is None:
        d0, d1 = _period_dates(default_period)
        if sd is None: sd = d0
        if ed is None: ed = d1
    return sd, ed


def _invoice_filter_clauses(args, default_period='month', sd=None, ed=None, allow_product=True):
    if sd is None and ed is None:
        sd, ed = _report_dates(args, default_period)
    filters = {'start_date': sd, 'end_date': ed}
    clauses = [Invoice.invoice_date >= sd, Invoice.invoice_date <= ed]

    status = (args.get('status') or '').strip()
    if status:
        filters['status'] = status
        clauses.append(Invoice.status == status)

    pstatus = (args.get('payment_status') or '').strip()
    if pstatus:
        filters['payment_status'] = pstatus
        if pstatus == 'paid':
            clauses.append(Invoice.status != 'cancelled', Invoice.amount_paid >= Invoice.grand_total)
        elif pstatus == 'partial':
            clauses.append(Invoice.status != 'cancelled', Invoice.amount_paid >= Decimal('0.01'), Invoice.amount_paid < Invoice.grand_total)
        elif pstatus in ('pending', 'due'):
            clauses.append(Invoice.status != 'cancelled', Invoice.amount_paid < Decimal('0.01'))
        elif pstatus == 'cancelled':
            clauses.append(Invoice.status == 'cancelled')

    customer_id = to_int(args.get('customer_id'))
    if customer_id:
        clauses.append(Invoice.customer_id == customer_id); filters['customer_id'] = customer_id

    salesperson_id = to_int(args.get('salesperson_id'))
    if salesperson_id:
        clauses.append(Invoice.created_by == salesperson_id); filters['salesperson_id'] = salesperson_id

    method = (args.get('payment_method') or '').strip()
    if method:
        clauses.append(Invoice.payment_method == method); filters['payment_method'] = method

    gst_type = (args.get('gst_type') or '').strip()
    if gst_type in ('intra', 'inter'):
        clauses.append(Invoice.is_intra_state == (gst_type == 'intra')); filters['gst_type'] = gst_type

    if allow_product:
        product_id = to_int(args.get('product_id'))
        if product_id:
            clauses.append(Invoice.id.in_(db.session.query(InvoiceItem.invoice_id).filter(InvoiceItem.product_id == product_id)))
            filters['product_id'] = product_id

    q = (args.get('q') or '').strip()
    if q:
        filters['q'] = q
        like = '%%%s%%' % q
        clauses.append(or_(
            Invoice.invoice_number.ilike(like),
            Invoice.customer_name.ilike(like),
            Invoice.customer_mobile.ilike(like),
            Invoice.customer_gstin.ilike(like),
        ))
    return clauses, filters


def _money(v):
    try: return float(v or 0)
    except (TypeError, ValueError): return 0.0


def _user_name_map():
    return {u.id: u.full_name for u in User.query.all()}


_PAYMENT_HEADERS = ['Date', 'Invoice #', 'Customer', 'Amount', 'Method', 'Reference', 'Collected By', 'Remarks']


def _payment_filter_clauses(args, default_period='month'):
    """Clauses that filter Payment rows by payment_date (not invoice_date)."""
    sd, ed = _report_dates(args, default_period)
    filters = {'start_date': sd, 'end_date': ed}
    clauses = [func.date(Payment.payment_date) >= sd, func.date(Payment.payment_date) <= ed]
    method = (args.get('payment_method') or '').strip()
    if method:
        clauses.append(Payment.payment_method == method); filters['payment_method'] = method
    customer_id = to_int(args.get('customer_id'))
    if customer_id:
        clauses.append(Invoice.customer_id == customer_id); filters['customer_id'] = customer_id
    q = (args.get('q') or '').strip()
    if q:
        filters['q'] = q
        like = '%%%s%%' % q
        clauses.append(or_(
            Invoice.invoice_number.ilike(like),
            Invoice.customer_name.ilike(like),
            Invoice.customer_mobile.ilike(like),
        ))
    return clauses, filters


def _fetch_payments(clauses):
    rows = (db.session.query(Payment, Invoice)
            .join(Invoice, Invoice.id == Payment.invoice_id)
            .filter(*clauses, Invoice.status != 'cancelled')
            .order_by(desc(Payment.payment_date), desc(Payment.id)).all())
    out = []
    for p, inv in rows:
        out.append({
            'date': p.payment_date.isoformat() if p.payment_date else '',
            'invoice_number': inv.invoice_number, 'customer_name': inv.customer_name or '',
            'amount': _money(p.amount), 'method': _payment_method_label(p.payment_method),
            'reference': _payment_display_ref(p), 'received_by': p.received_by or '',
            'remarks': p.remarks or '',
        })
    return out


def _fetch_outstanding():
    invoices = (Invoice.query.filter(Invoice.status != 'cancelled', Invoice.balance_due > 0)
                .order_by(desc(Invoice.balance_due)).all())
    out = []
    for i in invoices:
        out.append({
            'invoice_number': i.invoice_number,
            'invoice_date': i.invoice_date.isoformat() if i.invoice_date else '',
            'due_date': i.due_date.isoformat() if i.due_date else '',
            'customer_name': i.customer_name or '', 'customer_mobile': i.customer_mobile or '',
            'grand_total': _money(i.grand_total), 'amount_paid': _money(i.amount_paid),
            'balance_due': _money(i.balance_due),
            'status': _payment_status_label(i.payment_status),
            'overdue': bool(i.due_date and i.due_date < date.today()),
        })
    return out


def _fetch_low_stock_rows():
    products = (Product.query.filter(Product.is_active == True, Product.min_stock > 0,
                                     Product.stock_quantity <= Product.min_stock)
                .order_by(Product.stock_quantity.asc(), Product.name.asc()).all())
    rows = []
    for p in products:
        rows.append({
            'id': p.id, 'name': p.name, 'sku': p.sku or '', 'brand': p.brand or '',
            'category': p.category.name if p.category else '',
            'stock': p.stock_quantity or 0, 'min_stock': p.min_stock or 0, 'max_stock': p.max_stock or 0,
            'suggested': _low_stock_suggested_qty(p),
            'purchase_price': _money(p.purchase_price),
            'stock_value': _money((p.stock_quantity or 0) * (p.purchase_price or 0)),
            'status': p.stock_status,
        })
    return rows


# ------------------------------------------------------------------
# REPORTS DASHBOARD  (/reports)
# ------------------------------------------------------------------

@app.route('/reports')
@login_required
def reports_page():
    if not current_user.is_admin:
        abort(403)
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    month_start = date(today.year, today.month, 1)
    year_start = date(today.year, 1, 1)

    def _sum(sd, ed):
        return _money(db.session.query(func.coalesce(func.sum(Invoice.grand_total), 0)).filter(
            Invoice.invoice_date >= sd, Invoice.invoice_date <= ed, Invoice.status != 'cancelled').scalar())

    def _cnt(sd, ed):
        return db.session.query(func.count(Invoice.id)).filter(
            Invoice.invoice_date >= sd, Invoice.invoice_date <= ed, Invoice.status != 'cancelled').scalar() or 0

    total_rev = _money(db.session.query(func.coalesce(func.sum(Invoice.grand_total), 0)).filter(Invoice.status != 'cancelled').scalar())
    paid = _money(db.session.query(func.coalesce(func.sum(Invoice.amount_paid), 0)).filter(Invoice.status != 'cancelled').scalar())
    outstanding = _money(db.session.query(func.coalesce(func.sum(Invoice.grand_total - Invoice.amount_paid), 0)).filter(
        Invoice.status != 'cancelled', Invoice.grand_total > Invoice.amount_paid).scalar())
    cancelled = _money(db.session.query(func.coalesce(func.sum(Invoice.grand_total), 0)).filter(Invoice.status == 'cancelled').scalar())

    top_row = db.session.query(InvoiceItem.product_name, func.sum(InvoiceItem.qty)).join(
        Invoice, Invoice.id == InvoiceItem.invoice_id).filter(Invoice.status != 'cancelled').group_by(
        InvoiceItem.product_name).order_by(desc(func.sum(InvoiceItem.qty))).first()

    stats = {
        'today_sales': _sum(today, today), 'week_sales': _sum(week_start, today),
        'month_sales': _sum(month_start, today), 'year_sales': _sum(year_start, today),
        'total_revenue': total_rev, 'paid_amount': paid, 'outstanding': outstanding, 'cancelled_amount': cancelled,
        'today_invoices': _cnt(today, today), 'month_invoices': _cnt(month_start, today),
        'recent_invoice_count': _cnt(today - timedelta(days=30), today),
        'total_customers': Customer.query.count(), 'active_customers': Customer.query.filter(Customer.invoice_count > 0).count(),
        'total_products': Product.query.filter(Product.is_active == True).count(),
        'low_stock_count': Product.query.filter(Product.stock_quantity <= Product.min_stock, Product.stock_quantity > 0, Product.is_active == True).count(),
        'out_of_stock_count': Product.query.filter(Product.stock_quantity == 0, Product.is_active == True).count(),
        'top_product': {'name': top_row[0], 'qty': int(top_row[1] or 0)} if top_row else None,
    }
    stats['recent_invoices'] = Invoice.query.order_by(desc(Invoice.invoice_date), desc(Invoice.id)).limit(8).all()
    return render_template('reports/reports.html', stats=stats, today=today)


# ------------------------------------------------------------------
# SALES REPORT
# ------------------------------------------------------------------

_SALES_HEADERS = ['Invoice #', 'Date', 'Customer', 'Mobile', 'GSTIN', 'Subtotal', 'Discount', 'Taxable',
                  'CGST', 'SGST', 'IGST', 'Round Off', 'Grand Total', 'Amount Paid', 'Balance', 'Status',
                  'Payment Method', 'Payment Status', 'Salesperson']


def _sales_row(inv, user_map):
    return {
        'invoice_number': inv.invoice_number,
        'invoice_date': inv.invoice_date.isoformat() if inv.invoice_date else '',
        'customer_name': inv.customer_name or '', 'customer_mobile': inv.customer_mobile or '',
        'customer_gstin': inv.customer_gstin or '', 'subtotal': _money(inv.subtotal),
        'total_discount': _money(inv.total_discount), 'total_taxable': _money(inv.total_taxable),
        'total_cgst': _money(inv.total_cgst), 'total_sgst': _money(inv.total_sgst),
        'total_igst': _money(inv.total_igst), 'round_off': _money(inv.round_off),
        'grand_total': _money(inv.grand_total), 'amount_paid': _money(inv.amount_paid),
        'balance': _money((inv.grand_total or 0) - (inv.amount_paid or 0)),
        'status': inv.status, 'payment_method': inv.payment_method or '',
        'payment_status': inv.payment_status, 'salesperson': user_map.get(inv.created_by, ''),
    }


@app.route('/reports/sales')
@login_required
def sales_report():
    if not current_user.is_admin:
        abort(403)
    args = request.args
    clauses, filters = _invoice_filter_clauses(args, default_period='month')
    q = Invoice.query.filter(*clauses)
    item_q = db.session.query(InvoiceItem, Invoice).join(Invoice, Invoice.id == InvoiceItem.invoice_id).filter(*clauses)

    agg = q.with_entities(
        func.count(Invoice.id),
        func.coalesce(func.sum(Invoice.grand_total), 0),
        func.coalesce(func.sum(Invoice.total_taxable), 0),
        func.coalesce(func.sum(Invoice.total_cgst + Invoice.total_sgst + Invoice.total_igst), 0),
        func.coalesce(func.sum(Invoice.total_discount), 0),
        func.coalesce(func.sum(Invoice.amount_paid), 0),
        func.coalesce(func.sum(Invoice.grand_total - Invoice.amount_paid), 0)).first()
    count = int(agg[0])
    gross = _money(agg[1]); taxable = _money(agg[2]); tax = _money(agg[3])
    discount = _money(agg[4]); paid = _money(agg[5]); balance = _money(agg[6])
    avg_sale = gross / count if count else 0
    highest = _money(q.with_entities(func.max(Invoice.grand_total)).scalar())
    lowest = _money(q.with_entities(func.min(Invoice.grand_total)).scalar())

    status_data = []
    for r in q.with_entities(Invoice.payment_status, func.count(), func.coalesce(func.sum(Invoice.grand_total), 0)).group_by(Invoice.payment_status).all():
        status_data.append({'status': r[0] or 'na', 'count': int(r[1]), 'total': _money(r[2])})

    daily_rows = q.with_entities(func.date(Invoice.invoice_date).label('d'), func.coalesce(func.sum(Invoice.grand_total), 0), func.count()).group_by('d').order_by('d').all()
    daily_data = {str(r[0]): {'total': round(_money(r[1]), 2), 'count': int(r[2])} for r in daily_rows}

    monthly_rows = q.with_entities(extract('year', Invoice.invoice_date).label('y'), extract('month', Invoice.invoice_date).label('m'),
                                   func.coalesce(func.sum(Invoice.grand_total), 0), func.count()).group_by('y', 'm').order_by('y', 'm').all()
    monthly_data = {("%d-%02d" % (int(r[0]), int(r[1]))): {'total': round(_money(r[2]), 2), 'count': int(r[3])} for r in monthly_rows}

    method_rows = q.with_entities(Invoice.payment_method, func.coalesce(func.sum(Invoice.grand_total), 0), func.count()).group_by(Invoice.payment_method).all()
    method_data = [{'method': r[0] or 'Unknown', 'total': _money(r[1]), 'count': int(r[2])} for r in method_rows]

    top_customers = q.with_entities(Invoice.customer_name, func.coalesce(func.sum(Invoice.grand_total), 0), func.count()).group_by(
        Invoice.customer_name).order_by(desc(func.sum(Invoice.grand_total))).limit(10).all()
    top_products = item_q.with_entities(InvoiceItem.product_name, func.sum(InvoiceItem.qty), func.coalesce(func.sum(InvoiceItem.total), 0)).group_by(
        InvoiceItem.product_name).order_by(desc(func.sum(InvoiceItem.total))).limit(10).all()

    page = args.get('page', 1, type=int)
    invoices = q.order_by(desc(Invoice.invoice_date), desc(Invoice.id)).all()

    summary = {
        'count': count, 'total': round(gross, 2), 'taxable': round(taxable, 2), 'tax': round(tax, 2),
        'discount': round(discount, 2), 'paid': round(paid, 2), 'balance': round(balance, 2),
        'avg': round(avg_sale, 2), 'highest': round(highest, 2), 'lowest': round(lowest, 2),
    }
    return render_template('reports/sales_report.html', invoices=invoices, summary=summary,
        daily_data=daily_data, monthly_data=monthly_data, status_data=status_data,
        method_data=method_data, top_customers=top_customers, top_products=top_products,
        filters=filters, sales_users=User.query.filter(User.is_active == True).order_by(User.full_name).all(),
        customers=Customer.query.filter(Customer.invoice_count > 0).order_by(Customer.name).all(),
        products=Product.query.filter(Product.is_active == True).order_by(Product.name).all())


# ------------------------------------------------------------------
# COLLECTION REPORT  (/reports/payments)
# ------------------------------------------------------------------

@app.route('/reports/payments')
@login_required
def payments_report():
    if not current_user.is_admin:
        abort(403)
    args = request.args
    clauses, filters = _payment_filter_clauses(args, default_period='month')
    q = (db.session.query(Payment, Invoice)
         .join(Invoice, Invoice.id == Payment.invoice_id)
         .filter(*clauses, Invoice.status != 'cancelled'))
    total_collected = _money(q.with_entities(func.coalesce(func.sum(Payment.amount), 0)).scalar())
    method_rows = (q.with_entities(Payment.payment_method, func.coalesce(func.sum(Payment.amount), 0), func.count())
                   .group_by(Payment.payment_method).all())
    method_data = [{'method': _payment_method_label(m), 'total': _money(t), 'count': int(n)} for m, t, n in method_rows]
    daily_rows = (q.with_entities(func.date(Payment.payment_date).label('d'), func.coalesce(func.sum(Payment.amount), 0), func.count())
                  .group_by('d').order_by('d').all())
    daily_data = [{'date': str(r[0]), 'total': _money(r[1]), 'count': int(r[2])} for r in daily_rows]
    payments = q.order_by(desc(Payment.payment_date), desc(Payment.id)).all()
    return render_template('reports/payments_report.html', payments=payments, total_collected=total_collected,
                           method_data=method_data, daily_data=daily_data, filters=filters,
                           customers=Customer.query.order_by(Customer.name).all())


# ------------------------------------------------------------------
# OUTSTANDING REPORT  (/reports/outstanding)
# ------------------------------------------------------------------

@app.route('/reports/outstanding')
@login_required
def outstanding_report():
    if not current_user.is_admin:
        abort(403)
    args = request.args
    clauses = [Invoice.status != 'cancelled', Invoice.balance_due > 0]
    filters = {}
    customer_id = to_int(args.get('customer_id'))
    if customer_id:
        clauses.append(Invoice.customer_id == customer_id); filters['customer_id'] = customer_id
    invoices = Invoice.query.filter(*clauses).order_by(desc(Invoice.balance_due), desc(Invoice.due_date)).all()
    total_outstanding = _money(db.session.query(func.coalesce(func.sum(Invoice.balance_due), 0)).filter(*clauses).scalar())
    overdue = [i for i in invoices if i.due_date and i.due_date < date.today()]
    overdue_total = _money(sum((i.balance_due or 0) for i in overdue))
    return render_template('reports/outstanding_report.html', invoices=invoices, total_outstanding=total_outstanding,
                           overdue=overdue, overdue_total=overdue_total, filters=filters, today=date.today(),
                           customers=Customer.query.order_by(Customer.name).all())


# ------------------------------------------------------------------
# LOW STOCK REPORT  (/reports/low-stock)
# ------------------------------------------------------------------

@app.route('/reports/low-stock')
@login_required
def low_stock_report():
    if not current_user.is_admin:
        abort(403)
    _run_low_stock_check()
    products = _fetch_low_stock_rows()
    total_value = _money(sum(p['stock_value'] for p in products))
    return render_template('reports/low_stock_report.html', products=products, total_value=total_value,
                           today=date.today())


# ------------------------------------------------------------------
# GST REPORT
# ------------------------------------------------------------------

@app.route('/reports/gst')
@login_required
def gst_report():
    if not current_user.is_admin:
        abort(403)
    now_dt = datetime.now()
    try:
        month = int(request.args.get('month', now_dt.month))
        year = int(request.args.get('year', now_dt.year))
    except (TypeError, ValueError):
        month, year = now_dt.month, now_dt.year
    from calendar import monthrange
    sd = date(year, month, 1)
    ed = date(year, month, monthrange(year, month)[1])

    clauses, filters = _invoice_filter_clauses(request.args, sd=sd, ed=ed)
    q = Invoice.query.filter(*clauses)
    item_q = db.session.query(InvoiceItem, Invoice).join(Invoice, Invoice.id == InvoiceItem.invoice_id).filter(*clauses)

    _cnt, _taxable, _cgst, _sgst, _igst = (_money(v) for v in q.with_entities(
        func.count(Invoice.id),
        func.coalesce(func.sum(Invoice.total_taxable), 0),
        func.coalesce(func.sum(Invoice.total_cgst), 0),
        func.coalesce(func.sum(Invoice.total_sgst), 0),
        func.coalesce(func.sum(Invoice.total_igst), 0)).first())
    inv_count = int(_cnt)
    grand_total = _taxable + _cgst + _sgst + _igst

    hsn_summary = []
    for r in item_q.with_entities(InvoiceItem.hsn, InvoiceItem.gst_rate, func.sum(InvoiceItem.qty),
            func.sum(InvoiceItem.taxable_value), func.sum(InvoiceItem.cgst), func.sum(InvoiceItem.sgst),
            func.sum(InvoiceItem.igst), func.count(func.distinct(Invoice.id))).group_by(
            InvoiceItem.hsn, InvoiceItem.gst_rate).order_by(desc(func.sum(InvoiceItem.taxable_value))).all():
        tv, cg, sg, ig = (_money(v) for v in r[3:7])
        hsn_summary.append({'hsn': r[0] or 'N/A', 'rate': _money(r[1]), 'qty': int(r[2] or 0),
                            'taxable': tv, 'cgst': cg, 'sgst': sg, 'igst': ig,
                            'total': tv + cg + sg + ig, 'invoices': int(r[7] or 0)})

    rate_summary = []
    for r in item_q.with_entities(InvoiceItem.gst_rate, func.sum(InvoiceItem.qty),
            func.sum(InvoiceItem.taxable_value), func.sum(InvoiceItem.cgst), func.sum(InvoiceItem.sgst),
            func.sum(InvoiceItem.igst)).group_by(InvoiceItem.gst_rate).order_by(InvoiceItem.gst_rate).all():
        rate_summary.append({'rate': _money(r[0]), 'qty': int(r[1] or 0), 'taxable': _money(r[2]),
                             'cgst': _money(r[3]), 'sgst': _money(r[4]), 'igst': _money(r[5]),
                             'total': _money(r[2]) + _money(r[3]) + _money(r[4]) + _money(r[5])})

    state_wise = []
    for r in q.with_entities(Invoice.customer_state_code, Invoice.customer_state, func.count(),
            func.coalesce(func.sum(Invoice.total_taxable), 0),
            func.coalesce(func.sum(Invoice.total_cgst + Invoice.total_sgst + Invoice.total_igst), 0)).group_by(
            Invoice.customer_state_code, Invoice.customer_state).all():
        state_wise.append({'code': r[0] or 0, 'name': r[1] or 'Unknown', 'count': int(r[2] or 0),
                           'taxable': _money(r[3]), 'tax': _money(r[4])})

    gstin_wise = []
    for r in q.with_entities(Invoice.customer_gstin, Invoice.customer_name, func.count(),
            func.coalesce(func.sum(Invoice.total_taxable), 0),
            func.coalesce(func.sum(Invoice.total_cgst + Invoice.total_sgst + Invoice.total_igst), 0),
            func.coalesce(func.sum(Invoice.grand_total), 0)).group_by(Invoice.customer_gstin, Invoice.customer_name).all():
        gstin_wise.append({'gstin': r[0] or 'B2C', 'name': r[1] or 'N/A', 'count': int(r[2] or 0),
                           'taxable': _money(r[3]), 'tax': _money(r[4]), 'total': _money(r[5])})

    customer_wise = []
    for r in q.with_entities(Invoice.customer_name, func.count(), func.coalesce(func.sum(Invoice.total_taxable), 0),
            func.coalesce(func.sum(Invoice.total_cgst + Invoice.total_sgst + Invoice.total_igst), 0),
            func.coalesce(func.sum(Invoice.grand_total), 0)).group_by(Invoice.customer_name).order_by(desc(func.sum(Invoice.grand_total))).all():
        customer_wise.append({'name': r[0] or 'N/A', 'count': int(r[1] or 0), 'taxable': _money(r[2]),
                              'tax': _money(r[3]), 'total': _money(r[4])})

    intra_count = q.filter(Invoice.is_intra_state == True).count()
    inter_count = inv_count - intra_count
    b2b_count = q.filter(Invoice.customer_gstin.isnot(None), Invoice.customer_gstin != '').count()
    b2c_count = inv_count - b2b_count

    page = request.args.get('page', 1, type=int)
    invoices = q.order_by(desc(Invoice.invoice_date), desc(Invoice.id)).all()

    return render_template('reports/gst_report.html', invoices=invoices,
        total_taxable=_taxable, total_cgst=_cgst, total_sgst=_sgst, total_igst=_igst,
        grand_total=grand_total, inv_count=inv_count, month=month, year=year, now=now_dt,
        hsn_summary=hsn_summary, rate_summary=rate_summary, state_wise=state_wise,
        gstin_wise=gstin_wise, customer_wise=customer_wise, filters=filters,
        intra_count=intra_count, inter_count=inter_count, b2b_count=b2b_count, b2c_count=b2c_count,
        customers=Customer.query.order_by(Customer.name).all())


# ------------------------------------------------------------------
# INVENTORY REPORT
# ------------------------------------------------------------------

def _inventory_row(p, purch, sold):
    bp = _money(p.purchase_price); sp = _money(p.selling_price)
    stock = p.stock_quantity if p.stock_quantity is not None else 0
    min_s = p.min_stock if p.min_stock is not None else 0
    margin = ((sp - bp) / sp * 100) if sp > 0 else 0
    return {
        'id': p.id, 'name': p.name, 'sku': p.sku or '', 'hsn': p.hsn or '',
        'category': p.category.name if p.category else '', 'brand': p.brand or '',
        'purchase_price': bp, 'selling_price': sp, 'gst_rate': _money(p.gst_rate),
        'opening_stock': p.opening_stock or 0, 'purchased': int(purch or 0), 'sold': int(sold or 0),
        'returned': 0, 'current_stock': stock, 'reserved': 0, 'available': stock,
        'stock_value': stock * bp, 'profit_margin': round(margin, 2), 'min_stock': min_s,
        'low_stock': 0 < stock <= min_s, 'out_of_stock': stock == 0,
    }


@app.route('/reports/inventory')
@login_required
def inventory_report():
    if not current_user.is_admin:
        abort(403)
    args = request.args
    search = (args.get('q') or '').strip()
    sort = (args.get('sort') or 'name').strip()

    q = Product.query.filter(Product.is_active == True)
    if search:
        q = q.filter(or_(Product.name.ilike('%%%s%%' % search), Product.sku.ilike('%%%s%%' % search),
                         Product.hsn.ilike('%%%s%%' % search), Product.brand.ilike('%%%s%%' % search)))

    sold_map = dict(db.session.query(InvoiceItem.product_id, func.sum(InvoiceItem.qty)).join(
        Invoice, Invoice.id == InvoiceItem.invoice_id).filter(Invoice.status != 'cancelled').group_by(InvoiceItem.product_id).all())
    purch_map = dict(db.session.query(PurchaseItem.product_id, func.sum(PurchaseItem.qty)).group_by(PurchaseItem.product_id).all())

    products = [_inventory_row(p, purch_map.get(p.id, 0), sold_map.get(p.id, 0)) for p in q.all()]

    if sort == 'stock':
        products.sort(key=lambda x: x['current_stock'], reverse=True)
    elif sort == 'value':
        products.sort(key=lambda x: x['stock_value'], reverse=True)
    elif sort == 'low':
        products.sort(key=lambda x: (0 if x['out_of_stock'] else (1 if x['low_stock'] else 2)))
    else:
        products.sort(key=lambda x: (x['name'] or '').lower())

    total_value = round(sum(p['stock_value'] for p in products), 2)
    low_stock = [p for p in products if p['low_stock']]
    out_of_stock = [p for p in products if p['out_of_stock']]
    never_sold = [p for p in products if (p['sold'] or 0) == 0]
    total_stock = sum(p['current_stock'] or 0 for p in products)

    cat_rows = db.session.query(Category.name, func.coalesce(func.sum(Product.stock_quantity * Product.purchase_price), 0)).join(
        Product, Product.category_id == Category.id).filter(Product.is_active == True).group_by(Category.name).order_by(desc(func.sum(Product.stock_quantity * Product.purchase_price))).all()
    category_data = [{'name': c[0] or 'Uncategorised', 'value': _money(c[1])} for c in cat_rows]

    return render_template('reports/inventory_report.html', products=products, total_value=total_value,
        low_stock=low_stock, out_of_stock=out_of_stock, never_sold=never_sold,
        total_products=len(products), total_stock=total_stock,
        low_stock_count=len(low_stock), out_of_stock_count=len(out_of_stock),
        category_data=category_data, search=search, sort=sort, top_products=products[:10])


# ------------------------------------------------------------------
# PROFIT REPORT
# ------------------------------------------------------------------

@app.route('/reports/profit')
@login_required
def profit_report():
    if not current_user.is_admin:
        abort(403)
    args = request.args
    clauses, filters = _invoice_filter_clauses(args, default_period='month')
    q = Invoice.query.filter(*clauses)
    item_q = db.session.query(InvoiceItem, Invoice).join(Invoice, Invoice.id == InvoiceItem.invoice_id).filter(*clauses)

    revenue = _money(q.with_entities(func.coalesce(func.sum(Invoice.grand_total), 0)).scalar())
    discount = _money(q.with_entities(func.coalesce(func.sum(Invoice.total_discount), 0)).scalar())
    gst_liability = _money(q.with_entities(func.coalesce(func.sum(Invoice.total_cgst + Invoice.total_sgst + Invoice.total_igst), 0)).scalar())

    prod_map = {p.id: _money(p.purchase_price) for p in db.session.query(Product).all()}
    purchase_cost = 0.0
    for pid, qt in item_q.with_entities(InvoiceItem.product_id, func.sum(InvoiceItem.qty)).group_by(InvoiceItem.product_id).all():
        purchase_cost += (prod_map.get(pid, 0) if pid else 0) * float(qt or 0)

    expenses = _money(get_setting('business_expenses', 0) or 0)
    gross_profit = revenue - purchase_cost
    net_profit = revenue - purchase_cost - gst_liability - discount - expenses

    product_list = []
    for pid, pname, qt, rev in item_q.with_entities(InvoiceItem.product_id, InvoiceItem.product_name,
            func.sum(InvoiceItem.qty), func.coalesce(func.sum(InvoiceItem.total), 0)).group_by(
            InvoiceItem.product_id, InvoiceItem.product_name).all():
        cost_v = (prod_map.get(pid, 0) if pid else 0) * float(qt or 0)
        product_list.append({'name': pname, 'qty': int(qt or 0), 'revenue': _money(rev),
                             'cost': cost_v, 'profit': _money(rev) - cost_v})
    product_list.sort(key=lambda x: x['profit'], reverse=True)
    highest_profit = product_list[0] if product_list else None
    lowest_profit = product_list[-1] if product_list else None

    daily_profit = {}
    for d, rev in q.with_entities(func.date(Invoice.invoice_date).label('d'), func.coalesce(func.sum(Invoice.grand_total), 0)).group_by('d').order_by('d').all():
        daily_profit[str(d)] = round(float(rev), 2)

    monthly_profit = {}
    for r in q.with_entities(extract('year', Invoice.invoice_date).label('y'), extract('month', Invoice.invoice_date).label('m'),
            func.coalesce(func.sum(Invoice.grand_total), 0)).group_by('y', 'm').order_by('y', 'm').all():
        monthly_profit[("%d-%02d" % (int(r[0]), int(r[1])))] = round(float(r[2]), 2)

    category_profit = []
    for r in item_q.join(Product, Product.id == InvoiceItem.product_id).outerjoin(Category, Category.id == Product.category_id).with_entities(
            Category.name, func.sum(InvoiceItem.total), func.sum(InvoiceItem.qty)).group_by(Category.name).all():
        category_profit.append({'name': r[0] or 'Uncategorised', 'revenue': _money(r[1]), 'qty': int(r[2] or 0)})

    customer_profit = []
    for r in q.with_entities(Invoice.customer_name, func.coalesce(func.sum(Invoice.grand_total), 0), func.count()).group_by(
            Invoice.customer_name).order_by(desc(func.sum(Invoice.grand_total))).all():
        customer_profit.append({'name': r[0] or 'N/A', 'revenue': _money(r[1]), 'count': int(r[2] or 0)})
    for cp in customer_profit:
        cp['profit'] = round(cp['revenue'] - (cp['revenue'] * 0.7), 2)

    inv_cost_map = {}
    for iid, pid, qt in db.session.query(InvoiceItem.invoice_id, InvoiceItem.product_id, func.sum(InvoiceItem.qty)).group_by(
            InvoiceItem.invoice_id, InvoiceItem.product_id).all():
        inv_cost_map[iid] = inv_cost_map.get(iid, 0) + (prod_map.get(pid, 0) if pid else 0) * float(qt or 0)

    page = args.get('page', 1, type=int)
    invoices = q.order_by(desc(Invoice.invoice_date), desc(Invoice.id)).all()

    return render_template('reports/profit_report.html', invoices=invoices,
        revenue=revenue, purchase_cost=round(purchase_cost, 2), gst_liability=gst_liability,
        discount=discount, expenses=expenses, gross_profit=round(gross_profit, 2),
        net_profit=round(net_profit, 2), profit_pct=round((net_profit / revenue * 100) if revenue else 0, 2),
        product_list=product_list, highest_profit=highest_profit, lowest_profit=lowest_profit,
        daily_profit=daily_profit, monthly_profit=monthly_profit, category_profit=category_profit,
        customer_profit=customer_profit, invoice_costs=inv_cost_map, filters=filters,
        sales_users=User.query.filter(User.is_active == True).order_by(User.full_name).all(),
        customers=Customer.query.order_by(Customer.name).all())


# ------------------------------------------------------------------
# CUSTOMER REPORT
# ------------------------------------------------------------------

@app.route('/reports/customer')
@login_required
def customer_report():
    if not current_user.is_admin:
        abort(403)
    args = request.args
    clauses, filters = _invoice_filter_clauses(args, default_period='year')
    q = Invoice.query.filter(*clauses)

    rows = q.with_entities(Invoice.customer_id, Invoice.customer_name, Invoice.customer_mobile, Invoice.customer_gstin,
        func.count(Invoice.id), func.coalesce(func.sum(Invoice.grand_total), 0),
        func.coalesce(func.sum(Invoice.grand_total - Invoice.amount_paid), 0),
        func.max(Invoice.invoice_date), func.min(Invoice.invoice_date)).group_by(
        Invoice.customer_id, Invoice.customer_name, Invoice.customer_mobile, Invoice.customer_gstin).all()

    fav_r = {}
    item_q = db.session.query(InvoiceItem, Invoice).join(Invoice, Invoice.id == InvoiceItem.invoice_id).filter(*clauses)
    for cid, pn, qt in item_q.with_entities(Invoice.customer_id, InvoiceItem.product_name, func.sum(InvoiceItem.qty)).group_by(
            Invoice.customer_id, InvoiceItem.product_name).all():
        if cid and pn:
            if cid not in fav_r or qt > fav_r[cid][1]:
                fav_r[cid] = (pn, qt)

    customer_data = []
    for cid, name, mob, gstin, cnt, tot, bal, lastd, firstd in rows:
        cnt = int(cnt or 0)
        if not cnt:
            continue
        customer_data.append({
            'customer': {'id': cid, 'name': name, 'mobile': mob, 'gstin': gstin},
            'count': cnt, 'total': _money(tot), 'outstanding': _money(bal),
            'average': _money(tot) / cnt, 'last_purchase': lastd or '', 'first_purchase': firstd or '',
            'favourite_product': (fav_r.get(cid, (None, 0))[0] or ''),
        })
    customer_data.sort(key=lambda x: x['total'], reverse=True)
    for i, c in enumerate(customer_data, 1):
        c['rank'] = i

    total_purchase = sum(c['total'] for c in customer_data)
    top10 = customer_data[:10]
    high_value = [c for c in customer_data if c['total'] >= 10000]
    cutoff = date.today() - timedelta(days=90)
    inactive = [c for c in customer_data if not c['last_purchase'] or c['last_purchase'] < cutoff]

    payment_rows = q.with_entities(Invoice.payment_method, func.count()).group_by(Invoice.payment_method).all()
    payment_methods = {m[0] or 'Unknown': int(m[1]) for m in payment_rows}

    return render_template('reports/customer_report.html', customer_data=customer_data,
        top10=top10, high_value=high_value, inactive=inactive,
        total_customers=len(customer_data), total_purchase=round(total_purchase, 2),
        top_customer=customer_data[0] if customer_data else None, payment_methods=payment_methods,
        filters=filters, customers=Customer.query.order_by(Customer.name).all())


# ------------------------------------------------------------------
# EXPORTS — CSV / EXCEL / PDF  (all filter-aware, shared dataset builders)
# ------------------------------------------------------------------

def _fetch_sales(clauses, user_map):
    for inv in Invoice.query.filter(*clauses).order_by(desc(Invoice.invoice_date), desc(Invoice.id)).yield_per(500):
        yield _sales_row(inv, user_map)


def _fetch_inventory_rows():
    sold_map = dict(db.session.query(InvoiceItem.product_id, func.sum(InvoiceItem.qty)).join(
        Invoice, Invoice.id == InvoiceItem.invoice_id).filter(Invoice.status != 'cancelled').group_by(InvoiceItem.product_id).all())
    purch_map = dict(db.session.query(PurchaseItem.product_id, func.sum(PurchaseItem.qty)).group_by(PurchaseItem.product_id).all())
    for p in Product.query.filter(Product.is_active == True).order_by(Product.name).all():
        yield _inventory_row(p, purch_map.get(p.id, 0), sold_map.get(p.id, 0))


def _fetch_customer_rows(clauses):
    q = db.session.query(Invoice.customer_id, Invoice.customer_name, Invoice.customer_mobile, Invoice.customer_gstin,
        func.count(Invoice.id), func.coalesce(func.sum(Invoice.grand_total), 0), func.coalesce(func.sum(Invoice.grand_total), 0)).filter(*clauses).group_by(
        Invoice.customer_id, Invoice.customer_name, Invoice.customer_mobile, Invoice.customer_gstin).all()
    out = []
    for cid, name, mob, gstin, cnt, tot, _t2 in q:
        cnt = int(cnt or 0)
        if cnt:
            out.append({'name': name or '', 'mobile': mob or '', 'gstin': gstin or '', 'count': cnt,
                        'total': _money(tot), 'average': _money(tot) / cnt})
    out.sort(key=lambda x: x['total'], reverse=True)
    for i, r in enumerate(out, 1):
        r['rank'] = i
    return out


@app.route('/reports/export/<fmt>/<report_type>')
@login_required
def report_export(fmt, report_type):
    if not current_user.is_admin:
        abort(403)
    if fmt not in ('excel', 'csv', 'pdf'):
        abort(404)
    if report_type == 'payments':
        clauses, filters = _payment_filter_clauses(request.args, default_period='year')
        user_map = _user_name_map()
    elif report_type == 'outstanding':
        clauses, filters = [Invoice.status != 'cancelled', Invoice.balance_due > 0], {}
        user_map = _user_name_map()
    elif report_type == 'low_stock':
        clauses, filters = [], {}
        user_map = {}
    else:
        clauses, filters = _invoice_filter_clauses(request.args, default_period='year', allow_product=False)
        user_map = _user_name_map()
    fname = "%s_report_%s" % (report_type, date.today().strftime('%Y%m%d'))
    if fmt == 'csv':
        return _export_csv(report_type, clauses, user_map, fname)
    if fmt == 'excel':
        return _export_excel(report_type, clauses, user_map, fname)
    return _export_pdf(report_type, clauses, user_map, filters, fname)


def _export_csv(report_type, clauses, user_map, fname):
    out = io.StringIO()
    w = csv.writer(out)
    if report_type == 'sales':
        w.writerow(_SALES_HEADERS)
        for r in _fetch_sales(clauses, user_map):
            w.writerow([r['invoice_number'], r['invoice_date'], r['customer_name'], r['customer_mobile'],
                        r['customer_gstin'], r['subtotal'], r['total_discount'], r['total_taxable'],
                        r['total_cgst'], r['total_sgst'], r['total_igst'], r['round_off'], r['grand_total'],
                        r['amount_paid'], r['balance'], r['status'], r['payment_method'], r['payment_status'],
                        r['salesperson']])
    elif report_type == 'inventory':
        w.writerow(['Product Name', 'SKU', 'HSN', 'Category', 'Brand', 'Purchase Price', 'Selling Price',
                    'Opening Stock', 'Purchased', 'Sold', 'Returned', 'Current Stock', 'Reserved', 'Available',
                    'Stock Value', 'Profit Margin %'])
        for r in _fetch_inventory_rows():
            w.writerow([r['name'], r['sku'], r['hsn'], r['category'], r['brand'], r['purchase_price'],
                        r['selling_price'], r['opening_stock'], r['purchased'], r['sold'], r['returned'],
                        r['current_stock'], r['reserved'], r['available'], r['stock_value'], r['profit_margin']])
    elif report_type == 'customers':
        w.writerow(['Rank', 'Name', 'Mobile', 'GSTIN', 'Invoice Count', 'Total Purchase', 'Average Purchase'])
        for r in _fetch_customer_rows(clauses):
            w.writerow([r['rank'], r['name'], r['mobile'], r['gstin'], r['count'], r['total'], r['average']])
    elif report_type == 'payments':
        w.writerow(_PAYMENT_HEADERS)
        for r in _fetch_payments(clauses):
            w.writerow([r['date'], r['invoice_number'], r['customer_name'], r['amount'], r['method'],
                        r['reference'], r['received_by'], r['remarks']])
    elif report_type == 'outstanding':
        w.writerow(['Invoice #', 'Invoice Date', 'Due Date', 'Customer', 'Mobile', 'Grand Total', 'Amount Paid', 'Balance Due', 'Status', 'Overdue'])
        for r in _fetch_outstanding():
            w.writerow([r['invoice_number'], r['invoice_date'], r['due_date'], r['customer_name'],
                        r['customer_mobile'], r['grand_total'], r['amount_paid'], r['balance_due'],
                        r['status'], 'Yes' if r['overdue'] else 'No'])
    elif report_type == 'low_stock':
        w.writerow(['Product', 'SKU', 'Brand', 'Category', 'Current Stock', 'Min Stock', 'Max Stock', 'Suggested Purchase'])
        for r in _fetch_low_stock_rows():
            w.writerow([r['name'], r['sku'], r['brand'], r['category'], r['stock'], r['min_stock'],
                        r['max_stock'], r['suggested']])
    else:
        abort(404)
    buf = io.BytesIO(out.getvalue().encode('utf-8-sig'))
    return send_file(buf, mimetype='text/csv', as_attachment=True, download_name=fname + '.csv')


def _export_excel(report_type, clauses, user_map, fname):
    import openpyxl
    wb = openpyxl.Workbook()
    if report_type == 'sales':
        ws = wb.active; ws.title = 'Sales Report'
        ws.append(_SALES_HEADERS)
        for r in _fetch_sales(clauses, user_map):
            ws.append([r['invoice_number'], r['invoice_date'], r['customer_name'], r['customer_mobile'],
                       r['customer_gstin'], r['subtotal'], r['total_discount'], r['total_taxable'],
                       r['total_cgst'], r['total_sgst'], r['total_igst'], r['round_off'], r['grand_total'],
                       r['amount_paid'], r['balance'], r['status'], r['payment_method'], r['payment_status'],
                       r['salesperson']])
    elif report_type == 'inventory':
        ws = wb.active; ws.title = 'Inventory Report'
        ws.append(['Product Name', 'SKU', 'HSN', 'Category', 'Brand', 'Purchase Price', 'Selling Price',
                   'Opening Stock', 'Purchased', 'Sold', 'Returned', 'Current Stock', 'Reserved', 'Available',
                   'Stock Value', 'Profit Margin %'])
        for r in _fetch_inventory_rows():
            ws.append([r['name'], r['sku'], r['hsn'], r['category'], r['brand'], r['purchase_price'],
                       r['selling_price'], r['opening_stock'], r['purchased'], r['sold'], r['returned'],
                       r['current_stock'], r['reserved'], r['available'], r['stock_value'], r['profit_margin']])
    elif report_type == 'customers':
        ws = wb.active; ws.title = 'Customers Report'
        ws.append(['Rank', 'Name', 'Mobile', 'GSTIN', 'Invoice Count', 'Total Purchase', 'Average Purchase'])
        for r in _fetch_customer_rows(clauses):
            ws.append([r['rank'], r['name'], r['mobile'], r['gstin'], r['count'], r['total'], r['average']])
    elif report_type == 'payments':
        ws = wb.active; ws.title = 'Collection Report'
        ws.append(_PAYMENT_HEADERS)
        for r in _fetch_payments(clauses):
            ws.append([r['date'], r['invoice_number'], r['customer_name'], r['amount'], r['method'],
                       r['reference'], r['received_by'], r['remarks']])
    elif report_type == 'outstanding':
        ws = wb.active; ws.title = 'Outstanding Report'
        ws.append(['Invoice #', 'Invoice Date', 'Due Date', 'Customer', 'Mobile', 'Grand Total', 'Amount Paid', 'Balance Due', 'Status', 'Overdue'])
        for r in _fetch_outstanding():
            ws.append([r['invoice_number'], r['invoice_date'], r['due_date'], r['customer_name'],
                       r['customer_mobile'], r['grand_total'], r['amount_paid'], r['balance_due'],
                       r['status'], 'Yes' if r['overdue'] else 'No'])
    elif report_type == 'low_stock':
        ws = wb.active; ws.title = 'Low Stock Report'
        ws.append(['Product', 'SKU', 'Brand', 'Category', 'Current Stock', 'Min Stock', 'Max Stock', 'Suggested Purchase'])
        for r in _fetch_low_stock_rows():
            ws.append([r['name'], r['sku'], r['brand'], r['category'], r['stock'], r['min_stock'],
                       r['max_stock'], r['suggested']])
    else:
        abort(404)
    _style_excel(ws)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname + '.xlsx')


def _style_excel(ws):
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            m = max([len(str(c.value)) for c in col if c.value is not None] or [0])
            ws.column_dimensions[col[0].column_letter].width = min(m + 2, 40)
    except Exception:
        pass


# ------------------------------------------------------------------
# PDF REPORTS
# ------------------------------------------------------------------

def _pdf_dataset(report_type, clauses, user_map):
    """Return (title, subtitle, headers, rows) for PDF generation."""
    if report_type == 'sales':
        headers = _SALES_HEADERS
        rows = []
        for r in _fetch_sales(clauses, user_map):
            rows.append([r[k] for k in ['invoice_number', 'invoice_date', 'customer_name', 'customer_mobile',
                        'customer_gstin', 'subtotal', 'total_discount', 'total_taxable', 'total_cgst',
                        'total_sgst', 'total_igst', 'round_off', 'grand_total', 'amount_paid', 'balance',
                        'status', 'payment_method', 'payment_status', 'salesperson']])
        return 'Sales Report', 'Invoice-level sales', headers, rows
    if report_type == 'inventory':
        headers = ['Product Name', 'SKU', 'HSN', 'Category', 'Brand', 'Purchase', 'Selling', 'Current Stock', 'Stock Value', 'Profit %']
        rows = [[r['name'], r['sku'], r['hsn'], r['category'], r['brand'], r['purchase_price'], r['selling_price'],
                 r['current_stock'], r['stock_value'], r['profit_margin']] for r in _fetch_inventory_rows()]
        return 'Inventory Report', 'Stock & valuation', headers, rows
    if report_type == 'customers':
        headers = ['Rank', 'Name', 'Mobile', 'GSTIN', 'Invoices', 'Total', 'Average']
        rows = [[r['rank'], r['name'], r['mobile'], r['gstin'], r['count'], round(r['total'], 2), round(r['average'], 2)]
                for r in _fetch_customer_rows(clauses)]
        return 'Customer Report', 'Customer purchase analytics', headers, rows
    if report_type == 'payments':
        headers = ['Date', 'Invoice #', 'Customer', 'Amount', 'Method', 'Reference', 'Collected By', 'Remarks']
        rows = [[r['date'], r['invoice_number'], r['customer_name'], round(r['amount'], 2), r['method'],
                 r['reference'], r['received_by'], r['remarks']] for r in _fetch_payments(clauses)]
        return 'Collection Report', 'Payments received by date & method', headers, rows
    if report_type == 'outstanding':
        headers = ['Invoice #', 'Invoice Date', 'Due Date', 'Customer', 'Mobile', 'Grand Total', 'Amount Paid', 'Balance Due', 'Status', 'Overdue']
        rows = [[r['invoice_number'], r['invoice_date'], r['due_date'], r['customer_name'], r['customer_mobile'],
                 round(r['grand_total'], 2), round(r['amount_paid'], 2), round(r['balance_due'], 2),
                 r['status'], 'Yes' if r['overdue'] else 'No'] for r in _fetch_outstanding()]
        return 'Outstanding Report', 'Invoices with unpaid balances', headers, rows
    if report_type == 'low_stock':
        headers = ['Product', 'SKU', 'Brand', 'Category', 'Current', 'Min', 'Max', 'Suggested']
        rows = [[r['name'], r['sku'], r['brand'], r['category'], r['stock'], r['min_stock'],
                 r['max_stock'], r['suggested']] for r in _fetch_low_stock_rows()]
        return 'Low Stock Report', 'Products at or below minimum stock', headers, rows
    abort(404)


def _export_pdf(report_type, clauses, user_map, filters, fname):
    if report_type == 'sales':
        return _export_sales_pdf(clauses, user_map, filters, fname)
    title, subtitle, headers, rows = _pdf_dataset(report_type, clauses, user_map)
    buf = _build_report_pdf(title, subtitle, headers, rows, filters)
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=fname + '.pdf')


def _build_report_pdf(title, subtitle, headers, rows, filters):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    co = current_app.config.get('COMPANY', {})
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    styles = getSampleStyleSheet()
    S = {
        'title': ParagraphStyle('title', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=15, textColor=rl_colors.HexColor('#081C3A')),
        'comp': ParagraphStyle('comp', parent=styles['Normal'], fontSize=7.5, textColor=rl_colors.HexColor('#4B5563')),
        'subtitle': ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', textColor=rl_colors.HexColor('#2563EB')),
        'summary': ParagraphStyle('summary', parent=styles['Normal'], fontSize=7.5, textColor=rl_colors.HexColor('#111827')),
        'th': ParagraphStyle('th', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7, textColor=rl_colors.white),
        'tc': ParagraphStyle('tc', parent=styles['Normal'], fontSize=7),
    }

    story = []
    logo_path = os.path.join(BASE_DIR, 'static', 'img', 'logo', 'img.png')
    if os.path.exists(logo_path):
        story.append(Image(logo_path, width=42, height=42, hAlign='LEFT'))
        story.append(Spacer(1, 4))
    story.append(Paragraph(co.get('name', 'GV POWERS'), S['title']))
    story.append(Paragraph('GSTIN: %s' % co.get('gstin', ''), S['comp']))
    story.append(Paragraph('%s | Phone: %s %s | Email: %s | %s' % (
        co.get('address', ''), co.get('phone', ''), co.get('mobile', ''), co.get('email', ''), co.get('website', '')), S['comp']))
    story.append(Spacer(1, 2))
    story.append(Paragraph('%s &nbsp;—&nbsp; %s' % (title, subtitle), S['subtitle']))
    sd = filters.get('start_date'); ed = filters.get('end_date')
    if sd or ed:
        story.append(Paragraph('Period: %s to %s' % (sd or '—', ed or '—'), S['summary']))
    story.append(Spacer(1, 10))
    data = [[Paragraph(h, S['th']) for h in headers]]
    for r in rows:
        data.append([Paragraph(str(c), S['tc']) for c in r])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#2563EB')),
        ('GRID', (0, 0), (-1, -1), 0.3, rl_colors.HexColor('#E5E7EB')),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#F8FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(table)

    def _footer(canvas, docobj):
        canvas.saveState()
        canvas.setFont('Helvetica', 6.5)
        canvas.setFillColor(rl_colors.HexColor('#6B7280'))
        canvas.drawCentredString(docobj.pagesize[0] / 2, 15,
            'Generated by GV POWERS ERP | Generated: %s | Page %d | %s' % (
                datetime.now().strftime('%d %b %Y %H:%M'), docobj.page, co.get('website', '')))
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return buf


# ------------------------------------------------------------------
# PRIEMIUM SALES REPORT PDF  (Tally / Zoho / Marg / Busy / SAP style)
# ------------------------------------------------------------------

def _inr(v):
    """Format as Indian-Rupee currency: Rs. 2,950.00"""
    try:
        v = round(float(v or 0), 2)
    except (TypeError, ValueError):
        v = 0.0
    neg = v < 0
    v = abs(v)
    cents = int(round(v * 100))
    whole, frac = divmod(cents, 100)
    ws = str(whole)
    if len(ws) <= 3:
        out = ws
    else:
        head = ws[:-3]
        tail = ws[-3:]
        groups = [head[max(0, i - 2):i] for i in range(len(head), 0, -2)]
        groups.reverse()
        out = ",".join(groups) + "," + tail
    return ("-" if neg else "") + "Rs. " + out + "." + "%02d" % frac


def _pdf_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%d-%m-%Y")
    return str(v or "")


def _esc(s):
    return (str(s or "")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _export_sales_pdf(clauses, user_map, filters, fname):
    rows = list(_fetch_sales(clauses, user_map))
    total = {
        "invoices": len(rows), "revenue": 0.0, "taxable": 0.0, "cgst": 0.0, "sgst": 0.0,
        "igst": 0.0, "discount": 0.0, "round_off": 0.0, "grand": 0.0, "paid": 0.0,
        "bal": 0.0, "cancelled": 0.0,
    }
    for r in rows:
        g = r["grand_total"] or 0
        total["revenue"] += g
        total["taxable"] += r["total_taxable"] or 0
        total["cgst"] += r["total_cgst"] or 0
        total["sgst"] += r["total_sgst"] or 0
        total["igst"] += r["total_igst"] or 0
        total["discount"] += r["total_discount"] or 0
        total["round_off"] += r["round_off"] or 0
        total["grand"] += g
        total["paid"] += r["amount_paid"] or 0
        total["bal"] += r["balance"] or 0
        if (r["status"] or "").lower() == "cancelled":
            total["cancelled"] += g
    total["net_gst"] = total["cgst"] + total["sgst"] + total["igst"]
    total["outstanding"] = total["bal"]

    gen_by = current_user.full_name if current_user.is_authenticated else "Administrator"
    buf = _build_sales_pdf(rows, total, filters, gen_by)
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname + ".pdf")


def _build_sales_pdf(rows, total, filters, generated_by):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as C
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    FN_R = "Helvetica"
    FB_R = "Helvetica-Bold"
    _font_home = r"C:/Windows/Fonts"
    _seg = os.path.join(os.environ.get("WINDIR", "C:/Windows") + "/Fonts", "segoeui.ttf")
    _segb = os.path.join(os.environ.get("WINDIR", "C:/Windows") + "/Fonts", "segoeuib.ttf")
    try:
        if os.path.exists(_seg):
            pdfmetrics.registerFont(TTFont("SeqUI", _seg))
            if os.path.exists(_segb):
                pdfmetrics.registerFont(TTFont("SeqUIB", _segb))
            FN_R, FB_R = "SeqUI", "SeqUIB"
    except Exception:
        pass

    styles = getSampleStyleSheet()
    co = current_app.config.get("COMPANY", {})
    NAVY = C.HexColor("#0F2747")
    BLUE = C.HexColor("#2563EB")
    ZEBRA = C.HexColor("#F8FAFC")
    GRID = C.HexColor("#E5E7EB")
    TEXT = C.HexColor("#1F2937")
    MUTE = C.HexColor("#4B5563")
    nm_buf = io.BytesIO()

    doc = SimpleDocTemplate(nm_buf, pagesize=landscape(A4),
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    content_w = doc.width

    thL = ParagraphStyle("thL", parent=styles["Normal"], fontName=FB_R, fontSize=6.5,
                         leading=8, textColor=C.white, alignment=TA_LEFT)
    thR = ParagraphStyle("thR", parent=styles["Normal"], fontName=FB_R, fontSize=6.5,
                         leading=8, textColor=C.white, alignment=TA_RIGHT)
    tdL = ParagraphStyle("tdL", parent=styles["Normal"], fontSize=6.8, leading=8.4, textColor=TEXT, alignment=TA_LEFT)
    tdR = ParagraphStyle("tdR", parent=styles["Normal"], fontName=FN_R, fontSize=6.8, leading=8.4, textColor=TEXT, alignment=TA_RIGHT)
    tdB = ParagraphStyle("tdB", parent=styles["Normal"], fontName=FB_R, fontSize=6.8,
                         leading=8.4, textColor=NAVY, alignment=TA_RIGHT)
    toL = ParagraphStyle("toL", parent=styles["Normal"], fontName=FB_R, fontSize=6.8,
                         leading=8.4, textColor=C.white, alignment=TA_LEFT)
    toR = ParagraphStyle("toR", parent=styles["Normal"], fontName=FB_R, fontSize=7,
                         leading=8.4, textColor=C.white, alignment=TA_RIGHT)
    cLab = ParagraphStyle("cLab", parent=styles["Normal"], fontSize=6, leading=7.5, textColor=MUTE, alignment=TA_CENTER)
    cVal = ParagraphStyle("cVal", parent=styles["Normal"], fontName=FB_R, fontSize=9.5,
                          leading=11, textColor=NAVY, alignment=TA_CENTER)
    ct = ParagraphStyle("ct", parent=styles["Normal"], fontName=FB_R, fontSize=16,
                        leading=19, textColor=NAVY, alignment=TA_CENTER)
    cs = ParagraphStyle("cs", parent=styles["Normal"], fontSize=6.5, leading=8.5, textColor=MUTE, alignment=TA_CENTER)
    st_meta_t = ParagraphStyle("mt", parent=styles["Normal"], fontName=FB_R, fontSize=12,
                               leading=15, textColor=NAVY, alignment=TA_RIGHT)
    st_meta_s = ParagraphStyle("ms", parent=styles["Normal"], fontSize=7.5, leading=10, textColor=BLUE, alignment=TA_RIGHT)
    st_meta = ParagraphStyle("mm", parent=styles["Normal"], fontSize=6.8, leading=9.4, textColor=TEXT, alignment=TA_RIGHT)

    story = []
    logo = os.path.join(BASE_DIR, "static", "img", "logo", "img.png")

    # ---------------- header ----------------
    dl = ".."
    lc = Table([[""]], colWidths=[72])
    if os.path.exists(logo):
        lc = Table([[Image(logo, width=58, height=58, hAlign="CENTER")]], colWidths=[72])
    lc.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 0), (-1, -1), "CENTER")]))

    center = [Paragraph(_esc(co.get("name", "GV POWERS")), ct), Paragraph(_esc("GSTIN: %s" % (co.get("gstin", ""))), cs)]
    for al in str(co.get("address", "")).split("\n"):
        al = al.strip()
        if al:
            center.append(Paragraph(_esc(al), cs))
    phon = " | ".join([pth for pth in [str(co.get("phone", "")), str(co.get("mobile", ""))] if str(pth).strip()])
    if phon:
        center.append(Paragraph(_esc("Phone: %s" % phon), cs))
    if co.get("email"):
        center.append(Paragraph(_esc("Email: %s" % co.get("email", "")), cs))
    if co.get("website"):
        center.append(Paragraph(_esc("Website: %s" % co.get("website", "")), cs))
    cc = Table([[center]], colWidths=[content_w * 0.62])
    cc.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 2)]))

    nowd = datetime.now()
    meta = [Paragraph("Sales Report", st_meta_t), Paragraph("Invoice Level Sales Report", st_meta_s), Spacer(1, 2)]
    def meta_row(label, value):
        meta.append(Paragraph('<font color="#64748B"><b>%s:</b></font> %s' % (_esc(label), _esc(_pdf_date(value))), st_meta))
    meta_row("Report ID", "SR-%s%s" % (date.today().strftime("%Y%m%d"), str(total["invoices"]).zfill(3)))
    meta_row("Generated Date", nowd.strftime("%d-%m-%Y"))
    meta_row("Generated Time", nowd.strftime("%I:%M %p"))
    meta_row("Generated By", generated_by)
    meta_row("Selected Range", "%s to %s" % (_pdf_date(filters.get("start_date")), _pdf_date(filters.get("end_date"))))
    mc = Table([[meta]], colWidths=[content_w * 0.28])
    mc.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 2)]))

    header = Table([[lc, cc, mc]], colWidths=[content_w * 0.10, content_w * 0.62, content_w * 0.28])
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LINEBELOW", (0, 0), (-1, 0), 5, BLUE),
        ("LINEBELOW", (0, 0), (-1, -1), 0.8, GRID),
    ]))
    story.append(header)
    story.append(Spacer(1, 8))

    # ---------------- summary cards ----------------
    deck1 = [("Total Invoices", _inr(total["invoices"])), ("Total Revenue", _inr(total["revenue"])),
             ("Taxable Value", _inr(total["taxable"])), ("Net GST", _inr(total["net_gst"])),
             ("Grand Total", _inr(total["grand"])), ("Outstanding", _inr(total["outstanding"])),
             ("Paid Amount", _inr(total["paid"]))]
    deck2 = [("CGST", _inr(total["cgst"])), ("SGST", _inr(total["sgst"])), ("IGST", _inr(total["igst"])),
             ("Discount", _inr(total["discount"])), ("Round Off", _inr(total["round_off"])),
             ("Cancelled", _inr(total["cancelled"]))]

    def cards(deck):
        n = len(deck)
        cw = content_w / n
        out = []
        for label, value in deck:
            cell = Table([[Paragraph(_esc(label), cLab)], [Paragraph(value, cVal)]], colWidths=[cw])
            cell.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), C.HexColor("#F1F5F9")),
                ("LINEABOVE", (0, 0), (-1, 0), 2.4, C.HexColor("#CBD5E1")),
                ("BOX", (0, 0), (-1, -1), 0.5, GRID),
                ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ]))
            out.append(cell)
        t = Table([out], colWidths=[cw] * n)
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        return t

    story.append(cards(deck1))
    story.append(Spacer(1, 3))
    story.append(cards(deck2))
    story.append(Spacer(1, 8))
    story.append(Paragraph("Invoice Level Report", ParagraphStyle("sec", parent=styles["Normal"],
                                       fontName=FB_R, fontSize=8.5, textColor=NAVY)))
    story.append(Spacer(1, 4))

    # ---------------- table ----------------
    heads = ["Invoice No", "Date", "Customer", "Mobile", "GSTIN", "Taxable", "CGST", "SGST", "IGST",
             "Discount", "Round Off", "Grand Total", "Paid", "Balance", "Payment", "Status", "Salesperson"]
    num_idx = set([5, 6, 7, 8, 9, 10, 11, 12, 13])
    base_w = [50, 46, 96, 46, 48, 34, 34, 34, 34, 34, 34, 40, 34, 34, 30, 34, 79]
    sc = content_w / float(sum(base_w))
    widths = [w * sc for w in base_w]

    data = [[Paragraph(_esc(h), thR if i in num_idx else thL) for i, h in enumerate(heads)]]

    for r in rows:
        stx = (r["status"] or "").lower()
        if stx in ("completed", "paid"):
            sc = C.HexColor("#16A34A")
        elif stx in ("pending", "partial"):
            sc = C.HexColor("#B45309")
        elif stx == "cancelled":
            sc = C.HexColor("#DC2626")
        else:
            sc = C.HexColor("#4B5563")
        sts = ParagraphStyle("st_%d" % len(data), parent=styles["Normal"], fontName=FB_R,
                             fontSize=6.8, textColor=sc, alignment=TA_LEFT)
        data.append([
            Paragraph(_esc(r["invoice_number"]), tdL),
            Paragraph(_esc(_pdf_date(r["invoice_date"])), tdL),
            Paragraph(_esc(r["customer_name"] or "Walk-in"), tdL),
            Paragraph(_esc(r["customer_mobile"] or "-"), tdL),
            Paragraph(_esc(r["customer_gstin"] or "-"), tdL),
            Paragraph(_inr(r["total_taxable"]), tdR), Paragraph(_inr(r["total_cgst"]), tdR),
            Paragraph(_inr(r["total_sgst"]), tdR), Paragraph(_inr(r["total_igst"]), tdR),
            Paragraph(_inr(r["total_discount"]), tdR), Paragraph(_inr(r["round_off"]), tdR),
            Paragraph(_inr(r["grand_total"]), tdB), Paragraph(_inr(r["amount_paid"]), tdR),
            Paragraph(_inr(r["balance"]), tdR),
            Paragraph(_esc((r["payment_method"] or "-").title()), tdL),
            Paragraph(_esc((r["status"] or "").title()), sts),
            Paragraph(_esc(r["salesperson"] or ""), tdL),
        ])
    data.append([
        Paragraph("Total", toL), Paragraph("", toL), Paragraph("", toL), Paragraph("", toL), Paragraph("", toL),
        Paragraph(_inr(total["taxable"]), toR), Paragraph(_inr(total["cgst"]), toR),
        Paragraph(_inr(total["sgst"]), toR), Paragraph(_inr(total["igst"]), toR),
        Paragraph(_inr(total["discount"]), toR), Paragraph(_inr(total["round_off"]), toR),
        Paragraph(_inr(total["grand"]), toR), Paragraph(_inr(total["paid"]), toR), Paragraph(_inr(total["bal"]), toR),
        Paragraph("", toL), Paragraph("", toL), Paragraph("", toL),
    ])

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [C.white, ZEBRA]),
        ("BACKGROUND", (0, -1), (-1, -1), NAVY),
        ("GRID", (0, 0), (-1, -1), 0.5, GRID),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("FONTSIZE", (0, 1), (-1, -2), 6.8),
        ("FONTSIZE", (0, -1), (-1, -1), 6.8),
    ]))
    story.append(table)

    # ---------------- footer ----------------
    class _NumCanvas(canvas.Canvas):
        def __init__(self, *a, **k):
            canvas.Canvas.__init__(self, *a, **k)
            self._saved = []
        def showPage(self):
            self._saved.append(dict(self.__dict__))
            self._startPage()
        def save(self):
            n = len(self._saved)
            for pst in self._saved:
                self.__dict__.update(pst)
                self._foot(n)
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)
        def _foot(self, n):
            w = self._pagesize[0]
            self.saveState()
            self.setFont("Helvetica", 6.2)
            self.setFillColor(C.HexColor("#6B7280"))
            self.drawString(15 * mm, 8, "Generated by GV POWERS ERP")
            self.drawCentredString(w / 2, 8, "This report is computer generated.")
            self.drawRightString(w - 15 * mm, 8, "Page %d of %d" % (self._pageNumber, n))
            self.drawRightString(w - 15 * mm, -1, "Generated %s | %s" % (nowd.strftime("%d-%m-%Y %H:%M"),
                                                                         _esc(co.get("website", ""))))
            self.restoreState()

    doc.build(story, canvasmaker=_NumCanvas)
    nm_buf.seek(0)
    return nm_buf


############################################################
# API ROUTES
############################################################
############################################################
# API ROUTES
############################################################

@app.route('/api/search')
@login_required
def global_search():
    q = request.args.get('q', '').strip()
    if len(q) < 2: return jsonify([])
    results = []
    if current_user.is_admin:
        for c in Customer.query.filter(or_(Customer.name.ilike(f'%{q}%'), Customer.mobile.ilike(f'%{q}%'))).limit(5).all():
            results.append({'type': 'customer', 'id': c.id, 'title': c.name, 'subtitle': c.mobile or '', 'url': url_for('customer_profile', cid=c.id)})
        for p in Product.query.filter(or_(Product.name.ilike(f'%{q}%'), Product.sku.ilike(f'%{q}%'))).limit(5).all():
            results.append({'type': 'product', 'id': p.id, 'title': p.name, 'subtitle': f'Stock: {p.stock_quantity}', 'url': url_for('product_profile', pid=p.id)})
    for i in Invoice.query.filter(or_(Invoice.invoice_number.ilike(f'%{q}%'), Invoice.customer_name.ilike(f'%{q}%'))).limit(5).all():
        results.append({'type': 'invoice', 'id': i.id, 'title': i.invoice_number, 'subtitle': f'{i.customer_name} - Rs. {i.grand_total}', 'url': url_for('view_invoice', iid=i.id)})
    return jsonify(results)


@app.route('/api/customers/search')
@login_required
def search_customers():
    q = request.args.get('q', '').strip()
    if not q or len(q) < 1: return jsonify([])
    like = f'%{q}%'
    rows = Customer.query.filter(
        or_(
            Customer.name.ilike(like),
            Customer.mobile.ilike(like),
            Customer.email.ilike(like),
            Customer.gstin.ilike(like),
        )
    ).order_by(Customer.name).limit(10).all()
    out = []
    for c in rows:
        outstanding = Decimal('0')
        last_inv = None
        for inv in c.invoices:
            o = (inv.grand_total or 0) - (inv.amount_paid or 0)
            if o > 0:
                outstanding += Decimal(str(o))
            if last_inv is None or inv.invoice_date >= last_inv.invoice_date:
                last_inv = inv
        out.append({
            'id': c.id, 'name': c.name, 'mobile': c.mobile, 'email': c.email,
            'gstin': c.gstin, 'state': c.state, 'state_code': c.state_code,
            'address': c.address, 'invoice_count': c.invoice_count or 0,
            'total_purchases': float(c.total_purchases or 0),
            'outstanding': float(outstanding),
            'last_invoice_number': last_inv.invoice_number if last_inv else None,
            'last_invoice_date': last_inv.invoice_date.isoformat() if (last_inv and last_inv.invoice_date) else None,
        })
    return jsonify(out)


@app.route('/api/products/search')
@login_required
def api_product_search():
    q = request.args.get('q', '').strip()
    allow_oos = request.args.get('allow_out_of_stock', '').lower() in ('1', 'true', 'yes', 'on')
    if len(q) < 1: return jsonify([])
    query = Product.query.outerjoin(Product.category).filter(
        or_(
            Product.name.ilike(f'%{q}%'),
            Product.sku.ilike(f'%{q}%'),
            Product.barcode.ilike(f'%{q}%'),
            Product.hsn.ilike(f'%{q}%'),
            Product.brand.ilike(f'%{q}%'),
            Product.description.ilike(f'%{q}%'),
            Category.name.ilike(f'%{q}%'),
        ),
        Product.is_active == True,
    )
    if not allow_oos:
        query = query.filter(Product.stock_quantity > 0)
    products = query.limit(20).all()
    payload = []
    for p in products:
        stock = p.current_stock or 0
        item = {
            "id": p.id, "name": p.name, "sku": p.sku, "barcode": p.barcode,
            "hsn": p.hsn, "brand": p.brand, "unit": p.unit,
            "selling_price": float(p.selling_price),
            "price": float(p.selling_price), "gst_rate": float(p.gst_rate),
            "quantity": stock, "current_stock": stock, "stock": stock,
            "image": p.image, "category": p.category.name if p.category else "",
            "purchase_price": float(p.purchase_price) if current_user.is_admin else 0,
        }
        payload.append(item)
    return jsonify(payload)


@app.route('/api/products/barcode/<barcode>')
@login_required
def api_product_barcode(barcode):
    """Exact barcode lookup — used by scanner for fast single-product resolution."""
    code = (barcode or '').strip()
    if not code:
        return jsonify({'found': False}), 400
    p = Product.query.filter(Product.barcode == code, Product.is_active == True).first()
    if not p:
        return jsonify({'found': False, 'message': f'No product found for barcode: {code}'}), 404
    stock = p.current_stock or 0
    return jsonify({
        'found': True,
        'product': {
            'id': p.id, 'name': p.name, 'sku': p.sku, 'barcode': p.barcode,
            'hsn': p.hsn, 'brand': p.brand, 'unit': p.unit,
            'selling_price': float(p.selling_price), 'price': float(p.selling_price),
            'gst_rate': float(p.gst_rate), 'current_stock': stock, 'stock': stock,
            'min_stock': p.min_stock or 0, 'max_stock': p.max_stock or 500,
            'image': p.image, 'purchase_price': float(p.purchase_price) if current_user.is_admin else 0,
        }
    })


@app.route('/api/products/generate-barcode')
@login_required
def api_generate_barcode():
    """Generate a unique EAN-13 barcode for a new product."""
    import random
    for _ in range(50):
        body = ''.join([str(random.randint(0, 9)) for _ in range(12)])
        code13 = barcode.ean.EAN13(body)
        full = code13.get_fullcode()
        if not Product.query.filter_by(barcode=full).first():
            return jsonify({'barcode': full})
    return jsonify({'error': 'Could not generate a unique barcode. Please try again.'}), 500


@app.route('/products/<int:pid>/barcode-image')
@login_required
def product_barcode_image(pid):
    """Generate a barcode SVG image for a product."""
    p = db.session.get(Product, pid)
    if not p or not p.barcode:
        abort(404)
    code = p.barcode.strip()
    try:
        if len(code) == 13:
            bc = barcode.ean.EAN13(code)
        elif len(code) == 12:
            bc = barcode.ean.EAN13(code)
        elif len(code) == 8:
            bc = barcode.ean.EAN8(code)
        else:
            bc = barcode.code128.Code128(code)
    except Exception:
        bc = barcode.code128.Code128(code)
    buf = io.BytesIO()
    bc.write(buf, writer=SVGWriter(), options={'module_width': 0.3, 'module_height': 15, 'font_size': 10, 'text_distance': 5, 'quiet_zone': 6.5})
    buf.seek(0)
    return send_file(buf, mimetype='image/svg+xml', download_name=f'barcode_{p.id}.svg')


@app.route('/products/<int:pid>/barcode-label')
@login_required
def product_barcode_label(pid):
    """Render a printable barcode label page for a product."""
    p = db.session.get(Product, pid)
    if not p:
        abort(404)
    barcode_url = url_for('product_barcode_image', pid=p.id) if p.barcode else None
    return render_template('products/barcode_label.html', product=p, barcode_url=barcode_url)


@app.route('/api/notifications')
@login_required
def get_notifications():
    notifs = Notification.query.filter_by(user_id=current_user.id, is_read=False).order_by(desc(Notification.created_at)).limit(10).all()
    return jsonify([{'id': n.id, 'title': n.title, 'message': n.message, 'type': n.notification_type, 'created': n.created_at.isoformat()} for n in notifs])


@app.route('/api/notifications/<int:nid>/read', methods=['POST'])
@login_required
def mark_notification_read(nid):
    n = db.session.get(Notification, nid)
    if n and n.user_id == current_user.id: n.is_read = True; db.session.commit()
    return jsonify({'success': True})


@app.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    today = date.today()
    today_collections = db.session.query(func.coalesce(func.sum(Payment.amount), 0)).join(
        Invoice, Invoice.id == Payment.invoice_id
    ).filter(func.date(Payment.payment_date) == today, Invoice.status != 'cancelled').scalar() or 0
    outstanding = db.session.query(func.coalesce(func.sum(Invoice.balance_due), 0)).filter(
        Invoice.status != 'cancelled', Invoice.balance_due > 0
    ).scalar() or 0
    pending = Invoice.query.filter(Invoice.status != 'cancelled', Invoice.amount_paid < Invoice.grand_total).count()
    paid = Invoice.query.filter(Invoice.status != 'cancelled', Invoice.amount_paid >= Invoice.grand_total).count()
    return jsonify({
        'total_invoices': Invoice.query.count(), 'total_customers': Customer.query.count(),
        'total_products': Product.query.count(), 'total_revenue': float(db.session.query(func.sum(Invoice.grand_total)).scalar() or 0),
        'pending_invoices': pending,
        'paid_invoices': paid,
        'outstanding_amount': float(outstanding),
        'today_collections': float(today_collections),
        'low_stock': Product.query.filter(Product.stock_quantity <= Product.min_stock).count(),
    })


############################################################
# ERROR HANDLERS
############################################################


@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403

@app.errorhandler(500)
def server_error(e):
    return render_template('errors/500.html'), 500


############################################################
# START APPLICATION
############################################################

app = create_app()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)

